# app/operation/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST # For delete view
from django.core.serializers.json import DjangoJSONEncoder

from django.contrib import messages
from django.utils import timezone
from django.urls import reverse

from django.db import transaction, IntegrityError, models
from django.db.models import Q, Count, Avg, Prefetch, F, Value, Case, When, Sum, ExpressionWrapper, DurationField, DecimalField
from django.db.models.functions import Coalesce
from django.utils.dateparse import parse_date
from django.http import JsonResponse, Http404, HttpResponse, HttpResponseForbidden
from django.template.loader import render_to_string
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

import logging
import traceback
import datetime
import calendar

from decimal import Decimal, InvalidOperation
from openpyxl import Workbook
from collections import Counter
from dateutil.relativedelta import relativedelta
from fuzzywuzzy import fuzz, process

import openpyxl
import xlrd
import json
import pandas as pd

from .forms import (
    ExcelImportForm, ParcelItemFormSet, InitialParcelItemFormSet,
    RemoveOrderItemForm, RemoveOrderItemFormSet,
    ParcelCustomsDetailForm, ParcelItemCustomsDetailFormSet, CustomsDeclarationForm,
    PackagingTypeForm, PackagingMaterialForm, PackagingTypeMaterialComponentFormSet, ReceivePackagingStockForm, AirwayBillForm, CourierInvoiceForm,
    DisputeForm, DisputeUpdateForm, ParcelEditForm, CourierInvoiceFilterForm
)
from .models import (Order,
                     OrderItem,
                     Parcel,
                     ParcelItem,
                     CustomsDeclaration,
                     CourierCompany,
                     PackagingType,
                     PackagingTypeMaterialComponent,
                     ParcelTrackingLog,
                     CourierInvoice,
                     CourierInvoiceItem,
                     ProductMapping)
from inventory.models import (Product,
                              InventoryBatchItem,
                              StockTransaction,
                              PackagingMaterial,
                              WarehousePackagingMaterial,
                              PackagingStockTransaction
                              )
from warehouse.models import Warehouse, WarehouseProduct
from inventory.services import get_suggested_batch_for_order_item
from customers.utils import get_or_create_customer_from_import
from customers.models import Customer
from .services import update_parcel_tracking_from_api, parse_invoice_file


logger = logging.getLogger(__name__)
DEFAULT_CUSTOMER_ORDERS_TAB = "customer_orders"
DEFAULT_PARCELS_TAB = "parcels_details"


@login_required
def order_list(request):
    logger.debug(f"[OrderListView] Request GET params: {request.GET}")
    user = request.user

    active_tab = request.GET.get('tab', DEFAULT_CUSTOMER_ORDERS_TAB)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    # Use fetch_dynamic_content_only for customer orders, and fetch_parcel_list_only for parcels
    fetch_dynamic_content_only_co = request.GET.get('fetch_dynamic_content_only') == 'true'

    all_warehouses_qs = Warehouse.objects.all().order_by('name')
    status_choices = Order.STATUS_CHOICES
    import_form = ExcelImportForm()

    context = {
        'status_choices': status_choices,
        'import_form': import_form,
        'active_tab': active_tab,
        'DEFAULT_CUSTOMER_ORDERS_TAB': DEFAULT_CUSTOMER_ORDERS_TAB,
        'DEFAULT_PARCELS_TAB': DEFAULT_PARCELS_TAB,
        'request': request,
    }

    if active_tab == DEFAULT_CUSTOMER_ORDERS_TAB:
        selected_warehouse_id = request.GET.get('warehouse')
        selected_status = request.GET.get('status')
        query = request.GET.get('q', '').strip()
        page_number = request.GET.get('page', 1)
        logger.debug(f"[CustomerOrdersTab] Filters: warehouse='{selected_warehouse_id}', status='{selected_status}', q='{query}', page='{page_number}'")

        # 1. Create a base queryset that only considers user permissions.
        base_orders_qs = Order.objects.all()
        if not user.is_superuser:
            if user.warehouse:
                base_orders_qs = base_orders_qs.filter(warehouse=user.warehouse)
            else:
                base_orders_qs = base_orders_qs.none()

        # 2. Calculate the static counts from this base queryset. These will not change with filters.
        new_orders_count = base_orders_qs.filter(status='NEW_ORDER').count()
        partial_orders_count = base_orders_qs.filter(status='PARTIALLY_SHIPPED').count()

        # 3. Create a separate queryset for filtering and displaying in the table.
        # This starts from the same base but will have more filters applied.
        filtered_orders_qs = base_orders_qs.select_related(
            'warehouse', 'imported_by', 'customer'
        ).prefetch_related(
            Prefetch('items', queryset=OrderItem.objects.select_related('product', 'suggested_batch_item', 'warehouse_product').order_by('product__name')),
            Prefetch('parcels', queryset=Parcel.objects.select_related('courier_company').order_by('-created_at'))
        )


        # 4. Apply user-selected filters (warehouse, query, status) to the 'filtered_orders_qs'.
        warehouses_for_co_filters = all_warehouses_qs
        if not user.is_superuser:
             if user.warehouse:
                warehouses_for_co_filters = Warehouse.objects.filter(pk=user.warehouse.pk)
                selected_warehouse_id = str(user.warehouse.pk)
             else:
                warehouses_for_co_filters = Warehouse.objects.none()
        context['warehouses'] = warehouses_for_co_filters


        if selected_warehouse_id and user.is_superuser:
            filtered_orders_qs = filtered_orders_qs.filter(warehouse_id=selected_warehouse_id)

        if query:
            filtered_orders_qs = filtered_orders_qs.filter(
                Q(erp_order_id__icontains=query) | Q(order_display_code__icontains=query) |
                Q(customer__customer_name__icontains=query) | Q(customer__company_name__icontains=query) |
                Q(items__product__sku__icontains=query) | Q(items__product__name__icontains=query) |
                Q(parcels__parcel_code_system__icontains=query) | Q(parcels__tracking_number__icontains=query)
            ).distinct()

        if selected_status:
            filtered_orders_qs = filtered_orders_qs.filter(status=selected_status)

        filtered_orders_qs = filtered_orders_qs.order_by('-order_date', '-imported_at')
        paginator = Paginator(filtered_orders_qs, 30)

        try:
            orders_page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            orders_page_obj = paginator.page(1)
        except EmptyPage:
            orders_page_obj = paginator.page(paginator.num_pages if paginator.num_pages > 0 else 1)

        for order_instance in orders_page_obj.object_list:
            for item_instance in order_instance.items.all():
                item_instance.quantity_notionally_removed = order_instance.get_total_removed_quantity_for_item(item_instance.id)

        all_courier_companies_for_js = list(CourierCompany.objects.all().values('id', 'name', 'code'))
        context['all_courier_companies_json'] = json.dumps(all_courier_companies_for_js, cls=DjangoJSONEncoder)

        # --- START: New logic to get total parcel counts for displayed customers ---
        # 1. Get the unique customer IDs from the orders on the current page.
        customer_ids_on_page = {order.customer.id for order in orders_page_obj.object_list if order.customer}

        # 2. Run one efficient query to get the total parcel count for each of those customers.
        parcel_counts = Customer.objects.filter(
            id__in=customer_ids_on_page
        ).annotate(
            total_parcel_count=Count('orders__parcels')
        )

        # 3. Create a dictionary for easy lookup in the template.
        customer_parcel_counts = {
            customer.id: customer.total_parcel_count for customer in parcel_counts
        }
        # --- END: New logic ---

        context.update({
            'orders_page_obj': orders_page_obj,
            'customer_parcel_counts': customer_parcel_counts,
            'total_orders_count': paginator.count,
            'new_orders_count': new_orders_count,
            'partial_orders_count': partial_orders_count,
            'selected_warehouse': selected_warehouse_id,
            'selected_status': selected_status,
            'query': query,
            'page_title': "Customer Orders",
        })

        if is_ajax:
            # Use the specific flag for customer orders
            if fetch_dynamic_content_only_co:
                template_to_render = 'operation/partials/_customer_orders_table_with_pagination.html'
            else:
                template_to_render = 'operation/partials/customer_orders_table.html'
            response = render(request, template_to_render, context)
            response['X-Total-Orders-Count'] = paginator.count
            response['X-New-Orders-Count'] = new_orders_count
            response['X-Partial-Orders-Count'] = partial_orders_count

            if orders_page_obj.has_next():
                response['HX-Trigger-After-Swap'] = 'loadMoreCustomerOrdersAvailable'
            else:
                response['HX-Trigger-After-Swap'] = 'loadMoreCustomerOrdersUnavailable'
            return response



    elif active_tab == DEFAULT_PARCELS_TAB:
        fetch_parcel_list_only = request.GET.get('fetch_parcel_list_only') == 'true'

        base_parcels_qs = Parcel.objects.all()
        if not user.is_superuser:
            if user.warehouse:
                base_parcels_qs = base_parcels_qs.filter(order__warehouse=user.warehouse)
            else:
                base_parcels_qs = base_parcels_qs.none()

        to_pack_count = base_parcels_qs.filter(status='PREPARING_TO_PACK').count()
        ready_to_ship_count = base_parcels_qs.filter(status='READY_TO_SHIP').count()
        in_transit_count = base_parcels_qs.filter(status='IN_TRANSIT').count()
        delivered_count = base_parcels_qs.filter(status='DELIVERED').count()

        filtered_parcels_qs = base_parcels_qs.select_related(
            'order__warehouse', 'order__imported_by', 'created_by', 'courier_company', 'billing_item'
        ).prefetch_related(
            Prefetch('items_in_parcel', queryset=ParcelItem.objects.select_related(
                'order_item__product', 'shipped_from_batch'
            ).order_by('order_item__product__name'))
        )

        courier_companies_for_filter = CourierCompany.objects.all().order_by('name')
        context['courier_companies'] = courier_companies_for_filter

        warehouses_for_parcel_filters_ui = all_warehouses_qs
        actual_selected_warehouse_id_for_query_and_ui = request.GET.get('parcel_warehouse')
        selected_parcel_courier_name = request.GET.get('parcel_courier')
        start_date_str = request.GET.get('parcel_start_date')
        end_date_str = request.GET.get('parcel_end_date')

        # --- START: CORRECTION ---
        # This section was causing the error. It now correctly modifies 'filtered_parcels_qs'.
        if not user.is_superuser:
            if user.warehouse:
                warehouses_for_parcel_filters_ui = Warehouse.objects.filter(pk=user.warehouse.pk)
                # The queryset is already filtered by the base_parcels_qs logic, so we just set the UI variable.
                actual_selected_warehouse_id_for_query_and_ui = str(user.warehouse.pk)
            else:
                warehouses_for_parcel_filters_ui = Warehouse.objects.none()
                # The queryset is already .none() from the base_parcels_qs logic.
                actual_selected_warehouse_id_for_query_and_ui = None
        # --- END: CORRECTION ---

        context['warehouses'] = warehouses_for_parcel_filters_ui
        selected_parcel_status = request.GET.get('parcel_status', None)

        not_billed_only = request.GET.get('not_billed_only') == 'true'

        # Make status and "not billed" filters mutually exclusive
        if not_billed_only:
            selected_parcel_status = '' # De-select status filter
            # Filter for parcels where the link to a billing item does not exist
            filtered_parcels_qs = filtered_parcels_qs.filter(billing_item__isnull=True)
        elif selected_parcel_status:
            not_billed_only = False # De-select not_billed filter
            filtered_parcels_qs = filtered_parcels_qs.filter(status=selected_parcel_status)
        # --- END OF THE FIX ---

        if user.is_superuser and actual_selected_warehouse_id_for_query_and_ui:
            filtered_parcels_qs = filtered_parcels_qs.filter(order__warehouse_id=actual_selected_warehouse_id_for_query_and_ui)

        if selected_parcel_courier_name:
            filtered_parcels_qs = filtered_parcels_qs.filter(courier_company__code=selected_parcel_courier_name)

        if start_date_str:
            try:
                start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
                # Use '__date' to correctly filter a DateTimeField with a date object
                filtered_parcels_qs = filtered_parcels_qs.filter(created_at__date__gte=start_date)
            except (ValueError, TypeError):
                # Silently ignore invalid date formats instead of crashing
                pass

        # Filter by end date only if a valid date string is provided
        if end_date_str:
            try:
                end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
                # Use '__date' to correctly filter a DateTimeField with a date object
                filtered_parcels_qs = filtered_parcels_qs.filter(created_at__date__lte=end_date)
            except (ValueError, TypeError):
                # Silently ignore invalid date formats
                pass

        parcel_query_param = request.GET.get('parcel_q', '').strip()
        page_number = request.GET.get('page', 1)
        logger.debug(f"[ParcelTab] Filters: parcel_warehouse='{actual_selected_warehouse_id_for_query_and_ui}', parcel_courier='{selected_parcel_courier_name}', parcel_q='{parcel_query_param}', page='{page_number}'")

        if parcel_query_param:
            filtered_parcels_qs = filtered_parcels_qs.filter(
                Q(parcel_code_system__icontains=parcel_query_param) | Q(tracking_number__icontains=parcel_query_param) |
                Q(order__erp_order_id__icontains=parcel_query_param) | Q(order__customer__customer_name__icontains=parcel_query_param) |
                Q(items_in_parcel__order_item__product__name__icontains=parcel_query_param) | Q(items_in_parcel__order_item__product__sku__icontains=parcel_query_param) |
                Q(items_in_parcel__shipped_from_batch__batch_number__icontains=parcel_query_param) | Q(items_in_parcel__shipped_from_batch__location_label__icontains=parcel_query_param)
            ).distinct()

        # --- START: CORRECTION ---
        # Use the correct variable 'filtered_parcels_qs' for ordering and pagination.
        filtered_parcels_qs = filtered_parcels_qs.order_by('-created_at')
        parcel_paginator = Paginator(filtered_parcels_qs, 30)
        # --- END: CORRECTION ---

        try:
            parcels_page = parcel_paginator.page(page_number)
        except PageNotAnInteger:
            parcels_page = parcel_paginator.page(1)
        except EmptyPage:
            parcels_page = parcel_paginator.page(parcel_paginator.num_pages if parcel_paginator.num_pages > 0 else 1)

        context.update({
            'parcels': parcels_page,
            'total_parcels_count': parcel_paginator.count,
            'to_pack_count': to_pack_count,
            'ready_to_ship_count': ready_to_ship_count,
            'in_transit_count': in_transit_count,
            'delivered_count': delivered_count,
            'selected_parcel_warehouse': actual_selected_warehouse_id_for_query_and_ui,
            'selected_parcel_courier': selected_parcel_courier_name,
            'selected_parcel_status': selected_parcel_status,
            'not_billed_only': not_billed_only,
            'parcel_query': parcel_query_param,
            'page_title': "Parcel Details",
            'start_date': start_date_str,
            'end_date': end_date_str,
        })

        if is_ajax:
            if fetch_parcel_list_only:
                template_to_render = 'operation/partials/_parcels_list_content.html'
            else:
                template_to_render = 'operation/partials/parcels_table.html'

            response = render(request, template_to_render, context)
            response['X-Total-Parcels-Count'] = parcel_paginator.count
            response['X-To-Pack-Count'] = to_pack_count
            response['X-Ready-To-Ship-Count'] = ready_to_ship_count
            response['X-In-Transit-Count'] = in_transit_count
            response['X-Delivered-Count'] = delivered_count
            return response

    return render(request, 'operation/order_management_base.html', context)


@login_required
def load_more_customer_orders(request):
    # --- START: Added extensive logging ---
    logger.debug("--- [load_more_customer_orders] Function Called ---")
    logger.debug(f"Request GET params: {request.GET}")

    user = request.user

    orders_qs = Order.objects.select_related(
        'warehouse', 'imported_by', 'customer'
    ).prefetch_related(
        Prefetch('items', queryset=OrderItem.objects.select_related('product', 'suggested_batch_item', 'warehouse_product').order_by('product__name')),
        Prefetch('parcels', queryset=Parcel.objects.select_related('courier_company').order_by('-created_at'))
    ).all()

    initial_count = orders_qs.count()
    logger.debug(f"Initial queryset count: {initial_count}")

    if not user.is_superuser and user.warehouse:
        orders_qs = orders_qs.filter(warehouse=user.warehouse)
        logger.debug(f"Filtered for non-superuser warehouse '{user.warehouse.name}'. New count: {orders_qs.count()}")
    elif not user.is_superuser:
        orders_qs = orders_qs.none()
        logger.debug("Non-superuser with no warehouse. Queryset is now empty.")

    selected_warehouse_id = request.GET.get('warehouse')
    selected_status = request.GET.get('status')
    query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page', 1)
    logger.debug(f"Extracted Filters -> Warehouse: '{selected_warehouse_id}', Status: '{selected_status}', Query: '{query}', Page: {page_number}")

    if user.is_superuser and selected_warehouse_id:
        orders_qs = orders_qs.filter(warehouse_id=selected_warehouse_id)
        logger.debug(f"Filtered for warehouse ID '{selected_warehouse_id}'. New count: {orders_qs.count()}")

    if selected_status:
        orders_qs = orders_qs.filter(status=selected_status)
        logger.debug(f"Filtered for status '{selected_status}'. New count: {orders_qs.count()}")

    if query:
        # This uses the corrected and more complete filtering logic
        orders_qs = orders_qs.filter(
            Q(erp_order_id__icontains=query) |
            Q(order_display_code__icontains=query) |
            Q(customer__customer_name__icontains=query) |
            Q(customer__company_name__icontains=query) |
            Q(items__product__sku__icontains=query) |
            Q(items__product__name__icontains=query) |
            Q(parcels__parcel_code_system__icontains=query) |
            Q(parcels__tracking_number__icontains=query)
        ).distinct()
        logger.debug(f"Filtered for query '{query}'. New (distinct) count: {orders_qs.count()}")

    orders_qs = orders_qs.order_by('-order_date', '-imported_at')

    paginator = Paginator(orders_qs, 30) # Using 10 for "load more" is fine
    logger.debug(f"Paginator created. Total items for pagination: {paginator.count}. Requesting page: {page_number}")

    try:
        orders_page_obj = paginator.page(page_number)
        logger.debug(f"Successfully fetched page {page_number}. Items on this page: {len(orders_page_obj.object_list)}")
    except PageNotAnInteger:
        logger.warning(f"PageNotAnInteger error for page '{page_number}'. Defaulting to page 1.")
        orders_page_obj = paginator.page(1)
        if not orders_page_obj.object_list:
            logger.debug("Defaulted to page 1, but it is empty. Returning empty response.")
            return HttpResponse("")
    except EmptyPage:
        logger.info(f"EmptyPage error for page '{page_number}'. No more items to load. Returning empty response.")
        return HttpResponse("") # This is expected when all items are loaded

    # --- This part for fetching related customer parcel counts is new and useful for the template ---
    customer_ids_on_page = {order.customer.id for order in orders_page_obj.object_list if order.customer}
    customer_parcel_counts = {
        customer.id: customer.total_parcel_count
        for customer in Customer.objects.filter(id__in=customer_ids_on_page).annotate(total_parcel_count=Count('orders__parcels'))
    }
    logger.debug(f"Fetched total parcel counts for {len(customer_parcel_counts)} customers on this page.")
    # ---

    for order_instance in orders_page_obj.object_list:
        for item_instance in order_instance.items.all():
            item_instance.quantity_notionally_removed = order_instance.get_total_removed_quantity_for_item(item_instance.id)

    context = {
        'orders': orders_page_obj.object_list,
        'customer_parcel_counts': customer_parcel_counts, # Pass this to the template
        'request': request,
    }

    html_rows = render_to_string('operation/partials/_customer_orders_list_items_only.html', context)

    response = HttpResponse(html_rows)
    if orders_page_obj.has_next():
        response['HX-Trigger'] = 'loadMoreCustomerOrdersAvailable'
        logger.debug("Sending response with HX-Trigger: loadMoreCustomerOrdersAvailable")
    else:
        response['HX-Trigger'] = 'loadMoreCustomerOrdersUnavailable'
        logger.debug("Sending response with HX-Trigger: loadMoreCustomerOrdersUnavailable")

    logger.debug("--- [load_more_customer_orders] Response Sent ---")
    # --- END: Added extensive logging ---
    return response



@login_required
def import_orders_from_excel(request):
    """
    Handles the initial Excel file upload, remembers past corrections,
    and shows a confirmation modal for new or uncertain matches.
    """
    if request.method != 'POST':
        return redirect('operation:order_list')

    form = ExcelImportForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "The uploaded form was not valid. Please try again.")
        return redirect('operation:order_list')

    excel_file = request.FILES['excel_file']

    try:
        # --- 1. Load Excel File and Get Iterator ---
        if excel_file.name.lower().endswith('.xlsx'):
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = workbook.active
            rows_iterator = sheet.iter_rows(min_row=1, values_only=True)
        elif excel_file.name.lower().endswith('.xls'):
            file_contents = excel_file.read()
            workbook = xlrd.open_workbook(file_contents=file_contents)
            sheet = workbook.sheet_by_index(0)
            rows_iterator = (sheet.row_values(r) for r in range(sheet.nrows))
        else:
            messages.error(request, "Unsupported file format. Please use .xlsx or .xls.")
            return redirect('operation:order_list')

        headers = [str(h).strip() if h is not None else '' for h in next(rows_iterator)]

        # --- 2. Map Headers ---
        header_map = {
            'Order_ID': 'Order ID', 'Order_Date': 'Order date', 'Warehouse_name': 'Warehouse name',
            'title': 'title', 'comment': 'comment', 'Customer_Name': 'Address name', 'company': 'company',
            'address': 'address', 'country': 'country', 'city': 'city', 'state': 'state',
            'zip': 'zip', 'phone': 'phone', 'Vat_number': 'Vat number',
            'Product_Name': 'Product name', 'Quantity': 'Product quantity', 'isCold': 'isCold'
        }

        index_map = {key: headers.index(value) for key, value in header_map.items() if value in headers}

        missing = [v for k, v in header_map.items() if k in ['Order_ID', 'Product_Name', 'Quantity'] and v not in headers]
        if missing:
            messages.error(request, f"Required headers not found in file: {', '.join(missing)}")
            return redirect('operation:order_list')

        # --- 3. Process Rows & Prepare for Confirmation ---
        items_to_confirm = []
        last_order_details = {}

        for row_idx, row_values in enumerate(rows_iterator, start=2):
            if not any(c is not None and str(c).strip() for c in row_values):
                continue

            row_dict = {key: row_values[idx] if idx < len(row_values) else None for key, idx in index_map.items()}

            # Clean Order ID
            order_id_raw = row_dict.get('Order_ID')
            if order_id_raw:
                try:
                    row_dict['Order_ID'] = str(int(float(order_id_raw)))
                except (ValueError, TypeError):
                    row_dict['Order_ID'] = str(order_id_raw).strip()

            # Clean Company Name
            company_raw = row_dict.get('company')
            row_dict['company'] = str(company_raw).strip() if company_raw and pd.notna(company_raw) else '-'


            if not row_dict.get('Order_ID'):
                row_dict.update({k: v for k, v in last_order_details.items() if k not in ['Product_Name', 'Quantity', 'isCold']})
            else:
                last_order_details = row_dict.copy()

            if not all([row_dict.get('Order_ID'), row_dict.get('Product_Name'), row_dict.get('Quantity')]):
                continue

            items_to_confirm.append(row_dict)

        if not items_to_confirm:
            messages.warning(request, "No valid order items could be parsed from the file.")
            return redirect('operation:order_list')

        # --- 4. Perform Matching ("Remember" Logic) ---
        all_products = list(Product.objects.all())
        product_choices = {p.name: p.id for p in all_products}
        learned_mappings = {mapping.imported_name: mapping.mapped_product_id for mapping in ProductMapping.objects.all()}

        for item in items_to_confirm:
            name_to_match = str(item.get('Product_Name', '')).strip()
            item['suggested_product_id'] = None
            item['similarity_score'] = 0

            # Prioritize learned mappings
            if name_to_match in learned_mappings:
                item['suggested_product_id'] = learned_mappings[name_to_match]
                item['similarity_score'] = 100 # Perfect score for a learned match
            elif name_to_match:
                # Fallback to fuzzy matching
                match = process.extractOne(name_to_match, product_choices.keys(), scorer=fuzz.ratio)
                if match and match[1] >= 60:
                    item['suggested_product_id'] = product_choices[match[0]]
                    item['similarity_score'] = match[1]

        # --- 5. Render Confirmation Modal ---
        request.session['imported_data'] = items_to_confirm
        return render(request, 'operation/import_orders_from_excel.html', {
            'show_confirmation_modal': True,
            'items_to_confirm': items_to_confirm,
            'all_products': all_products,
        })

    except Exception as e:
        logger.error(f"Error during Excel import: {e}", exc_info=True)
        messages.error(request, f"An unexpected error occurred: {e}")
        return redirect('operation:order_list')


@login_required
@transaction.atomic
def create_orders_from_import(request):
    """
    Creates orders from the session data and "learns" any product
    corrections made by the user.
    """
    if request.method != 'POST':
        return redirect('operation:import_orders_from_excel')

    imported_data = request.session.get('imported_data')
    if not imported_data:
        messages.error(request, "Import session has expired. Please re-upload your file.")
        return redirect('operation:import_orders_from_excel')

    # --- START: "Learn" from User Corrections ---
    for i, item_row in enumerate(imported_data):
        product_id_str = request.POST.get(f'product_selection_{i}')
        if not product_id_str:
            messages.error(request, f"A product was not selected for '{item_row.get('Product_Name')}'.")
            return redirect('operation:import_orders_from_excel')

        confirmed_product_id = int(product_id_str)
        item_row['confirmed_product_id'] = confirmed_product_id

        imported_name = str(item_row.get('Product_Name', '')).strip()
        if imported_name:
            ProductMapping.objects.update_or_create(
                imported_name=imported_name,
                defaults={'mapped_product_id': confirmed_product_id}
            )
            logger.info(f"Learned/updated mapping: '{imported_name}' -> Product ID {confirmed_product_id}")
    # --- END: "Learn" from User Corrections ---

    # Group items by Order ID
    orders_to_process = {}
    for item in imported_data:
        order_id = str(item.get('Order_ID'))
        if order_id not in orders_to_process:
            orders_to_process[order_id] = []
        orders_to_process[order_id].append(item)

    # Final database creation loop
    for order_id, items in orders_to_process.items():
        first_item = items[0]

        # --- Validate and Get Warehouse ---
        warehouse_name = first_item.get('Warehouse_name')
        if pd.isna(warehouse_name) or not warehouse_name:
            messages.error(request, f"Import failed: 'Warehouse name' is missing for order {order_id}.")
            raise transaction.TransactionManagementError("Rolling back due to missing warehouse name.")
        try:
            warehouse = Warehouse.objects.get(name__iexact=str(warehouse_name))
        except Warehouse.DoesNotExist:
            messages.error(request, f"Import failed: Warehouse '{warehouse_name}' not found for order {order_id}.")
            raise transaction.TransactionManagementError("Rolling back due to non-existent warehouse.")

        # --- Robust Customer Lookup ---
        customer = None
        phone_number_raw = first_item.get('phone')
        phone_number_clean = None
        if phone_number_raw and pd.notna(phone_number_raw):
            try:
                phone_number_clean = str(int(float(phone_number_raw)))
            except (ValueError, TypeError):
                pass
        if phone_number_clean:
            customer = Customer.objects.filter(phone_number=phone_number_clean).first()
        if not customer:
            customer, _ = Customer.objects.get_or_create(
                customer_name=first_item.get('Customer_Name', ''),
                defaults={
                    'email': first_item.get('email') if pd.notna(first_item.get('email')) else None,
                    'company_name': first_item.get('company', ''), 'phone_number': phone_number_clean,
                    'address_line1': first_item.get('address'), 'city': first_item.get('city'),
                    'state': first_item.get('state'), 'zip_code': first_item.get('zip'),
                    'country': first_item.get('country'), 'vat_number': first_item.get('Vat_number')
                }
            )

        # --- Create or Update Order ---
        order_date = pd.to_datetime(first_item.get('Order_Date'), errors='coerce').date() or timezone.now().date()
        order, created = Order.objects.update_or_create(
            erp_order_id=order_id,
            defaults={
                'customer': customer, 'order_date': order_date, 'warehouse': warehouse,
                'title_notes': first_item.get('title'), 'shipping_notes': first_item.get('comment'),
                'status': 'NEW_ORDER', 'imported_by': request.user
            }
        )
        if not created:
            order.items.all().delete()

        # --- Create Order Items ---
        order_is_cold = False
        for item_data in items:
            product = Product.objects.get(id=item_data['confirmed_product_id'])
            warehouse_product, _ = WarehouseProduct.objects.get_or_create(
                warehouse=warehouse, product=product, defaults={'quantity': 0}
            )
            OrderItem.objects.create(
                order=order, product=product, warehouse_product=warehouse_product,
                quantity_ordered=int(item_data.get('Quantity', 0))
            )
            if str(item_data.get('isCold')).lower() in ['yes', 'true', '1']:
                order_is_cold = True

        if order_is_cold:
            order.is_cold_chain = True
            order.save()

    # Clean up and redirect
    if 'imported_data' in request.session:
        del request.session['imported_data']
    messages.success(request, f"{len(orders_to_process)} orders have been successfully imported or updated.")
    return redirect('operation:order_list')


@login_required
def get_order_items_for_packing(request, order_pk):
    """
    Gets all data needed to populate the 'Pack Order' modal.
    This version uses a more robust method to count parcels for the courier dashboard.
    """
    try:
        order = get_object_or_404(
            Order.objects.prefetch_related(
                Prefetch(
                    'items',
                    queryset=OrderItem.objects.filter(
                        Q(status='PENDING_PROCESSING') | (Q(status='PACKED') & Q(quantity_packed__lt=F('quantity_ordered')))
                    ).select_related('product', 'warehouse_product')
                )
            ).select_related('warehouse', 'customer'), # Ensure customer is also selected
            pk=order_pk
        )

        if not request.user.is_superuser and (not request.user.warehouse or order.warehouse != request.user.warehouse):
            return JsonResponse({'success': False, 'message': 'Permission denied for this order.'}, status=403)

        initial_form_data = []
        for item in order.items.all():
            total_removed_for_this_item = order.get_total_removed_quantity_for_item(item.id)
            quantity_remaining_to_pack_for_this_item = item.quantity_ordered - item.quantity_packed - total_removed_for_this_item

            if quantity_remaining_to_pack_for_this_item > 0:
                best_suggested_batch = get_suggested_batch_for_order_item(item, quantity_remaining_to_pack_for_this_item)

                # ++ MODIFICATION: Get the warehouse_product to access shipping limits ++
                wp = item.warehouse_product

                initial_form_data.append({
                    'order_item_id': item.pk,
                    'product_name': item.product.name if item.product else item.erp_product_name,
                    'sku': item.product.sku if item.product else "N/A",
                    'quantity_to_pack': quantity_remaining_to_pack_for_this_item,
                    'selected_batch_item_id': best_suggested_batch.pk if best_suggested_batch else None,
                    # ++ ADDED: Pass the shipping quantity limits to the form's initial data ++
                    'max_ship_qty_a': wp.max_ship_qty_a if wp else None,
                    'max_ship_qty_b': wp.max_ship_qty_b if wp else None,
                })

        formset_html_content = ""
        message_for_modal = ""
        if not initial_form_data:
            message_for_modal = 'All items for this order are already fully packed or have no remaining quantity.'
            formset_html_content = f'<p class="text-center py-4 text-gray-500">{message_for_modal}</p>'
        else:
            packing_items_formset = InitialParcelItemFormSet(initial=initial_form_data, prefix='packitems')
            formset_html_content = render_to_string(
                'operation/partials/pack_order_formset.html',
                {'formset': packing_items_formset, 'order': order},
                request=request
            )

        env_type = 'COLD' if order.is_cold_chain else 'AMBIENT'
        available_packaging = list(PackagingType.objects.filter(
            environment_type=env_type, is_active=True
        ).values('id', 'name', 'type_code'))

        now = timezone.now()
        start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = now.replace(hour=23, minute=59, second=59, microsecond=999999)

        todays_parcels = Parcel.objects.filter(
            created_at__date=timezone.now().date(),
            order__warehouse=order.warehouse, # This is the crucial filter
            courier_company__isnull=False
        ).select_related('courier_company')


        courier_names = [p.courier_company.name for p in todays_parcels]
        todays_parcel_counts = Counter(courier_names)

        active_couriers = CourierCompany.objects.filter(is_active=True).order_by('name')

        final_daily_courier_counts = {
            courier.name: todays_parcel_counts.get(courier.name, 0)
            for courier in active_couriers
        }

        courier_list_for_modal = [
            {'id': c.id, 'name': c.name, 'code': c.code or ''}
            for c in active_couriers
        ]

        customer_name_display = order.customer.customer_name if order.customer else "N/A"

        return JsonResponse({
            'success': True,
            'order_id': order.pk,
            'erp_order_id': order.erp_order_id,
            'customer_name': customer_name_display,
            'formset_html': formset_html_content,
            'message': message_for_modal,
            'shipping_notes_for_parcel': order.shipping_notes or '',
            'is_cold_chain': order.is_cold_chain,
            'daily_courier_counts_object': final_daily_courier_counts,
            'available_couriers': courier_list_for_modal,
            'available_packaging': available_packaging
        })

    except Http404:
        logger.warning(f"Order PK {order_pk} not found in get_order_items_for_packing.")
        return JsonResponse({'success': False, 'message': 'Order not found.'}, status=404)
    except Exception as e:
        logger.error(f"Unexpected error in get_order_items_for_packing for order_pk {order_pk}: {e}\n{traceback.format_exc()}")
        return JsonResponse({'success': False, 'message': 'An unexpected server error occurred while preparing packing information.'}, status=500)


@login_required
@require_POST # Ensures this view only accepts POST requests
@transaction.atomic # Ensures all database operations are atomic
def process_packing_for_order(request, order_pk):
    """
    Processes the packing form submission to create a new Parcel.
    This version is redesigned to handle a single OrderItem being packed
    from MULTIPLE inventory batches within the same parcel.
    """
    order = get_object_or_404(Order.objects.select_related('warehouse'), pk=order_pk)

    logger.info("=====================================================")
    logger.info(f"Initiating packing process for Order PK: {order_pk}")

    # Log the entire raw POST data from the form
    # This is the MOST IMPORTANT log to check.
    logger.info("--- RAW REQUEST.POST DATA ---")
    for key, value in request.POST.items():
        logger.info(f"  {key}: {value}")
    logger.info("-----------------------------")


    try:
        # 1. Permission Check
        if not request.user.is_superuser and (not request.user.warehouse or order.warehouse != request.user.warehouse):
            return JsonResponse({'success': False, 'message': 'Permission denied for this order.'}, status=403)

        # 2. Get and Validate Parcel-Level Data
        courier_id = request.POST.get('parcel-courier_id')
        packaging_id = request.POST.get('parcel-packaging_type')
        if not courier_id or not packaging_id:
            return JsonResponse({'success': False, 'message': 'A Courier and Packaging selection are required.'}, status=400)

        courier_instance = get_object_or_404(CourierCompany, pk=courier_id)
        packaging_type = get_object_or_404(PackagingType, pk=packaging_id)
        parcel_notes = request.POST.get('parcel-notes', order.shipping_notes or '')

        # 3. Aggregate and Pre-validate Item Data from Formset
        items_to_pack_grouped = {}
        batch_pack_requests = {}
        total_forms = int(request.POST.get('packitems-TOTAL_FORMS', 0))

        logger.info(f"Found 'packitems-TOTAL_FORMS': {total_forms}. Looping through forms...")

        for i in range(total_forms):
            prefix = f'packitems-{i}-'

            # Use .get() to avoid errors if a field is missing for some reason
            quantity_str = request.POST.get(f'{prefix}quantity_to_pack', '0')
            order_item_id = request.POST.get(f'{prefix}order_item_id')
            # IMPORTANT: The name of your batch select might be different!
            # I am using `selected_batch_item_id` based on previous context.
            # Double-check this name in your HTML template.
            batch_id_str = request.POST.get(f'{prefix}selected_batch_item_id')

            logger.info(f"--- Processing Form Index {i} ('{prefix}') ---")
            logger.info(f"  Extracted quantity_to_pack: '{quantity_str}'")
            logger.info(f"  Extracted order_item_id: '{order_item_id}'")
            logger.info(f"  Extracted selected_batch_item_id: '{batch_id_str}'")

            try:
                quantity_to_pack = int(quantity_str)
                if quantity_to_pack <= 0:
                    logger.info("  Quantity is 0 or less, skipping this form.")
                    continue
            except (ValueError, TypeError):
                logger.warning(f"  Could not parse quantity '{quantity_str}' to an integer. Skipping.")
                continue

            if not order_item_id or not batch_id_str:
                logger.warning("  Missing order_item_id or batch_id_str. Skipping.")
                continue

            batch_id = int(batch_id_str)

            # Aggregate total quantity requested per batch
            batch_pack_requests[batch_id] = batch_pack_requests.get(batch_id, 0) + quantity_to_pack

            # Group items for creating ParcelItem objects later
            if order_item_id not in items_to_pack_grouped:
                items_to_pack_grouped[order_item_id] = {
                    'order_item_obj': get_object_or_404(OrderItem, pk=order_item_id, order=order),
                    'total_quantity_to_pack': 0,
                    'packs_from_batches': []
                }

            items_to_pack_grouped[order_item_id]['packs_from_batches'].append({
                'batch_id': batch_id,
                'quantity': quantity_to_pack
            })
            items_to_pack_grouped[order_item_id]['total_quantity_to_pack'] += quantity_to_pack

        # =========================================================================
        # START: AGGREGATION LOGGING
        # =========================================================================
        logger.info("--- AGGREGATION RESULTS ---")
        logger.info(f"Final `batch_pack_requests` dictionary: {batch_pack_requests}")
        logger.info(f"Final `items_to_pack_grouped` dictionary: {items_to_pack_grouped}")
        logger.info("---------------------------")
        # =========================================================================

        if not items_to_pack_grouped:
            logger.error("Validation failed: No items with a quantity to pack were found after processing.")
            return JsonResponse({'success': False, 'message': 'No items were submitted with a quantity to pack.'}, status=400)

        # 4. Comprehensive Stock Validation (Post-Aggregation)
        logger.info("--- PERFORMING STOCK VALIDATION ---")
        for batch_id, requested_qty in batch_pack_requests.items():
            batch_item = get_object_or_404(InventoryBatchItem, pk=batch_id)
            logger.info(f"  Validating Batch ID {batch_id} ('{batch_item.batch_number}'): Requested {requested_qty}, Available {batch_item.quantity}")
            if requested_qty > batch_item.quantity:
                sku = batch_item.warehouse_product.product.sku
                error_msg = f"Not enough stock for {sku} in batch {batch_item.batch_number}. Requested: {requested_qty}, Available: {batch_item.quantity}"
                logger.error(f"  STOCK VALIDATION FAILED: {error_msg}")
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
        logger.info("--- STOCK VALIDATION PASSED ---")

        for data in items_to_pack_grouped.values():
            order_item = data['order_item_obj']
            total_packed = data['total_quantity_to_pack']
            total_removed_for_item = order.get_total_removed_quantity_for_item(order_item.id)
            quantity_remaining_on_order = (order_item.quantity_ordered - total_removed_for_item) - order_item.quantity_packed

            if total_packed > quantity_remaining_on_order:
                sku = order_item.product.sku
                return JsonResponse({'success': False, 'message': f"Cannot pack a total of {total_packed} for {sku}. Only {quantity_remaining_on_order} are remaining on the order."}, status=400)

        # 5. Create the Parcel
        new_parcel = Parcel.objects.create(
            order=order,
            created_by=request.user,
            notes=parcel_notes,
            courier_company=courier_instance,
            packaging_type=packaging_type
        )
        logger.info(f"Parcel {new_parcel.pk} created for order {order_pk}.")

        # 6. Create ParcelItems and Update Stock
        for data in items_to_pack_grouped.values():
            order_item_obj = data['order_item_obj']

            for pack_info in data['packs_from_batches']:
                batch = get_object_or_404(InventoryBatchItem, pk=pack_info['batch_id'])
                qty = pack_info['quantity']

                ParcelItem.objects.create(
                    parcel=new_parcel,
                    order_item=order_item_obj,
                    quantity_shipped_in_this_parcel=qty,
                    shipped_from_batch=batch
                )

                # Deduct stock from both the specific batch AND the parent WarehouseProduct
                batch.quantity = F('quantity') - qty
                batch.save(update_fields=['quantity'])

                if batch.warehouse_product:
                    warehouse_product = batch.warehouse_product
                    warehouse_product.quantity = F('quantity') - qty
                    warehouse_product.save(update_fields=['quantity'])

                StockTransaction.objects.create(
                    warehouse=batch.warehouse_product.warehouse,
                    transaction_type=StockTransaction.TransactionTypes.SALE_PACKED_OUT,
                    warehouse_product=batch.warehouse_product,
                    product=batch.warehouse_product.product,
                    batch_item_involved=batch,
                    quantity=-qty,
                    reference_note=f"LWA Order {order.erp_order_id}, Parcel {new_parcel.parcel_code_system}, Batch {batch.batch_number}",
                    related_order=order,
                    recorded_by=request.user
                )

        # 6b. Deduct Packaging Material Stock
        packaging_components = packaging_type.packagingtypematerialcomponent_set.select_related('packaging_material')
        # First, validate all materials before making any changes
        for component in packaging_components:
            material_to_check = component.packaging_material
            quantity_needed = component.quantity

            try:
                # Query the WAREHOUSE-SPECIFIC stock model
                warehouse_stock = WarehousePackagingMaterial.objects.get(
                    packaging_material=material_to_check,
                    warehouse=order.warehouse
                )

                logger.info(f"  Validating stock for '{material_to_check.name}': Needed {quantity_needed}, Available {warehouse_stock.current_stock}")
                if warehouse_stock.current_stock < quantity_needed:
                    error_msg = f"Insufficient stock for packaging material: '{material_to_check.name}'. Required: {quantity_needed}, Available: {warehouse_stock.current_stock}"
                    logger.error(f"  PACKAGING STOCK VALIDATION FAILED: {error_msg}")
                    return JsonResponse({'success': False, 'message': error_msg}, status=400)

            except WarehousePackagingMaterial.DoesNotExist:
                error_msg = f"Stock for packaging material '{material_to_check.name}' is not configured in warehouse '{order.warehouse.name}'."
                logger.error(f"  PACKAGING STOCK VALIDATION FAILED: {error_msg}")
                return JsonResponse({'success': False, 'message': error_msg}, status=400)

        for component in packaging_components:
            warehouse_stock = WarehousePackagingMaterial.objects.get(
                packaging_material=component.packaging_material,
                warehouse=order.warehouse
            )
            quantity_to_deduct = component.quantity

            # 1. Update the main stock level
            warehouse_stock.current_stock = F('current_stock') - quantity_to_deduct
            warehouse_stock.save(update_fields=['current_stock'])

            # 2. Create the historical USAGE transaction record
            PackagingStockTransaction.objects.create(
                warehouse_packaging_material=warehouse_stock,
                transaction_type=PackagingStockTransaction.TransactionTypes.STOCK_OUT,
                quantity=-quantity_to_deduct, # Log usage as a negative number
                related_parcel=new_parcel,
                notes=f"Used for Parcel {new_parcel.parcel_code_system}",
                recorded_by=request.user
            )
            logger.info(f"  Deducted and logged {quantity_to_deduct} of '{component.packaging_material.name}' for Parcel {new_parcel.pk}.")



        # 7. Update Order Status
        order.update_status_based_on_items_and_parcels()

        messages.success(request, f"Parcel {new_parcel.parcel_code_system} created successfully for order {order.erp_order_id}.")
        return JsonResponse({
            'success': True,
            'message': f'Parcel {new_parcel.parcel_code_system} created successfully.',
            'redirect_url': reverse('operation:order_list') + f"?tab={DEFAULT_CUSTOMER_ORDERS_TAB}"
        })

    except Exception as e:
        logger.error(f"Critical error in process_packing_for_order for order_pk {order_pk}: {e}\n{traceback.format_exc()}")
        return JsonResponse({'success': False, 'message': f'An unexpected server error occurred: {str(e)}'}, status=500)



@login_required
def get_available_batches_for_order_item(request, order_item_pk):
    try:
        order_item = get_object_or_404(OrderItem.objects.select_related('product', 'warehouse_product__warehouse', 'order__warehouse'), pk=order_item_pk)

        logger.info(f"[get_available_batches] Called for OI PK: {order_item_pk}. WP ID: {order_item.warehouse_product_id if order_item.warehouse_product else 'None'}")

        # Permission Check: User must be superuser or assigned to the order's warehouse
        if not request.user.is_superuser and \
           (not request.user.warehouse or order_item.order.warehouse != request.user.warehouse):
            logger.error(f"Permission denied for user {request.user.id} on OI {order_item_pk} in get_available_batches. User WH: {request.user.warehouse}, Order WH: {order_item.order.warehouse}")
            return JsonResponse({'success': False, 'message': 'Permission denied to access batches for this order item.'}, status=403)

        if not order_item.warehouse_product:
            logger.warning(f"OrderItem {order_item.pk} (Product: {order_item.product.sku if order_item.product else 'N/A'}) has no linked WarehouseProduct. Cannot fetch batches.")
            return JsonResponse({'success': True, 'batches': [], 'message': 'Order item is not linked to a specific warehouse product.'})

        warehouse_product_for_item = order_item.warehouse_product
        today = timezone.now().date()

        logger.info(f"Fetching batches for WP: {warehouse_product_for_item.id} (Product: {warehouse_product_for_item.product.sku}), Criteria: Qty > 0, Not Expired (Today: {today})")

        # Query for available batches
        batches_qs = InventoryBatchItem.objects.filter(
            warehouse_product=warehouse_product_for_item,
            quantity__gt=0 # Only batches with stock
        ).exclude(
            expiry_date__isnull=False, expiry_date__lt=today # Exclude expired batches
        ).order_by(
            F('pick_priority').asc(nulls_last=True), # Default/Secondary picks first
            F('expiry_date').asc(nulls_last=True),   # Then FEFO (First-Expiry, First-Out)
            'date_received'                          # Then by received date as a tie-breaker
        )

        logger.info(f"Found {batches_qs.count()} total batches matching quantity/expiry for WP {warehouse_product_for_item.id} before serialization for OI {order_item.pk}.")


        batches_data = []
        for batch in batches_qs:
            priority_label = ""
            if batch.pick_priority == 0: priority_label = " [Default]"
            elif batch.pick_priority == 1: priority_label = " [Secondary]"

            location_display = f"[{batch.location_label}]" if batch.location_label else "NoLoc"
            batch_display = f"Batch: {batch.batch_number}" if batch.batch_number else "NoBatch" # Ensure batch_number is shown
            expiry_display = f"Exp: {batch.expiry_date.strftime('%d/%m/%y')}" if batch.expiry_date else "NoExp"
            qty_display = f"Qty: {batch.quantity}"

            batch_data_entry = {
                'id': batch.pk,
                'display_name': f"{location_display} | {batch_display} | {expiry_display} | {qty_display}{priority_label}",
                'quantity_available': batch.quantity,
                'pick_priority': batch.pick_priority # Send pick_priority for potential JS logic
            }
            batches_data.append(batch_data_entry)
            logger.debug(f"Added batch to dropdown data: ID {batch.pk}, Prio {batch.pick_priority}, Qty {batch.quantity}, Label: ...{priority_label}")


        if not batches_data:
            logger.info(f"No available batches were serialized for OI {order_item.pk} (WP: {warehouse_product_for_item.id}). Dropdown will indicate no batches.")
            return JsonResponse({'success': True, 'batches': [], 'message': 'No available stock batches found for this item.'})

        logger.info(f"Successfully prepared {len(batches_data)} batch options for OI {order_item.pk} dropdown.")
        return JsonResponse({'success': True, 'batches': batches_data})

    except Http404:
        logger.error(f"[get_available_batches] OrderItem PK {order_item_pk} not found.")
        return JsonResponse({'success': False, 'message': 'Order item not found.'}, status=404)
    except Exception as e:
        logger.error(f"[get_available_batches] Unexpected error for OI PK {order_item_pk}: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'An unexpected server error occurred: {str(e)}'}, status=500)


@login_required
def get_order_items_for_editing(request, order_pk):
    # This view remains largely the same as provided by the user.
    # Ensure it correctly calculates balance_quantity_to_pack for the form.
    logger.debug(f"--- get_order_items_for_editing for order_pk: {order_pk} ---")
    order = get_object_or_404(Order.objects.prefetch_related(
        Prefetch('items', queryset=OrderItem.objects.select_related('product').order_by('product__name'))
    ), pk=order_pk)
    logger.debug(f"Order found: {order.erp_order_id}, Status: {order.get_status_display()}")

    if not request.user.is_superuser and (not request.user.warehouse or order.warehouse != request.user.warehouse):
        logger.warning(f"Permission denied for user {request.user.email} to edit order {order_pk}")
        return JsonResponse({'success': False, 'message': 'Permission denied for this order.'}, status=403)

    if order.status != 'PARTIALLY_SHIPPED': # Only allow editing for partially shipped orders
        logger.info(f"Order {order_pk} status is '{order.get_status_display()}', cannot edit items for removal.")
        return JsonResponse({'success': False, 'message': f'Order status is "{order.get_status_display()}", cannot edit items for removal.'}, status=400)

    initial_form_data = []
    logger.debug(f"Processing order items for order {order.erp_order_id}:")
    order_items_all = order.items.all() # Get all items for this order
    logger.debug(f"Total items in order: {order_items_all.count()}")

    for item in order_items_all:
        total_removed_for_this_item = order.get_total_removed_quantity_for_item(item.id)
        # Balance is ordered - packed - (already notionally removed)
        balance_qty = item.quantity_ordered - item.quantity_packed - total_removed_for_this_item

        logger.debug(f"  Item PK: {item.pk}, Prod: {item.product.sku if item.product else 'N/A'}, "
                     f"Ordered: {item.quantity_ordered}, Packed: {item.quantity_packed}, "
                     f"Total Removed Logged: {total_removed_for_this_item}, Calculated Balance: {balance_qty}")

        if balance_qty > 0: # Only include items that still have a balance to be packed/accounted for
            initial_form_data.append({
                'order_item_id': item.pk,
                'product_name': item.product.name if item.product else item.erp_product_name,
                'sku': item.product.sku if item.product else "N/A",
                # 'balance_quantity_to_pack' will be set when creating form instances
            })
            logger.debug(f"    -> ADDED to initial_form_data (PK: {item.pk}, Balance: {balance_qty})")
        else:
            logger.debug(f"    -> SKIPPED for initial_form_data (PK: {item.pk}, Balance: {balance_qty})")


    # Prepare initial data for the formset, including the calculated balance
    formset_initial_with_balance = []
    if not initial_form_data:
        logger.debug("initial_form_data is EMPTY. No items will be shown in the formset.")
    else:
        logger.debug("Populating formset_initial_with_balance...")
        for item_data_initial in initial_form_data:
            # Find the corresponding OrderItem instance to get its current state
            oi = next((i for i in order_items_all if i.pk == item_data_initial['order_item_id']), None)
            if not oi:
                logger.error(f"Could not find OrderItem with PK {item_data_initial['order_item_id']} in prefetched items.")
                continue # Should not happen if initial_form_data was built correctly

            total_removed_for_oi = order.get_total_removed_quantity_for_item(oi.id)
            balance_for_form = oi.quantity_ordered - oi.quantity_packed - total_removed_for_oi
            logger.debug(f"  For Form (OI PK: {oi.pk}): balance_for_form = {balance_for_form}")

            formset_initial_with_balance.append({
                **item_data_initial, # Includes order_item_id, product_name, sku
                'balance_quantity_to_pack': balance_for_form # This will be used by the form's __init__
            })

    logger.debug(f"Final formset_initial_with_balance count: {len(formset_initial_with_balance)}")

    edit_items_formset = RemoveOrderItemFormSet(initial=formset_initial_with_balance, prefix='edititems')
    logger.debug(f"edit_items_formset created. Number of forms: {len(edit_items_formset.forms)}")

    # Log if no forms are generated (e.g., all items are fully packed or accounted for)
    if not edit_items_formset.forms:
        logger.debug("No forms in edit_items_formset. formset_html will likely be empty or show 'no items'.")


    formset_template_name = 'operation/partials/edit_order_formset.html'
    logger.debug(f"Rendering template: {formset_template_name}")
    formset_html_content = render_to_string(
        formset_template_name,
        {'formset': edit_items_formset, 'order': order}, # Pass order for context if template needs it
        request=request
    )

    logger.debug(f"Length of rendered formset_html_content: {len(formset_html_content)}")
    if len(formset_html_content) < 200: # Arbitrary short length to log more detail
        logger.debug(f"Short formset_html_content: '{formset_html_content[:500]}...'")


    # Get the log of previously removed items for display
    removed_items_log_display = order.items_removed_log or [] # Ensure it's a list
    logger.debug(f"Removed items log being sent to client: {removed_items_log_display}")

    logger.debug(f"--- get_order_items_for_editing END for order_pk: {order_pk} ---")
    return JsonResponse({
        'success': True,
        'order_id': order.pk,
        'erp_order_id': order.erp_order_id,
        'customer_name': order.customer.customer_name,
        'formset_html': formset_html_content,
        'removed_items_log': removed_items_log_display, # Send the log
        'message': 'Items loaded for editing/removal.'
    })


@login_required
@transaction.atomic
def process_order_item_removal(request, order_pk):
    # This view remains largely the same as provided by the user.
    # It processes the RemoveOrderItemFormSet.
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

    order = get_object_or_404(Order, pk=order_pk)

    # Permission check
    if not request.user.is_superuser and (not request.user.warehouse or order.warehouse != request.user.warehouse):
        return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)

    # Security check: Only allow removal if order is PARTIALLY_SHIPPED
    if order.status != 'PARTIALLY_SHIPPED':
        return JsonResponse({'success': False, 'message': f'Order status is "{order.get_status_display()}", cannot process removals.'}, status=400)

    # Reconstruct initial data for formset validation, including the correct balance
    # This is crucial because the form's `clean_quantity_to_remove` depends on `self.balance_quantity_to_pack`
    temp_formset_data_from_post = {} # To store POSTed order_item_id and quantity_to_remove
    for key, value in request.POST.items():
        if key.startswith('edititems-') and key.endswith('-order_item_id'):
            form_idx = key.split('-')[1]
            if form_idx not in temp_formset_data_from_post: temp_formset_data_from_post[form_idx] = {}
            temp_formset_data_from_post[form_idx]['order_item_id'] = value
        elif key.startswith('edititems-') and key.endswith('-quantity_to_remove'):
            form_idx = key.split('-')[1]
            if form_idx not in temp_formset_data_from_post: temp_formset_data_from_post[form_idx] = {}
            temp_formset_data_from_post[form_idx]['quantity_to_remove'] = value

    reconstructed_initial_for_formset = []
    for idx_str, data_dict in temp_formset_data_from_post.items():
        try:
            oi_id = int(data_dict['order_item_id'])
            oi = OrderItem.objects.get(pk=oi_id, order=order) # Ensure item belongs to this order
            total_removed_for_oi = order.get_total_removed_quantity_for_item(oi.id)
            balance = oi.quantity_ordered - oi.quantity_packed - total_removed_for_oi
            reconstructed_initial_for_formset.append({
                'order_item_id': oi_id,
                'product_name': oi.product.name, # For display if needed by form, though form makes it readonly
                'sku': oi.product.sku,
                'balance_quantity_to_pack': balance, # This is what the form needs for its max validation
                'quantity_to_remove': data_dict.get('quantity_to_remove', 0) # The value user entered
            })
        except (OrderItem.DoesNotExist, ValueError, KeyError) as e:
            logger.error(f"Error reconstructing formset initial data for validation: {e}")
            return JsonResponse({'success': False, 'message': 'Error processing form data.'}, status=400)


    # Now, create the formset with the POST data and the reconstructed initial data (for balance validation)
    # The formset factory will create forms, and each form's __init__ will use its specific initial data.
    # Note: The `initial` kwarg to formset_factory is for *extra* forms, not for existing ones.
    # We need to pass the data such that each form gets its correct `balance_quantity_to_pack`.
    # This is typically handled by passing `form_kwargs` to the formset if all forms need the same extra arg,
    # or by customizing the formset's `_construct_form` method.
    # For simplicity here, we'll validate each form individually.

    is_formset_valid = True
    cleaned_forms_data = [] # Store cleaned data from each valid form

    # Manually iterate and validate each form
    for i in range(int(request.POST.get('edititems-TOTAL_FORMS', 0))):
        form_data_for_instance = {k.replace(f'edititems-{i}-', ''): v for k, v in request.POST.items() if k.startswith(f'edititems-{i}-')}

        # Find the corresponding reconstructed initial data for this form instance
        current_form_oi_id = int(form_data_for_instance.get('order_item_id', 0))
        balance_for_this_form = 0
        for init_data in reconstructed_initial_for_formset:
            if init_data['order_item_id'] == current_form_oi_id:
                balance_for_this_form = init_data['balance_quantity_to_pack']
                break

        # Instantiate the form with POST data and the specific balance for validation
        form = RemoveOrderItemForm(
            form_data_for_instance, # This is the POST data for this form
            balance_quantity_to_pack=balance_for_this_form # Pass the balance to __init__
        )

        if form.is_valid():
            cleaned_forms_data.append(form.cleaned_data)
        else:
            is_formset_valid = False
            logger.warning(f"Form {i} errors: {form.errors.as_json()}")
            # Collect all form errors to return
            # For now, just return a generic message on first error
            return JsonResponse({'success': False, 'message': 'Invalid data in removal form. Please check quantities.'}, status=400)


    if is_formset_valid:
        removed_items_summary_for_log = order.items_removed_log or [] # Get existing log or start new
        any_actual_removal = False

        for data in cleaned_forms_data:
            order_item_id = data.get('order_item_id')
            qty_removed = data.get('quantity_to_remove', 0) # Default to 0 if not present

            if qty_removed > 0:
                try:
                    order_item = OrderItem.objects.get(pk=order_item_id, order=order)
                    # Log the removal
                    removed_items_summary_for_log.append({
                        'order_item_id': order_item.id,
                        'product_sku': order_item.product.sku,
                        'product_name': order_item.product.name,
                        'removed_qty': qty_removed,
                        'removed_at': timezone.now().isoformat() # Store timestamp of removal
                    })
                    any_actual_removal = True
                except OrderItem.DoesNotExist:
                    # This should have been caught if form_data_for_instance was correctly linked
                    messages.error(request, f"Error: Order Item ID {order_item_id} not found for this order.")
                    return JsonResponse({'success': False, 'message': f"Item ID {order_item_id} not found."}, status=400)

        if any_actual_removal:
            order.items_removed_log = removed_items_summary_for_log
            # The order.save() will trigger update_status_based_on_items_and_parcels
            # which now needs to consider items_removed_log.

        order.save() # This will call update_status_based_on_items_and_parcels
        order.refresh_from_db() # Get the potentially updated status

        messages.success(request, "Order items updated successfully. Status refreshed.")
        return JsonResponse({
            'success': True,
            'message': 'Items processed successfully! Order status may have been updated.',
            'new_order_status': order.get_status_display(), # Send back new status
            'order_id': order.pk
        })
    else:
        # This case should ideally be caught by individual form validation returning early
        logger.error(f"Order item removal formset was not valid for order {order_pk}.")
        return JsonResponse({'success': False, 'message': 'Invalid form data submitted.'}, status=400)


@require_POST
@login_required
@transaction.atomic
def remove_parcel(request, parcel_pk):
    """
    Removes a parcel, returns its items to inventory, updates the
    OrderItem's packed quantity, and recalculates the order status.
    """
    logger.info(f"User {request.user.email} initiated removal for Parcel PK: {parcel_pk}")
    parcel = get_object_or_404(Parcel.objects.select_related('order', 'order__warehouse'), pk=parcel_pk)
    order = parcel.order

    # --- Permission & Status Checks ---
    if not request.user.is_superuser and (not request.user.warehouse or order.warehouse != request.user.warehouse):
        logger.warning(f"Permission denied for user {request.user.email} on Parcel PK: {parcel_pk}")
        return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)

    if parcel.status not in ['PREPARING_TO_PACK', 'READY_TO_SHIP']:
        logger.warning(f"Attempted to remove parcel {parcel.parcel_code_system} with invalid status: {parcel.get_status_display()}")
        return JsonResponse({
            'success': False,
            'message': f'Parcel cannot be removed as it is already {parcel.get_status_display()}.'
        }, status=400)

    try:
        logger.debug(f"Starting stock return process for parcel {parcel.parcel_code_system}")

        parcel_items_to_process = parcel.items_in_parcel.select_related(
            'order_item',
            'shipped_from_batch',
            'shipped_from_batch__warehouse_product'
        ).all()

        for parcel_item in parcel_items_to_process:
            order_item = parcel_item.order_item
            batch_item = parcel_item.shipped_from_batch
            quantity_to_return = parcel_item.quantity_shipped_in_this_parcel

            logger.debug(f"Processing ParcelItem PK: {parcel_item.pk}. Qty to return: {quantity_to_return}")

            if order_item:
                order_item.quantity_packed = F('quantity_packed') - quantity_to_return
                order_item.save(update_fields=['quantity_packed'])

            if batch_item:
                # --- START: THE FIX ---
                # Add stock back to both the specific batch AND the parent WarehouseProduct
                batch_item.quantity = F('quantity') + quantity_to_return
                batch_item.save(update_fields=['quantity'])

                warehouse_product = batch_item.warehouse_product
                if warehouse_product:
                    warehouse_product.quantity = F('quantity') + quantity_to_return
                    warehouse_product.save(update_fields=['quantity'])
                # --- END: THE FIX ---

                if not warehouse_product:
                    error_message = f"Data integrity error: BatchItem {batch_item.pk} is missing its WarehouseProduct link."
                    logger.error(error_message)
                    raise IntegrityError(error_message)

                StockTransaction.objects.create(
                    warehouse=order.warehouse,
                    warehouse_product=warehouse_product,
                    product=order_item.product,
                    transaction_type=StockTransaction.TransactionTypes.RETURN_IN,
                    quantity=quantity_to_return,
                    batch_item_involved=batch_item,
                    reference_note=f"Return from deleted parcel {parcel.parcel_code_system} for Order {order.erp_order_id}",
                    related_order=order,
                    recorded_by=request.user
                )
            else:
                logger.warning(f"ParcelItem PK: {parcel_item.pk} has no associated batch. Cannot return stock automatically.")

        parcel_code_system = parcel.parcel_code_system
        parcel.delete()
        logger.info(f"Parcel {parcel_code_system} (PK: {parcel_pk}) deleted successfully.")

        order.update_status_based_on_items_and_parcels()
        order.save()
        logger.info(f"Order {order.erp_order_id} status updated to '{order.get_status_display()}' after parcel removal.")

        messages.success(request, f"Parcel {parcel_code_system} has been removed.")
        return JsonResponse({
            'success': True,
            'message': 'Parcel removed successfully!',
            'redirect_url': reverse('operation:order_list') + f"?tab={DEFAULT_PARCELS_TAB}"
        })

    except Exception as e:
        logger.error(f"Error removing parcel {parcel_pk}: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': 'An unexpected server error occurred.'}, status=500)

# --- NEW VIEWS FOR PARCEL CUSTOMS DETAILS ---
@login_required
def get_parcel_details_for_editing(request, parcel_pk):
    """
    Gets all data needed to populate the 'View/Edit Parcel' modal.
    This version is updated to use the new customers.Customer model.
    """
    # --- UPDATED: Added 'order__customer' to select_related for efficiency ---
    parcel = get_object_or_404(
        Parcel.objects.select_related(
            'order__warehouse', 'order__customer', 'packaging_type',
            'courier_company', 'customs_declaration'
        ).prefetch_related(models.Prefetch('items_in_parcel', queryset=ParcelItem.objects.select_related('order_item__product'))),
        pk=parcel_pk
    )

    if not request.user.is_superuser and (not request.user.warehouse or parcel.order.warehouse != request.user.warehouse):
        return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)

    # --- Filtering logic for declarations remains the same and is correct ---
    courier_filter = Q(courier_companies__isnull=True)
    if parcel.courier_company:
        courier_filter |= Q(courier_companies=parcel.courier_company)

    effective_env_type = None
    if parcel.packaging_type and parcel.packaging_type.environment_type in ['COLD', 'AMBIENT']:
        effective_env_type = parcel.packaging_type.environment_type
    elif parcel.order.is_cold_chain:
        effective_env_type = 'COLD'
    else:
        effective_env_type = 'AMBIENT'
    shipment_type_filter = Q()
    if effective_env_type == 'COLD':
        shipment_type_filter = Q(applies_to_cold_chain=True) | Q(applies_to_mix=True)
    elif effective_env_type == 'AMBIENT':
        shipment_type_filter = Q(applies_to_ambient=True) | Q(applies_to_mix=True)
    declarations_queryset = CustomsDeclaration.objects.filter(
        Q(warehouse=parcel.order.warehouse) & courier_filter & shipment_type_filter
    ).distinct().order_by('description')

    declarations_json = list(declarations_queryset.values('pk', 'description', 'hs_code'))

    # Instantiate forms
    parcel_form = ParcelCustomsDetailForm(instance=parcel, declarations_queryset=declarations_queryset)
    item_formset = ParcelItemCustomsDetailFormSet(instance=parcel, prefix='parcelitems')

    # Prepare other data
    env_type = 'COLD' if parcel.order.is_cold_chain else 'AMBIENT'
    available_packaging = list(PackagingType.objects.filter(environment_type=env_type, is_active=True).values('id', 'name', 'default_length_cm', 'default_width_cm', 'default_height_cm'))

    # --- THE FIX: Get customer info from the new 'order.customer' relationship ---
    parcel_data = {
        'parcel_code_system': parcel.parcel_code_system,
        'courier_name': parcel.courier_company.name if parcel.courier_company else "N/A",
        'tracking_number': parcel.tracking_number,
        'status_display': parcel.get_status_display(),
        'packaging_type_display': parcel.get_packaging_type_display(),
        'created_at': parcel.created_at.strftime('%Y-%m-%d %H:%M') if parcel.created_at else "N/A",
        'shipped_at': parcel.shipped_at.strftime('%Y-%m-%d %H:%M') if parcel.shipped_at else "N/A",

        # Correctly access customer data
        'customer_name': parcel.order.customer.customer_name if parcel.order.customer else "N/A",
        'company_name': parcel.order.customer.company_name if parcel.order.customer else "",
        'recipient_address_line1': parcel.order.customer.address_line1 if parcel.order.customer else "",
        'recipient_address_city': parcel.order.customer.city if parcel.order.customer else "",
        'recipient_address_state': parcel.order.customer.state if parcel.order.customer else "",
        'recipient_address_zip': parcel.order.customer.zip_code if parcel.order.customer else "",
        'recipient_address_country': parcel.order.customer.country if parcel.order.customer else "",
        'recipient_phone': parcel.order.customer.phone_number if parcel.order.customer else "",
    }

    # Prepare context for rendering the template
    context = {
        'parcel': parcel,
        'parcel_form': parcel_form,
        'item_formset': item_formset,
        'parcel_data': parcel_data,
        'available_packaging': available_packaging,
        'declarations_for_template': declarations_queryset,
    }
    form_html = render_to_string('operation/partials/_parcel_edit_form_content.html', context, request=request)

    # Return all data as a JSON response
    return JsonResponse({
        'success': True,
        'form_html': form_html,
        'parcel_data': parcel_data,
        'available_packaging': available_packaging,
        'declarations_json': declarations_json,
    })


@login_required
@transaction.atomic
def update_parcel_customs_details(request, parcel_pk):
    """
    Handles the AJAX POST request from the parcel edit modal to save customs details.
    This view now includes the corrected filtering logic and automatic status update.
    """
    if request.method != 'POST':
        logger.warning(f"update_parcel_customs_details received non-POST request for parcel {parcel_pk}")
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

    parcel = get_object_or_404(
        Parcel.objects.select_related('order__warehouse', 'courier_company', 'packaging_type'),
        pk=parcel_pk
    )
    logger.info(f"Updating customs details for Parcel PK: {parcel_pk}, System Code: {parcel.parcel_code_system}")

    if not request.user.is_superuser and (not request.user.warehouse or parcel.order.warehouse != request.user.warehouse):
        logger.warning(f"User {request.user.email} permission denied for parcel {parcel_pk}.")
        return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)

    courier_filter = Q(courier_companies__isnull=True)
    if parcel.courier_company:
        courier_filter |= Q(courier_companies=parcel.courier_company)

    effective_env_type = None
    if parcel.packaging_type and parcel.packaging_type.environment_type in ['COLD', 'AMBIENT']:
        effective_env_type = parcel.packaging_type.environment_type
    elif parcel.order.is_cold_chain:
        effective_env_type = 'COLD'
    else:
        effective_env_type = 'AMBIENT'

    shipment_type_filter = Q()
    if effective_env_type == 'COLD':
        shipment_type_filter = Q(applies_to_cold_chain=True) | Q(applies_to_mix=True)
    elif effective_env_type == 'AMBIENT':
        shipment_type_filter = Q(applies_to_ambient=True) | Q(applies_to_mix=True)

    valid_declarations_qs = CustomsDeclaration.objects.filter(
        Q(warehouse=parcel.order.warehouse) & courier_filter & shipment_type_filter
    ).distinct()

    parcel_form = ParcelCustomsDetailForm(request.POST, instance=parcel, declarations_queryset=valid_declarations_qs)
    item_formset = ParcelItemCustomsDetailFormSet(request.POST, instance=parcel, prefix='parcelitems')

    if parcel_form.is_valid() and item_formset.is_valid():
        logger.info(f"Forms are valid for parcel {parcel_pk}.")

        updated_parcel = parcel_form.save(commit=False)

        length = parcel_form.cleaned_data.get('length')
        width = parcel_form.cleaned_data.get('width')
        height = parcel_form.cleaned_data.get('height')

        if length and width and height:
            try:
                updated_parcel.dimensional_weight_kg = (Decimal(str(length)) * Decimal(str(width)) * Decimal(str(height))) / Decimal('5000')
            except Exception as e:
                logger.error(f"Error calculating dimensional weight for parcel {parcel_pk}: {e}")
                updated_parcel.dimensional_weight_kg = None
        else:
            updated_parcel.dimensional_weight_kg = None

        updated_parcel.save()
        item_formset.save()

        # --- START: NEW LOGIC FOR STATUS UPDATE ---
        # If the status is 'PREPARING_TO_PACK', update it to 'READY_TO_SHIP'
        if updated_parcel.status == 'PREPARING_TO_PACK':
            updated_parcel.status = 'READY_TO_SHIP'
            updated_parcel.save(update_fields=['status'])
            logger.info(f"Parcel {updated_parcel.pk} status automatically updated to 'READY_TO_SHIP'.")
        # --- END: NEW LOGIC ---

        logger.info(f"Successfully updated customs details for parcel {parcel_pk}.")
        messages.success(request, f"Customs details for Parcel {parcel.parcel_code_system} updated successfully.")
        return JsonResponse({'success': True, 'message': 'Customs details updated successfully.'})
    else:
        errors = {}
        if parcel_form.errors:
            errors['parcel_form'] = parcel_form.errors.as_json()
        if item_formset.errors:
            formset_errors_list = []
            for form_errors in item_formset.errors:
                if form_errors:
                    formset_errors_list.append(form_errors.as_json())
            if formset_errors_list:
                errors['item_formset'] = formset_errors_list
        if item_formset.non_form_errors():
            errors['item_formset_non_form'] = item_formset.non_form_errors()

        return JsonResponse({'success': False, 'message': "Validation failed. Please check the form details.", 'errors': errors}, status=400)

@login_required
def get_declarations_for_courier(request, parcel_pk):
    """
    Returns a JSON list of suitable customs declarations for a given parcel
    and a newly selected courier company.
    """
    courier_id = request.GET.get('courier_id')
    if not courier_id:
        return JsonResponse({'success': False, 'message': 'Courier ID is required.'}, status=400)

    try:
        parcel = get_object_or_404(Parcel.objects.select_related('order', 'packaging_type'), pk=parcel_pk)
        courier = get_object_or_404(CourierCompany, pk=courier_id)

        # Re-use the same reliable filtering logic from get_parcel_details_for_editing
        courier_filter = Q(courier_companies=courier) | Q(courier_companies__isnull=True)

        effective_env_type = None
        if parcel.packaging_type and parcel.packaging_type.environment_type in ['COLD', 'AMBIENT']:
            effective_env_type = parcel.packaging_type.environment_type
        elif parcel.order.is_cold_chain:
            effective_env_type = 'COLD'
        else:
            effective_env_type = 'AMBIENT'

        shipment_type_filter = Q()
        if effective_env_type == 'COLD':
            shipment_type_filter = Q(applies_to_cold_chain=True) | Q(applies_to_mix=True)
        else: # AMBIENT
            shipment_type_filter = Q(applies_to_ambient=True) | Q(applies_to_mix=True)

        declarations_queryset = CustomsDeclaration.objects.filter(
            Q(warehouse=parcel.order.warehouse) & courier_filter & shipment_type_filter
        ).distinct().order_by('description')

        declarations_json = list(declarations_queryset.values('pk', 'description', 'hs_code'))

        return JsonResponse({'success': True, 'declarations': declarations_json})

    except (Parcel.DoesNotExist, CourierCompany.DoesNotExist):
        return JsonResponse({'success': False, 'message': 'Invalid Parcel or Courier.'}, status=404)
    except Exception as e:
        logger.error(f"Error fetching declarations for courier {courier_id}: {e}")
        return JsonResponse({'success': False, 'message': 'An unexpected error occurred.'}, status=500)

@login_required
def manage_customs_declarations(request):
    """
    Handles the display and filtering of Customs Declarations without a status filter.
    """
    user = request.user

    # --- Permission Check ---
    if not (user.is_superuser or user.warehouse):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('inventory:inventory_batch_list_view')

    # --- Form Handling for Creating New Declarations (POST) ---
    if request.method == 'POST':
        form = CustomsDeclarationForm(request.POST, user=user)
        if form.is_valid():
            declaration = form.save(commit=False)
            if not user.is_superuser:
                declaration.warehouse = user.warehouse
            declaration.save()
            form.save_m2m()
            messages.success(request, "Customs declaration added successfully.")
            return redirect('operation:manage_customs_declarations')
        else:
            messages.error(request, "Could not add declaration. Please correct the errors below.")
    else:
        # For a GET request, create a fresh form for the "Add" modal
        form = CustomsDeclarationForm(user=user)

    # --- Filtering Logic (GET) ---
    declarations_qs = CustomsDeclaration.objects.select_related('warehouse').prefetch_related('courier_companies').all()

    # Get filter parameters from the URL
    selected_courier_id = request.GET.get('courier_company')
    selected_shipment_type = request.GET.get('shipment_type')

    if not user.is_superuser:
        declarations_qs = declarations_qs.filter(warehouse=user.warehouse)

    # Apply courier filter (simplified logic)
    if selected_courier_id:
        if selected_courier_id == "generic":
            declarations_qs = declarations_qs.filter(courier_companies__isnull=True)
        else:
            # Directly filter by the selected courier ID
            declarations_qs = declarations_qs.filter(courier_companies__id=selected_courier_id)

    # Apply shipment type filter
    if selected_shipment_type:
        shipment_filters = {
            'AMBIENT': Q(applies_to_ambient=True),
            'COLD_CHAIN': Q(applies_to_cold_chain=True),
            'MIX': Q(applies_to_mix=True)
        }
        if selected_shipment_type in shipment_filters:
            declarations_qs = declarations_qs.filter(shipment_filters[selected_shipment_type])

    # Final queryset ordering
    declarations = declarations_qs.order_by('-is_active', 'description', 'hs_code')

    # --- START OF THE FIX ---
    # Prepare the data structure the template expects
    declaration_data = []
    for decl in declarations:
        declaration_data.append({
            'declaration': decl,
            'edit_form': CustomsDeclarationForm(instance=decl, user=user)
        })
    # --- END OF THE FIX ---


    # --- Prepare Context for Template ---
    context = {
        'form': form,
        'declarations': declarations,
        'declaration_data': declaration_data, # Pass the corrected data structure
        'couriers': CourierCompany.objects.all().order_by('name'),
        'warehouses': Warehouse.objects.all().order_by('name'),
        'shipment_type_choices': [('AMBIENT', 'Ambient Only'), ('COLD_CHAIN', 'Cold Chain Only'), ('MIX', 'Mixed')],
        'selected_courier_id': selected_courier_id,
        'selected_shipment_type': selected_shipment_type,
        'page_title': "Manage Customs Declarations",
    }
    return render(request, 'operation/manage_customs_declarations.html', context)


@login_required
@require_POST
def edit_customs_declaration(request, pk):
    """
    Handles updating an existing customs declaration.
    """
    declaration = get_object_or_404(CustomsDeclaration, pk=pk)
    user = request.user

    # Security check: Ensure non-superusers can only edit declarations for their own warehouse
    if not user.is_superuser and declaration.warehouse != user.warehouse:
        messages.error(request, "You do not have permission to edit this declaration.")
        return redirect('operation:manage_customs_declarations')

    if request.method == 'POST':
        # Create a mutable copy of the POST data
        post_data = request.POST.copy()

        # Handle the 'is_active' checkbox
        # If 'is_active' is not in the POST data, it means the box was unchecked.
        if 'is_active' not in post_data:
            post_data['is_active'] = False

        form = CustomsDeclarationForm(post_data, instance=declaration, user=user)

        if form.is_valid():
            updated_declaration = form.save(commit=False)
            # Ensure warehouse is not accidentally changed by a non-superuser
            if not user.is_superuser:
                updated_declaration.warehouse = user.warehouse

            updated_declaration.save()
            form.save_m2m() # Save many-to-many fields
            messages.success(request, f"Successfully updated declaration: '{declaration.description}'.")
            return redirect('operation:manage_customs_declarations')
        else:
            # Combine form errors into a single message
            error_message = "Could not update declaration. "
            for field, errors in form.errors.items():
                error_message += f"{field.replace('_', ' ').title()}: {', '.join(errors)} "
            messages.error(request, error_message)

    # This redirect is a fallback for GET requests or failed POSTs to avoid rendering a separate page
    return redirect('operation:manage_customs_declarations')


# delete_customs_declaration view remains the same as provided in the previous step.
@login_required
@require_POST
def delete_customs_declaration(request, pk):
    if not (request.user.is_superuser or request.user.warehouse):
        messages.error(request, "You do not have permission to perform this action.")
        return redirect('operation:manage_customs_declarations')

    declaration = get_object_or_404(CustomsDeclaration, pk=pk)
    try:
        desc_preview = declaration.description[:30]
        declaration.delete()
        messages.success(request, f"Declaration '{desc_preview}...' deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting declaration: {str(e)}")
    return redirect('operation:manage_customs_declarations')


@login_required
def packaging_management(request):
    """
    Manages the display and creation of PackagingTypes and PackagingMaterials.
    The form for adding a new packaging type is simplified by removing dimension fields.
    """
    user = request.user

    all_global_materials = PackagingMaterial.objects.order_by('name')

    # Prepare initial data for the formset: one entry for each global material
    component_formset_initial = [{'packaging_material': mat} for mat in all_global_materials]
    component_formset = PackagingTypeMaterialComponentFormSet(
        initial=component_formset_initial,
        prefix='components'
    )

    packaging_type_form = PackagingTypeForm(user=user)
    packaging_material_form = PackagingMaterialForm()
    receive_stock_form = ReceivePackagingStockForm()


    if request.method == 'POST':
        if 'submit_packaging_type' in request.POST:
            form = PackagingTypeForm(request.POST, user=user)
            if form.is_valid():
                try:
                    with transaction.atomic():
                        # Save the main PackagingType instance
                        packaging_type_instance = form.save(commit=False)
                        if not user.is_superuser and user.warehouse:
                            packaging_type_instance.warehouse = user.warehouse
                        packaging_type_instance.save()

                        # Process the material components
                        for i in range(len(all_global_materials)):
                            material_id = request.POST.get(f'components-{i}-packaging_material')
                            quantity_str = request.POST.get(f'components-{i}-quantity')
                            if quantity_str and int(quantity_str) > 0:
                                material = PackagingMaterial.objects.get(pk=material_id)
                                PackagingTypeMaterialComponent.objects.create(
                                    packaging_type=packaging_type_instance,
                                    packaging_material=material,
                                    quantity=int(quantity_str)
                                )

                        messages.success(request, 'New packaging type added successfully.')
                        return redirect('operation:packaging_management')
                except Exception as e:
                    messages.error(request, f"An error occurred: {e}")
            else:
                messages.error(request, 'Error adding packaging type. Please check the form.')
                packaging_type_form = form # Return form with errors

        elif 'submit_global_material' in request.POST:
            # --- NEW: Logic for adding a global material ---
            form = PackagingMaterialForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'New global packaging material created successfully.')
                return redirect('operation:packaging_management')
            else:
                packaging_material_form = form
                messages.error(request, 'Error creating global material.')

        elif 'submit_receive_stock' in request.POST:
            form = ReceivePackagingStockForm(request.POST)
            if form.is_valid():
                try:
                    with transaction.atomic():
                        # --- START OF THE FIX ---
                        # 1. Get the selected global material and determine the warehouse
                        global_material = form.cleaned_data['packaging_material']
                        quantity_to_add = form.cleaned_data['quantity']

                        warehouse = None
                        if request.user.is_superuser:
                            # For superusers, you might need a warehouse selector in the form.
                            # For now, let's assume it defaults to the first one or you add that field.
                            # This is a placeholder for that logic.
                            warehouse = Warehouse.objects.first() # Or handle via a form field
                        else:
                            warehouse = request.user.warehouse

                        if not warehouse:
                            raise Exception("Cannot determine the target warehouse.")

                        # 2. Get or Create the warehouse-specific stock record
                        stock_item, created = WarehousePackagingMaterial.objects.get_or_create(
                            packaging_material=global_material,
                            warehouse=warehouse,
                            defaults={'current_stock': 0} # Default stock is 0 if created
                        )

                        # 3. Create the historical transaction record
                        PackagingStockTransaction.objects.create(
                            warehouse_packaging_material=stock_item,
                            transaction_type=PackagingStockTransaction.TransactionTypes.STOCK_IN,
                            quantity=quantity_to_add,
                            notes=form.cleaned_data['notes'],
                            recorded_by=request.user
                        )

                        # 4. Update the main stock level
                        stock_item.current_stock = F('current_stock') + quantity_to_add
                        stock_item.save()
                        # --- END OF THE FIX ---

                    messages.success(request, f"Successfully received {quantity_to_add} units of {global_material.name}.")

                except Exception as e:
                    messages.error(request, f"An unexpected server error occurred: {e}")

                return redirect('operation:packaging_management')
            else:
                receive_stock_form = form
                messages.error(request, 'Error receiving stock. Please check the form.')

    # For both GET requests and POST requests with errors,
    # remove dimension fields from the "Add" form.
    if 'default_length_cm' in packaging_type_form.fields:
        del packaging_type_form.fields['default_length_cm']
    if 'default_width_cm' in packaging_type_form.fields:
        del packaging_type_form.fields['default_width_cm']
    if 'default_height_cm' in packaging_type_form.fields:
        del packaging_type_form.fields['default_height_cm']

    # Querysets for display and data preparation
    packaging_types_qs = PackagingType.objects.select_related('warehouse').prefetch_related(
        'packagingtypematerialcomponent_set__packaging_material'
    ).order_by('name')

    # This now gets the global materials for dropdowns
    all_materials_qs = PackagingMaterial.objects.order_by('name')

    # This gets the warehouse-specific stock for display
    warehouse_stock_qs = WarehousePackagingMaterial.objects.select_related(
        'packaging_material', 'warehouse'
    ).order_by('packaging_material__name', 'warehouse__name')


    if not request.user.is_superuser and request.user.warehouse:
        packaging_types_qs = packaging_types_qs.filter(warehouse=request.user.warehouse)
        warehouse_stock_qs = warehouse_stock_qs.filter(warehouse=request.user.warehouse)

    all_materials_json = json.dumps(list(all_global_materials.values('pk', 'name')), cls=DjangoJSONEncoder)

    packaging_types_data = []
    for pt in packaging_types_qs:
        components = pt.packagingtypematerialcomponent_set.all()
        component_map = {comp.packaging_material.id: comp.quantity for comp in components}

        packaging_type_dict = {
            'pk': pt.pk,
            'name': pt.name,
            'type_code': pt.type_code or "",
            'environment_type': pt.environment_type,
            'is_active': pt.is_active,
            'default_length_cm': pt.default_length_cm,
            'default_width_cm': pt.default_width_cm,
            'default_height_cm': pt.default_height_cm,
        }

        packaging_types_data.append({
            'instance': pt,
            'components': components,
            'packaging_type_json': json.dumps(packaging_type_dict, cls=DjangoJSONEncoder),
            'component_map_json': json.dumps(component_map, cls=DjangoJSONEncoder)
        })

    today = timezone.now().date()

    context = {
        'page_title': 'Packaging Management',
        'packaging_type_form': packaging_type_form,
        'packaging_material_form': packaging_material_form,
        'receive_stock_form': receive_stock_form,
        'packaging_types': packaging_types_qs,
        'warehouse_packaging_stocks': warehouse_stock_qs,
        'all_materials_json': all_materials_json,
        'all_global_materials': all_global_materials,
        'today': today, 
    }
    return render(request, 'operation/packaging_management.html', context)



@login_required
def load_edit_packaging_type_form(request, pk):
    """
    This view is called by HTMX to load the edit form into a modal.
    It now pre-processes the materials list to remove the need for custom template filters.
    """
    packaging_type = get_object_or_404(PackagingType, pk=pk)
    form = PackagingTypeForm(instance=packaging_type)
    all_materials = PackagingMaterial.objects.all()

    # Create a map of existing components for easy lookup
    component_map = {comp.packaging_material.id: comp.quantity for comp in packaging_type.packagingtypematerialcomponent_set.all()}

    # Build a list of dictionaries with all data the template needs
    materials_data = []
    for material in all_materials:
        materials_data.append({
            'id': material.id,
            'name': material.name,
            'quantity': component_map.get(material.id, '')  # Get current quantity or an empty string
        })

    context = {
        'form': form,
        'packaging_type': packaging_type,
        'materials_data': materials_data, # Pass this new, pre-processed list
    }
    return render(request, 'operation/partials/_edit_packaging_type_form.html', context)


@login_required
def edit_packaging_type(request, pk):
    """
    Handles the POST submission for updating a PackagingType from the modal form.
    """
    packaging_type = get_object_or_404(PackagingType, pk=pk)

    if request.method == 'POST':
        form = PackagingTypeForm(request.POST, instance=packaging_type)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Save the main form data (name, code, etc.)
                    updated_packaging_type = form.save()

                    # Keep track of materials that have a quantity in the form
                    processed_material_ids = set()

                    # Iterate through all POST items to find material quantities
                    for key, value in request.POST.items():
                        if key.startswith('quantity_') and value.strip():
                            try:
                                material_id = int(key.split('_')[1])
                                quantity = int(value)

                                if quantity > 0:
                                    # This material should be linked to the packaging type
                                    material = get_object_or_404(PackagingMaterial, id=material_id)

                                    # Use update_or_create for efficiency. It will either update an
                                    # existing component or create a new one.
                                    PackagingTypeMaterialComponent.objects.update_or_create(
                                        packaging_type=updated_packaging_type,
                                        packaging_material=material,
                                        defaults={'quantity': quantity}
                                    )
                                    processed_material_ids.add(material_id)

                            except (ValueError, IndexError, PackagingMaterial.DoesNotExist):
                                # Ignore if the key is malformed or material doesn't exist
                                continue

                    # Remove any components that were not submitted in the form
                    # (i.e., their quantity was cleared or set to 0)
                    PackagingTypeMaterialComponent.objects.filter(
                        packaging_type=updated_packaging_type
                    ).exclude(
                        packaging_material_id__in=processed_material_ids
                    ).delete()

                messages.success(request, f"Packaging Type '{updated_packaging_type.name}' updated successfully.")

            except Exception as e:
                messages.error(request, f"An error occurred while saving: {e}")

            return redirect('operation:packaging_management')
        else:
            # If the form has validation errors (e.g., name is blank)
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error in '{field}': {error}")
            return redirect('operation:packaging_management')

    # If someone tries to access this URL with a GET request, just redirect them
    # as the form is now loaded into the modal by a different view.
    return redirect('operation:packaging_management')


@login_required
def get_customer_shipment_history(request, customer_pk):
    """
    Fetches all parcels for a specific customer to display in a modal.
    """
    customer = get_object_or_404(Customer, pk=customer_pk)

    # Get all parcels for this customer, across all their orders
    parcels = Parcel.objects.filter(order__customer=customer).select_related(
        'order', 'courier_company'
    ).order_by('-created_at')

    context = {
        'customer': customer,
        'parcels': parcels,
    }
    return render(request, 'operation/partials/_customer_shipments_modal_content.html', context)


@login_required
def get_airway_bill_details(request, parcel_pk):
    """
    Fetches all necessary data to populate the Air Waybill modal.
    """
    parcel = get_object_or_404(
        Parcel.objects.select_related(
            'order__customer',
            'order__warehouse',
            'courier_company',
            'customs_declaration'
        ),
        pk=parcel_pk
    )

    context = {
        'parcel': parcel,
        'form': AirwayBillForm(instance=parcel) # Pass an instance of the new form
    }
    return render(request, 'operation/partials/_airway_bill_modal_content.html', context)


@require_POST
@login_required
def save_airway_bill(request, parcel_pk):
    """
    Saves the Tracking ID and Estimated Cost from the Air Waybill modal.
    The parcel status is automatically updated by the model's save method.
    """
    parcel = get_object_or_404(Parcel, pk=parcel_pk)
    form = AirwayBillForm(request.POST, instance=parcel)

    if form.is_valid():
        # The form.save() call will trigger the Parcel model's save() method,
        # which correctly updates the status to READY_TO_SHIP if a tracking
        # number was just added.
        updated_parcel = form.save()

        return JsonResponse({
            'success': True,
            'message': 'Air Waybill details saved successfully.',
            'parcel_pk': updated_parcel.pk  # Return parcel PK for JavaScript
        })
    else:
        error_message = ". ".join([f"{field}: {error[0]}" for field, error in form.errors.items()])
        return JsonResponse({'success': False, 'message': error_message or "Invalid data submitted."}, status=400)

@login_required
def get_parcel_tracking_history(request, parcel_pk):
    """
    Fetches the full tracking log for a parcel to display in a modal.
    """
    parcel = get_object_or_404(
        Parcel.objects.prefetch_related('tracking_logs'),
        pk=parcel_pk
    )
    # The tracking logs are already ordered by timestamp due to the model's Meta ordering
    return render(request, 'operation/partials/_parcel_tracking_history_modal.html', {'parcel': parcel})


@require_POST
@login_required
def trace_selected_parcels(request):
    """
    Handles the 'Trace Selected' button click for bulk manual tracking updates.
    Loops through selected parcels and calls the tracking service for each one.
    """
    try:
        data = json.loads(request.body)
        parcel_ids = data.get('parcel_ids', [])
        logger.debug(f"[TraceParcels] Received request to trace parcel IDs: {parcel_ids}")

        if not parcel_ids:
            return JsonResponse({'success': False, 'message': 'No parcels selected.'}, status=400)

        parcels_to_trace = Parcel.objects.filter(id__in=parcel_ids)
        success_count = 0
        error_count = 0
        status_updates = []

        for parcel in parcels_to_trace:
            try:
                # Call the centralized service function for each parcel
                success, message = update_parcel_tracking_from_api(parcel)
                if success:
                    success_count += 1
                    # Optional: Log the specific message for each parcel
                    logger.info(f"Tracking update for Parcel {parcel.id}: {message}")
                    if "Status updated" in message:
                        status_updates.append(parcel.parcel_code_system)
                else:
                    error_count += 1
                    logger.warning(f"Tracking update failed for Parcel {parcel.id}: {message}")

            except Exception as e:
                # Catch any unexpected errors during the service call for a single parcel
                logger.error(f"[TraceParcels] Unexpected error processing parcel {parcel.id}: {e}", exc_info=True)
                error_count += 1

        # --- Construct a final summary message for the user ---
        final_message = f"Tracking update complete. Success: {success_count}, Failed: {error_count}."
        if status_updates:
            final_message += f" Status changed for: {', '.join(status_updates)}."

        return JsonResponse({'success': True, 'message': final_message})

    except json.JSONDecodeError:
        logger.error("[TraceParcels] Invalid JSON in request body.")
        return JsonResponse({'success': False, 'message': 'Invalid JSON in request body.'}, status=400)
    except Exception as e:
        logger.error(f"[TraceParcels] Critical error in view: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': 'A critical server error occurred.'}, status=500)


@login_required
def print_selected_parcels(request):
    parcel_ids_str = request.GET.get('ids', '')
    if not parcel_ids_str:
        return HttpResponse("No parcel IDs provided.", status=400)

    parcel_ids = [int(id) for id in parcel_ids_str.split(',') if id.isdigit()]

    parcels = Parcel.objects.filter(id__in=parcel_ids).select_related(
        'order__customer', 'courier_company', 'packaging_type'
    ).prefetch_related(
        'items_in_parcel__order_item__product'
    )

    if not parcels:
        return HttpResponse("No valid parcels found for the given IDs.", status=404)

    # Summary Data Calculation
    total_parcels = parcels.count()
    cold_chain_count = parcels.filter(order__is_cold_chain=True).count()
    ambient_count = total_parcels - cold_chain_count

    courier_counts = parcels.values('courier_company__name').annotate(count=Count('courier_company__name')).order_by('-count')

    product_quantities = {}
    for parcel in parcels:
        for item in parcel.items_in_parcel.all():
            product_name = item.order_item.product.name
            quantity = item.quantity_shipped_in_this_parcel
            if product_name in product_quantities:
                product_quantities[product_name] += quantity
            else:
                product_quantities[product_name] = quantity

    sorted_product_quantities = sorted(product_quantities.items(), key=lambda x: x[1], reverse=True)


    context = {
        'parcels': parcels,
        'summary': {
            'total_parcels': total_parcels,
            'cold_chain_count': cold_chain_count,
            'ambient_count': ambient_count,
            'courier_counts': courier_counts,
            'product_quantities': sorted_product_quantities,
        }
    }
    return render(request, 'operation/printable_parcels.html', context)


# 3. Edit Invoice (for payment date)
@login_required
def edit_courier_invoice(request, pk):
    """
    Updates an invoice's payment details, but only if changes are detected.
    """
    invoice = get_object_or_404(CourierInvoice, pk=pk)
    has_changed = False
    update_fields = []

    # Get new values from the form
    payment_date_str = request.POST.get('payment_date')
    payment_amount_str = request.POST.get('payment_amount')

    # --- Check and update Payment Date ---
    new_payment_date = None
    if payment_date_str:
        new_payment_date = datetime.datetime.strptime(payment_date_str, '%Y-%m-%d').date()

    if new_payment_date != invoice.payment_date:
        invoice.payment_date = new_payment_date
        invoice.payment_status = 'PAID' if new_payment_date else 'UNPAID'
        update_fields.extend(['payment_date', 'payment_status'])
        has_changed = True

    # --- Check and update Payment Amount ---
    new_payment_amount = None
    if payment_amount_str:
        try:
            # Clean up the string just in case, then convert to Decimal
            new_payment_amount = Decimal(payment_amount_str.strip())
        except (InvalidOperation, TypeError):
            messages.error(request, "Invalid payment amount format.")
            return redirect('operation:courier_invoice_list')

    # Compare new amount with the existing one (handle None cases)
    # The 'or Decimal('0.00')' handles cases where the amount is None or 0
    if new_payment_amount != invoice.payment_amount:
        invoice.payment_amount = new_payment_amount
        update_fields.append('payment_amount')
        has_changed = True

    # --- Save and show message only if there were changes ---
    if has_changed:
        # Recalculate discount if payment amount was part of the change
        if 'payment_amount' in update_fields:
            if invoice.invoice_amount is not None and invoice.payment_amount is not None:
                invoice.discount_amount = invoice.invoice_amount - invoice.payment_amount
            else:
                invoice.discount_amount = Decimal('0.00')
            update_fields.append('discount_amount')

        # Save all accumulated changes to the database
        invoice.save(update_fields=update_fields)
        messages.success(request, f"Invoice #{invoice.invoice_number} updated successfully.")

    return redirect('operation:courier_invoice_list')




# 4. Display Billed Parcels
@login_required
def billed_parcels_list(request):
    billed_items = CourierInvoiceItem.objects.select_related('courier_invoice', 'parcel').all()
    return render(request, 'operation/billed_parcels_list.html', {'billed_items': billed_items})


# 6. Cost Comparison Report
@login_required
def cost_comparison_report(request):
    parcels_with_costs = Parcel.objects.filter(
        estimated_cost__isnull=False,
        actual_shipping_cost__isnull=False
    ).annotate(
        cost_difference=F('actual_shipping_cost') - F('estimated_cost')
    )
    return render(request, 'operation/cost_comparison_report.html', {'parcels': parcels_with_costs})


# 7. Generate Client Invoice
@login_required
def generate_client_invoice(request, parcel_id):
    parcel = get_object_or_404(Parcel.objects.select_related('order__customer'), pk=parcel_id)
    # This view would gather all necessary details and render them into an invoice template.
    # For PDF generation, you could integrate libraries like WeasyPrint or ReportLab.
    return render(request, 'operation/client_invoice_template.html', {'parcel': parcel})


def courier_invoice_list(request):
    """
    Handles both displaying the list of invoices and uploading a new one.
    Now validates the uploaded file against the selected courier.
    """
    if request.method == 'POST':
        # Pass the current user to the form for permission handling
        form = CourierInvoiceForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            invoice = form.save(commit=False)
            invoice.uploaded_by = request.user

            # --- START: THE FIX (Enforce Warehouse on Save) ---
            # If the user is not a superuser, their warehouse is automatically assigned.
            if not request.user.is_superuser and request.user.warehouse:
                invoice.warehouse = request.user.warehouse
            # --- END: THE FIX ---

            invoice.save()

            # The orchestrator now returns the detected courier and parsing results
            detected_courier, created_count, errors, updated_count, successes = parse_invoice_file(invoice)

            # If the detected_courier is None, it means validation failed or the file is unsupported
            if not detected_courier:
                invoice.delete() # Delete the placeholder invoice
                for error in errors:
                    messages.error(request, error)
                return redirect('operation:courier_invoice_list')

            # --- Process success and error messages as before ---
            for msg in successes:
                messages.success(request, msg)

            has_critical_error = False
            if errors:
                for error in errors:
                    if "Info:" in error:
                        messages.info(request, error)
                    else:
                        messages.error(request, error)
                        if "Critical" in error:
                            has_critical_error = True

            if not errors or not has_critical_error:
                summary_parts = []
                if created_count > 0:
                    summary_parts.append(f"Created {created_count} new item(s)")
                if updated_count > 0:
                    summary_parts.append(f"Updated {updated_count} existing item(s)")

                if summary_parts:
                    item_summary = ", ".join(summary_parts) + "."
                    messages.success(request, f"Processing complete. {item_summary}")

            return redirect('operation:courier_invoice_list')
        else:
            for field, field_errors in form.errors.items():
                for error in field_errors:
                    messages.error(request, f"{field.capitalize()}: {error}")
            return redirect('operation:courier_invoice_list')

    filter_form = CourierInvoiceFilterForm(request.GET, user=request.user)

    # --- START: THE FIX (Filter Invoice List) ---
    # Start with a base queryset and apply filters based on user role
    invoices_qs = CourierInvoice.objects.select_related('courier_company', 'warehouse').prefetch_related('items').all().order_by('-invoice_date')
    if not request.user.is_superuser and request.user.warehouse:
        invoices_qs = invoices_qs.filter(warehouse=request.user.warehouse)
    elif not request.user.is_superuser:
        invoices_qs = invoices_qs.none() # If not a superuser and no warehouse, show nothing
    # --- END: THE FIX ---

    if filter_form.is_valid():
        # Your existing filter logic can now apply to the permission-filtered queryset
        courier_company = filter_form.cleaned_data.get('courier_company')
        payment_status = filter_form.cleaned_data.get('payment_status')
        query = filter_form.cleaned_data.get('q')
        warehouse = filter_form.cleaned_data.get('warehouse')

        if courier_company:
            invoices_qs = invoices_qs.filter(courier_company=courier_company)
        if payment_status:
            invoices_qs = invoices_qs.filter(payment_status=payment_status)
        if query:
            invoices_qs = invoices_qs.filter(invoice_number__icontains=query)
        if warehouse: # This filter will only be used by superusers
            invoices_qs = invoices_qs.filter(warehouse=warehouse)

    paginator = Paginator(invoices_qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    upload_form = CourierInvoiceForm(user=request.user)

    # Get all active couriers for the filter buttons
    all_couriers = CourierCompany.objects.filter(is_active=True).order_by('name')
    all_warehouses = Warehouse.objects.all().order_by('name')

    context = {
        'invoices': page_obj,
        'form': upload_form,
        'filter_form': filter_form,
        'all_couriers': all_couriers,
        'payment_status_choices': CourierInvoice.PAYMENT_STATUS_CHOICES,
        'warehouses': all_warehouses,
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'operation/partials/_courier_invoice_list_content.html', context)

    return render(request, 'operation/courier_invoice_list.html', context)

@login_required
def invoice_item_report(request):
    """
    Handles the Invoice Item Report page with consolidated and corrected filter logic.
    """
    is_ajax_filter = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    # --- 1. Get all filter and sort parameters from the request ---
    sort_by = request.GET.get('sort_by', '-courier_invoice__invoice_date')
    selected_courier = request.GET.get('courier')
    selected_dispute_status = request.GET.get('dispute_status', '')
    unlinked_only = request.GET.get('unlinked_only') == 'true'
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    query = request.GET.get('q')
    selected_warehouse = request.GET.get('warehouse')

    # --- 2. Start with the base queryset and annotations ---
    allowed_sort_fields = [
        'courier_invoice__invoice_date', '-courier_invoice__invoice_date', 'actual_cost',
        '-actual_cost', 'tracking_number', '-tracking_number', 'parcel__estimated_cost',
        '-parcel__estimated_cost', 'cost_gap', '-cost_gap'
    ]
    if sort_by not in allowed_sort_fields:
        sort_by = '-courier_invoice__invoice_date'

    items_query = CourierInvoiceItem.objects.select_related(
        'parcel__order__customer',
        'courier_invoice__courier_company',
        'courier_invoice__warehouse' # Important for filtering and display
    ).annotate(
        cost_gap=Case(
            When(parcel__estimated_cost__isnull=False, then=F('actual_cost') - F('parcel__estimated_cost')),
            default=Value(None),
            output_field=DecimalField()
        )
    )

    # --- 3. Apply all filters sequentially ---

    # Warehouse filter (applied first for permissions)
    if request.user.is_superuser and selected_warehouse:
        items_query = items_query.filter(courier_invoice__warehouse_id=selected_warehouse)
    elif not request.user.is_superuser and request.user.warehouse:
        items_query = items_query.filter(courier_invoice__warehouse=request.user.warehouse)

    # Dispute and Unlinked filters (mutually exclusive)
    if unlinked_only:
        selected_dispute_status = ''
        items_query = items_query.filter(parcel__isnull=True)
    elif selected_dispute_status:
        if selected_dispute_status == 'pending':
            items_query = items_query.filter(dispute_date__isnull=False, final_amount_date__isnull=True)
        elif selected_dispute_status == 'finalized':
            items_query = items_query.filter(final_amount_date__isnull=False)
        elif selected_dispute_status == 'not_disputed':
            items_query = items_query.filter(dispute_date__isnull=True)

    # Courier filter
    if selected_courier:
        items_query = items_query.filter(courier_invoice__courier_company_id=selected_courier)

    # Search query
    if query:
        items_query = items_query.filter(
            Q(tracking_number__icontains=query) |
            Q(courier_invoice__invoice_number__icontains=query)
        ).distinct()

    # Date range filter
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            items_query = items_query.filter(parcel__created_at__gte=start_date)
        except ValueError: pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            end_of_day = datetime.combine(end_date, datetime.time.max)
            items_query = items_query.filter(parcel__created_at__lte=end_of_day)
        except ValueError: pass

    # Apply final sorting
    items_query = items_query.order_by(sort_by)

    # --- 4. Pagination and Context ---
    paginator = Paginator(items_query, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    for item in page_obj:
        # ... (cost history processing remains the same) ...
        cumulative_costs = []
        running_total = Decimal('0.0')
        if isinstance(item.cost_history, list):
            sorted_history = sorted(item.cost_history, key=lambda x: x.get('date', ''))
            for charge in sorted_history:
                running_total += Decimal(str(charge.get('cost', 0)))
                cumulative_costs.append(running_total)
        item.display_costs = cumulative_costs


    context = {
        'items': page_obj,
        'couriers': CourierCompany.objects.filter(is_active=True).order_by('name'),
        'warehouses': Warehouse.objects.all().order_by('name'),
        'selected_courier': int(selected_courier) if selected_courier else None,
        'selected_warehouse': int(selected_warehouse) if selected_warehouse else None,
        'selected_dispute_status': selected_dispute_status,
        'unlinked_only': unlinked_only,
        'query': query,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'current_sort': sort_by,
        'total_items_count': paginator.count,
    }

    if is_ajax_filter:
        return render(request, 'operation/partials/_invoice_item_report_table_with_pagination.html', context)

    return render(request, 'operation/invoice_item_report.html', context)


@login_required
def load_more_invoice_items(request):
    """
    Handles AJAX requests for the 'Explore More' button.
    FIX: Added date range filtering to match the main view.
    """
    sort_by = request.GET.get('sort_by', '-courier_invoice__invoice_date')
    allowed_sort_fields = [
        'courier_invoice__invoice_date', '-courier_invoice__invoice_date', 'actual_cost',
        '-actual_cost', 'tracking_number', '-tracking_number', 'parcel__estimated_cost',
        '-parcel__estimated_cost', 'receiver_state', '-receiver_state', 'destination_name',
        '-destination_name', 'parcel__order__customer__customer_name',
        '-parcel__order__customer__customer_name', 'cost_gap', '-cost_gap'
    ]

    items_query = CourierInvoiceItem.objects.select_related(
        'parcel__order__customer', 'courier_invoice__courier_company'
    ).annotate(
        cost_gap=Case(
            When(parcel__estimated_cost__isnull=False, then=F('actual_cost') - F('parcel__estimated_cost')),
            default=Value(None)
        )
    ).order_by(sort_by)

    # --- START OF THE FIX ---
    selected_dispute_status = request.GET.get('dispute_status', '')
    unlinked_only = request.GET.get('unlinked_only') == 'true'

    if unlinked_only:
        items_query = items_query.filter(parcel__isnull=True)
    elif selected_dispute_status:
        if selected_dispute_status == 'pending':
            items_query = items_query.filter(dispute_date__isnull=False, final_amount_date__isnull=True)
        elif selected_dispute_status == 'finalized':
            items_query = items_query.filter(final_amount_date__isnull=False)
        elif selected_dispute_status == 'not_disputed':
            items_query = items_query.filter(dispute_date__isnull=True)
    # --- END OF THE FIX ---

    courier_id = request.GET.get('courier')
    if courier_id:
        items_query = items_query.filter(courier_invoice__courier_company_id=courier_id)

    query = request.GET.get('q')
    if query:
        items_query = items_query.filter(
            Q(tracking_number__icontains=query) | Q(courier_invoice__invoice_number__icontains=query)
        ).distinct()

    # FIX: Add date range filtering logic here as well
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            items_query = items_query.filter(parcel__created_at__gte=start_date)
        except ValueError:
            pass

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            end_of_day = datetime.combine(end_date, datetime.time.max)
            items_query = items_query.filter(parcel__created_at__lte=end_of_day)
        except ValueError:
            pass

    paginator = Paginator(items_query, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # ... (cost history processing is the same) ...
    for item in page_obj:
        cumulative_costs = []
        running_total = Decimal('0.0')
        if isinstance(item.cost_history, list):
            sorted_history = sorted(item.cost_history, key=lambda x: x.get('date', ''))
            for charge in sorted_history:
                running_total += Decimal(str(charge.get('cost', 0)))
                cumulative_costs.append(running_total)
        item.display_costs = cumulative_costs


    context = {'items': page_obj.object_list}
    html_rows = render_to_string('operation/partials/_invoice_item_report_rows_only.html', context)

    return JsonResponse({'html': html_rows, 'has_next': page_obj.has_next()})


@login_required
def view_parcel_items(request, item_id):
    """
    Handles the AJAX request to fetch and display the contents of a specific parcel
    linked to a courier invoice item.
    """
    # Fetch the specific invoice item, which links to the parcel
    invoice_item = get_object_or_404(
        CourierInvoiceItem.objects.select_related(
            'parcel__order__customer', # Pre-fetch related data for efficiency
            'parcel__packaging_type'  # <-- Added this line
        ).prefetch_related(
            # Pre-fetch all items within the parcel, including product and batch details
            'parcel__items_in_parcel__order_item__product',
            'parcel__items_in_parcel__shipped_from_batch'
        ),
        pk=item_id
    )

    parcel = invoice_item.parcel

    # Render just the modal content partial
    return render(request, 'operation/partials/_view_parcel_items_modal_content.html', {'parcel': parcel})


@login_required
def get_dispute_details(request, item_id):
    """
    Handles the AJAX request to fetch and render the dispute form for the modal.
    This version pre-processes the dispute history to ensure dates are formatted correctly.
    """
    invoice_item = get_object_or_404(CourierInvoiceItem, pk=item_id)
    form = DisputeForm(instance=invoice_item)
    new_update_form = DisputeUpdateForm(prefix='new_update')

    # --- MODIFICATION START ---
    # Pre-process the dispute history to ensure dates can be filtered.
    processed_history = []
    if invoice_item.dispute_history:
        for entry in invoice_item.dispute_history:
            try:
                # Attempt to parse the date string from the JSON field.
                # This handles the 'YYYY-MM-DD' format.
                parsed_date = datetime.datetime.strptime(entry['update_date'], '%Y-%m-%d').date()
                processed_history.append({
                    'update_date': parsed_date,
                    'remarks': entry['remarks']
                })
            except (ValueError, TypeError):
                # If parsing fails, keep the original entry to avoid crashing.
                # This can happen if old data is in a different format.
                processed_history.append(entry)
    # --- MODIFICATION END ---

    context = {
        'item': invoice_item,
        'form': form,
        'new_update_form': new_update_form,
        # Pass the pre-processed history to the template instead of the raw one.
        'processed_history': processed_history,
    }
    return render(request, 'operation/partials/_dispute_modal_content.html', context)


@login_required
@require_POST
def save_dispute_details(request, item_id):
    """
    Handles the POST request to save dispute details from the modal form.
    This version reliably adds the initialization record to the history.
    """
    invoice_item = get_object_or_404(CourierInvoiceItem, pk=item_id)

    # --- MODIFICATION START ---
    # The most reliable way to check if it's a new dispute is to see if the
    # date is already set on the database object BEFORE processing the form.
    is_new_dispute = invoice_item.dispute_date is None
    # --- MODIFICATION END ---

    form = DisputeForm(request.POST, instance=invoice_item)

    if not form.is_valid():
        errors = form.errors.as_json()
        logger.error(f"Main dispute form save failed for item {item_id}: {errors}")
        return JsonResponse({'success': False, 'message': 'Please correct the main form errors.', 'errors': errors}, status=400)

    # Save the main form data. This will set the dispute_date on new disputes.
    dispute_item = form.save()

    # Get the history list to modify it.
    history = dispute_item.dispute_history or []

    # If it was a new dispute, add the initialization record.
    if is_new_dispute:
        history.insert(0, {
            'update_date': dispute_item.dispute_date.isoformat(),
            'remarks': 'Dispute Initialized.'
        })

    # Check for and process the optional "Add New Update" form.
    update_date_from_post = request.POST.get('new_update-update_date')
    remarks_from_post = request.POST.get('new_update-remarks')

    if update_date_from_post or remarks_from_post:
        new_update_form = DisputeUpdateForm(request.POST, prefix='new_update')
        if new_update_form.is_valid():
            history.append({
                'update_date': new_update_form.cleaned_data['update_date'].isoformat(),
                'remarks': new_update_form.cleaned_data.get('remarks', '')
            })
        else:
            errors = new_update_form.errors.as_json()
            logger.error(f"New update form validation failed for item {item_id}: {errors}")
            return JsonResponse({'success': False, 'message': 'Please correct the errors in the New Update section.', 'errors': errors}, status=400)

    # Save the history list back to the object and update the database.
    dispute_item.dispute_history = sorted(history, key=lambda x: x['update_date'])
    dispute_item.save(update_fields=['dispute_history'])

    return JsonResponse({'success': True, 'message': 'Dispute details saved successfully.'})


@login_required
@require_POST
def cancel_dispute(request, item_id):
    """
    Allows a superuser to cancel an existing dispute, clearing all
    dispute-related fields for the invoice item.
    """
    # 1. Check for superuser permission
    if not request.user.is_superuser:
        return HttpResponseForbidden("You do not have permission to perform this action.")

    # 2. Get the invoice item
    invoice_item = get_object_or_404(CourierInvoiceItem, pk=item_id)

    # 3. Clear all dispute-related fields
    invoice_item.dispute_date = None
    invoice_item.final_amount_after_dispute = None
    invoice_item.final_amount_date = None
    invoice_item.dispute_history = []  # Reset the history to an empty list

    # 4. Save the changes to the database
    invoice_item.save()

    # 5. Return a success response
    return JsonResponse({'success': True, 'message': 'Dispute has been successfully canceled.'})

STATE_MAP = {
    'AL': 'ALABAMA', 'AK': 'ALASKA', 'AZ': 'ARIZONA', 'AR': 'ARKANSAS', 'CA': 'CALIFORNIA',
    'CO': 'COLORADO', 'CT': 'CONNECTICUT', 'DE': 'DELAWARE', 'FL': 'FLORIDA', 'GA': 'GEORGIA',
    'HI': 'HAWAII', 'ID': 'IDAHO', 'IL': 'ILLINOIS', 'IN': 'INDIANA', 'IA': 'IOWA',
    'KS': 'KANSAS', 'KY': 'KENTUCKY', 'LA': 'LOUISIANA', 'ME': 'MAINE', 'MD': 'MARYLAND',
    'MA': 'MASSACHUSETTS', 'MI': 'MICHIGAN', 'MN': 'MINNESOTA', 'MS': 'MISSISSIPPI',
    'MO': 'MISSOURI', 'MT': 'MONTANA', 'NE': 'NEBRASKA', 'NV': 'NEVADA', 'NH': 'NEW HAMPSHIRE',
    'NJ': 'NEW JERSEY', 'NM': 'NEW MEXICO', 'NY': 'NEW YORK', 'NC': 'NORTH CAROLINA',
    'ND': 'NORTH DAKOTA', 'OH': 'OHIO', 'OK': 'OKLAHOMA', 'OR': 'OREGON', 'PA': 'PENNSYLVANIA',
    'RI': 'RHODE ISLAND', 'SC': 'SOUTH CAROLINA', 'SD': 'SOUTH DAKOTA', 'TN': 'TENNESSEE',
    'TX': 'TEXAS', 'UT': 'UTAH', 'VT': 'VERMONT', 'VA': 'VIRGINIA', 'WA': 'WASHINGTON',
    'WV': 'WEST VIRGINIA', 'WI': 'WISCONSIN', 'WY': 'WYOMING', 'DC': 'DISTRICT OF COLUMBIA',
    'PR': 'PUERTO RICO', 'VI': 'VIRGIN ISLANDS'
}


# Add this import at the top of your views.py file if you don't have it
from dateutil.relativedelta import relativedelta

@login_required
def generate_report(request):
    """
    Handles the 'Generate Report' page.
    - GET: Displays both a single-month and a consolidated 6-month performance report.
    - POST: Generates and serves an Excel file based on the selected month.
    """
    if not request.user.is_superuser:
        # For non-superusers, automatically apply their assigned warehouse
        selected_warehouse_id = request.user.warehouse.id if request.user.warehouse else None
    else:
        # For superusers, get the selected warehouse from the GET parameters
        selected_warehouse_id = request.GET.get('warehouse')
    # --- POST: Handle Excel Report Generation (No changes needed here) ---
    if request.method == 'POST':
        month_str = request.POST.get('month')
        margin_str = request.POST.get('margin', '20')
        rate_str = request.POST.get('usd_exchange_rate', '7.25')
        try:
            year, month = map(int, month_str.split('-'))
            start_date = datetime.date(year, month, 1)
            end_date = start_date.replace(day=calendar.monthrange(year, month)[1])
            margin_multiplier = 1 + (Decimal(margin_str) / Decimal('100'))
            exchange_rate = Decimal(rate_str)
        except (ValueError, TypeError):
            return HttpResponse("Invalid data provided.", status=400)

        parcels = Parcel.objects.filter(created_at__date__range=[start_date, end_date]).select_related(
            'order__warehouse', 'packaging_type', 'billing_item').order_by('created_at')

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Parcels Export"
        headers = ["Order#", "Tracking Number", "Shipment Date", "Warehouse", "Type", "Shipment cost", "Dispute"]
        sheet.append(headers)

        for parcel in parcels:
            shipment_charges = Decimal('0.00')
            is_disputed = ""

            # --- START: THE FIX ---
            # The entire block accessing billing_item is now inside the try...except block.
            try:
                if parcel.billing_item.actual_cost is not None:
                    shipment_charges = (parcel.billing_item.actual_cost / exchange_rate) * margin_multiplier
                if parcel.billing_item.dispute_date:
                    is_disputed = "Yes"
            except Parcel.billing_item.RelatedObjectDoesNotExist:
                # This will be triggered if a parcel has no billing_item,
                # leaving shipment_charges and is_disputed at their default values.
                pass
            # --- END: THE FIX ---

            if parcel.packaging_type:
                packaging_info = parcel.packaging_type.get_environment_type_display()
            else:
                packaging_info = "N/A"

            row_data = [
                parcel.order.erp_order_id,
                parcel.tracking_number,
                parcel.created_at.strftime('%Y-%m-%d'),
                parcel.order.warehouse.name if parcel.order and parcel.order.warehouse else 'N/A',
                packaging_info,
                round(shipment_charges, 2),
                is_disputed
            ]
            sheet.append(row_data)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="parcels_report_{month_str}.xlsx"'
        workbook.save(response)
        return response

    # --- GET: Handle Dashboard Display ---
    today = datetime.date.today()
    selected_month_str = request.GET.get('month', today.strftime('%Y-%m'))
    try:
        selected_date = datetime.datetime.strptime(selected_month_str, '%Y-%m').date()
    except ValueError:
        selected_date = today

    months_for_selector = []
    current_month_for_selector = today.replace(day=1)
    for _ in range(12):
        months_for_selector.append(current_month_for_selector.strftime('%Y-%m'))
        last_month = current_month_for_selector - datetime.timedelta(days=1)
        current_month_for_selector = last_month.replace(day=1)

    all_couriers = CourierCompany.objects.filter(is_active=True).order_by('name')
    all_warehouses = Warehouse.objects.all().order_by('name') # Fetch warehouses for the filter

    # --- Reusable function to calculate stats for any given month ---
    def get_performance_stats_for_month(month_date, warehouse_id=None):
        parcels_in_month = Parcel.objects.filter(
            created_at__year=month_date.year,
            created_at__month=month_date.month
        )
        if warehouse_id:
            parcels_in_month = parcels_in_month.filter(order__warehouse_id=warehouse_id)

        monthly_stats = []
        for courier in all_couriers:
            courier_parcels = parcels_in_month.filter(courier_company=courier)

            simple_stats = courier_parcels.aggregate(
                total_parcels=Count('id'),
                total_successful=Count('id', filter=Q(status='DELIVERED')),
                total_billing_cost=Sum('billing_item__actual_cost'),
                total_cold_chain=Count('id', filter=Q(order__is_cold_chain=True)),
                total_ambient=Count('id', filter=Q(order__is_cold_chain=False)),
                avg_cold_chain_cost=Avg('billing_item__actual_cost', filter=Q(order__is_cold_chain=True)),
                avg_ambient_cost=Avg('billing_item__actual_cost', filter=Q(order__is_cold_chain=False))
            )

            duration_stats = courier_parcels.filter(
                status='DELIVERED',
                shipped_at__isnull=False,
                delivered_at__isnull=False
            ).aggregate(average_duration=Avg(F('delivered_at') - F('shipped_at')))

            total_parcels = simple_stats.get('total_parcels', 0)
            total_successful = simple_stats.get('total_successful', 0)
            success_rate = (total_successful / total_parcels * 100) if total_parcels > 0 else 0

            avg_days = 0.0
            avg_duration_timedelta = duration_stats.get('average_duration')
            if isinstance(avg_duration_timedelta, datetime.timedelta):
                avg_days = avg_duration_timedelta.total_seconds() / (24 * 3600)

            monthly_stats.append({
                'courier_name': courier.name,
                'total_parcels': total_parcels,
                'total_successful': total_successful,
                'success_rate': success_rate,
                'avg_delivery_days': avg_days,
                'total_billing_cost': simple_stats.get('total_billing_cost') or 0,
                'total_cold_chain': simple_stats.get('total_cold_chain', 0),
                'total_ambient': simple_stats.get('total_ambient', 0),
                'avg_cold_chain_cost': simple_stats.get('avg_cold_chain_cost') or 0,
                'avg_ambient_cost': simple_stats.get('avg_ambient_cost') or 0,
            })
        return monthly_stats

    # --- START OF THE FIX ---

    # 1. Calculate stats for the single selected month
    courier_performance_stats = get_performance_stats_for_month(selected_date, selected_warehouse_id)

    # 2. Calculate and consolidate stats for the past 6 months
    consolidated_courier_stats = {courier.name: [] for courier in all_couriers}

    for i in range(6):
        month_to_calculate = selected_date - relativedelta(months=i)
        monthly_stats = get_performance_stats_for_month(month_to_calculate)

        for stat in monthly_stats:
            courier_name = stat['courier_name']
            if courier_name in consolidated_courier_stats:
                # Append the entire dictionary of this month's stats to the courier's list
                stat['month'] = month_to_calculate.strftime('%B %Y')
                consolidated_courier_stats[courier_name].append(stat)

    # --- State statistics (now with warehouse filtering) ---
    stats_query_state = Parcel.objects.filter(
        created_at__year=selected_date.year, created_at__month=selected_date.month
    )
    if selected_warehouse_id:
        stats_query_state = stats_query_state.filter(order__warehouse_id=selected_warehouse_id)

    stats_query_state = stats_query_state.order_by().values(
        'billing_item__receiver_state', 'courier_company__name'
    ).annotate(total=Count('id')).order_by('billing_item__receiver_state', '-total')

    state_courier_counts = {}
    for item in stats_query_state:
        raw_state = item['billing_item__receiver_state']
        courier_name_from_query = item['courier_company__name']
        count = item['total']
        if not raw_state or not courier_name_from_query: continue
        state = STATE_MAP.get(raw_state.upper(), raw_state.upper())
        if state not in state_courier_counts:
            state_courier_counts[state] = {'total': 0, 'couriers': {}}
        state_courier_counts[state]['couriers'][courier_name_from_query] = count
        state_courier_counts[state]['total'] += count

    chart_labels = []
    chart_datasets = []
    colors = ['rgba(255, 99, 132, 0.2)', 'rgba(54, 162, 235, 0.2)', 'rgba(255, 206, 86, 0.2)', 'rgba(75, 192, 192, 0.2)', 'rgba(153, 102, 255, 0.2)', 'rgba(255, 159, 64, 0.2)']
    border_colors = ['rgba(255, 99, 132, 1)', 'rgba(54, 162, 235, 1)', 'rgba(255, 206, 86, 1)', 'rgba(75, 192, 192, 1)', 'rgba(153, 102, 255, 1)', 'rgba(255, 159, 64, 1)']

    # Get the month labels (in reverse chronological order)
    if consolidated_courier_stats:
        first_courier = next(iter(consolidated_courier_stats.values()), None)
        if first_courier:
            chart_labels = [stat['month'] for stat in first_courier]

    color_index = 0
    for courier_name, monthly_stats_list in consolidated_courier_stats.items():
        dataset = {
            'label': courier_name,
            'data': [stat.get('total_billing_cost', 0) or 0 for stat in monthly_stats_list], # handle None
            'backgroundColor': colors[color_index % len(colors)],
            'borderColor': border_colors[color_index % len(border_colors)],
            'borderWidth': 1
        }
        chart_datasets.append(dataset)
        color_index += 1

    chart_data = {
        'labels': chart_labels,
        'datasets': chart_datasets
    }

    context = {
        'warehouses': all_warehouses, # Add warehouses to the context
        'selected_warehouse': int(selected_warehouse_id) if selected_warehouse_id else None,
        'state_courier_counts': state_courier_counts,
        'courier_performance_stats': courier_performance_stats,
        'consolidated_courier_stats': consolidated_courier_stats,
        'months': months_for_selector,
        'selected_month': selected_month_str,
        'chart_data': json.dumps(chart_data, cls=DjangoJSONEncoder) # Use DjangoJSONEncoder for Decimals
    }

    return render(request, 'operation/generate_report.html', context)


# --- New View for the Dashboard ---
@login_required
def get_packaging_stock_dashboard(request, material_pk):
    """
    Handles fetching the data for the 6-month trend dashboard.
    """
    global_material = get_object_or_404(PackagingMaterial, pk=material_pk)
    today = timezone.now().date()

    # Base queryset for all transactions related to this material
    transactions_qs = PackagingStockTransaction.objects.filter(
        warehouse_packaging_material__packaging_material=global_material
    )

    # Filter by warehouse for non-superusers
    if not request.user.is_superuser and request.user.warehouse:
        transactions_qs = transactions_qs.filter(
            warehouse_packaging_material__warehouse=request.user.warehouse
        )

    # 6-Month Dashboard Calculation Logic
    monthly_dashboards = []
    for i in range(6):
        month_to_calculate = today - relativedelta(months=i)
        monthly_transactions = transactions_qs.filter(
            transaction_date__year=month_to_calculate.year,
            transaction_date__month=month_to_calculate.month
        )
        received = monthly_transactions.filter(transaction_type='IN').aggregate(total=Sum('quantity'))['total'] or 0
        usage = abs(monthly_transactions.filter(transaction_type='OUT').aggregate(total=Sum('quantity'))['total'] or 0)
        monthly_dashboards.append({
            'month_name': month_to_calculate.strftime("%B %Y"),
            'total_received': received, 'total_usage': usage
        })

    context = {
        'global_material': global_material,
        'monthly_dashboards': monthly_dashboards,
    }
    return render(request, 'operation/partials/_packaging_stock_dashboard.html', context)


@login_required
def get_packaging_receipt_log(request, material_pk):
    """
    Handles fetching the data for the month-filterable stock receipt log.
    """
    global_material = get_object_or_404(PackagingMaterial, pk=material_pk)
    today = timezone.now().date()
    selected_month_str = request.GET.get('month', today.strftime('%Y-%m'))

    try:
        selected_month_date = datetime.datetime.strptime(selected_month_str, '%Y-%m').date()
    except (ValueError, TypeError):
        selected_month_date = today

    # Base queryset for log
    log_qs = PackagingStockTransaction.objects.filter(
        warehouse_packaging_material__packaging_material=global_material,
        transaction_type=PackagingStockTransaction.TransactionTypes.STOCK_IN,
        transaction_date__year=selected_month_date.year,
        transaction_date__month=selected_month_date.month
    )

    # Filter by warehouse for non-superusers
    if not request.user.is_superuser and request.user.warehouse:
        log_qs = log_qs.filter(warehouse_packaging_material__warehouse=request.user.warehouse)

    context = {
        'transactions_for_log': log_qs.select_related('recorded_by', 'warehouse_packaging_material__warehouse').order_by('-transaction_date'),
        'is_superuser_view': request.user.is_superuser
    }
    return render(request, 'operation/partials/_packaging_receipt_log.html', context)
