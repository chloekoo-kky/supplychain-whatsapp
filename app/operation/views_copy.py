# app/operation/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.urls import reverse

from django.db import transaction, IntegrityError
from django.db.models import Q, Count, Prefetch, F, Value
from django.db.models.functions import Coalesce
from django.utils.dateparse import parse_date
from datetime import datetime
from django.http import JsonResponse, Http404, HttpResponse
from django.template.loader import render_to_string
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

import logging
import traceback


import openpyxl
import xlrd

from .forms import ExcelImportForm, ParcelForm, ParcelItemFormSet, InitialParcelItemFormSet, RemoveOrderItemForm, RemoveOrderItemFormSet
from .models import Order, OrderItem, Parcel, ParcelItem # Ensure Parcel is imported
from inventory.models import Product, InventoryBatchItem, StockTransaction
from warehouse.models import Warehouse, WarehouseProduct

logger = logging.getLogger(__name__)
DEFAULT_CUSTOMER_ORDERS_TAB = "customer_orders"
DEFAULT_PARCELS_TAB = "parcels_details"


@login_required
@login_required
def order_list_view(request):
    logger.debug(f"[OrderListView] Request GET params: {request.GET}")
    user = request.user

    active_tab = request.GET.get('tab', DEFAULT_CUSTOMER_ORDERS_TAB)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    fetch_dynamic_content_only = request.GET.get('fetch_dynamic_content_only') == 'true'

    all_warehouses_qs = Warehouse.objects.all().order_by('name')
    status_choices = Order.STATUS_CHOICES
    import_form = ExcelImportForm()

    context = {
        'status_choices': status_choices,
        'import_form': import_form,
        'active_tab': active_tab,
        'DEFAULT_CUSTOMER_ORDERS_TAB': DEFAULT_CUSTOMER_ORDERS_TAB,
        'DEFAULT_PARCELS_TAB': DEFAULT_PARCELS_TAB,
        'request': request,
    }

    if active_tab == DEFAULT_CUSTOMER_ORDERS_TAB:
        selected_warehouse_id = request.GET.get('warehouse')
        selected_status = request.GET.get('status')
        query = request.GET.get('q', '').strip()
        page_number = request.GET.get('page', 1)
        logger.debug(f"[CustomerOrdersTab] Filters: warehouse='{selected_warehouse_id}', status='{selected_status}', q='{query}', page='{page_number}'")

        orders_qs = Order.objects.select_related(
            'warehouse', 'imported_by'
        ).prefetch_related(
            Prefetch('items', queryset=OrderItem.objects.select_related('product', 'suggested_batch_item', 'warehouse_product').order_by('product__name')),
            Prefetch('parcels', queryset=Parcel.objects.all().order_by('-created_at'))
        ).all()

        warehouses_for_co_filters = all_warehouses_qs
        if not user.is_superuser:
            if user.warehouse:
                warehouses_for_co_filters = Warehouse.objects.filter(pk=user.warehouse.pk)
                orders_qs = orders_qs.filter(warehouse=user.warehouse)
                selected_warehouse_id = str(user.warehouse.pk)
            else:
                warehouses_for_co_filters = Warehouse.objects.none()
                orders_qs = orders_qs.none()
        context['warehouses'] = warehouses_for_co_filters

        if selected_warehouse_id and user.is_superuser:
            orders_qs = orders_qs.filter(warehouse_id=selected_warehouse_id)

        if selected_status:
            orders_qs = orders_qs.filter(status=selected_status)
        if query:
            orders_qs = orders_qs.filter(
                Q(erp_order_id__icontains=query) | Q(customer_name__icontains=query) |
                Q(order_display_code__icontains=query) | Q(items__product__sku__icontains=query) |
                Q(items__product__name__icontains=query) | Q(parcels__parcel_code_system__icontains=query) |
                Q(parcels__tracking_number__icontains=query)
            ).distinct()

        orders_qs = orders_qs.order_by('-order_date', '-imported_at')
        paginator = Paginator(orders_qs, 10)
        try:
            orders_page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            orders_page_obj = paginator.page(1)
        except EmptyPage:
            orders_page_obj = paginator.page(paginator.num_pages if paginator.num_pages > 0 else 1)

        for order_instance in orders_page_obj.object_list:
            for item_instance in order_instance.items.all():
                item_instance.quantity_notionally_removed = order_instance.get_total_removed_quantity_for_item(item_instance.id)

        context.update({
            'orders_page_obj': orders_page_obj,
            'total_orders_count': paginator.count,
            'selected_warehouse': selected_warehouse_id,
            'selected_status': selected_status,
            'query': query,
            'page_title': "Customer Orders",
        })

        if is_ajax:
            if fetch_dynamic_content_only:
                template_to_render = 'operation/partials/_customer_orders_table_with_pagination.html'
            else:
                template_to_render = 'operation/partials/customer_orders_table.html'
            response = render(request, template_to_render, context)
            response['X-Total-Orders-Count'] = paginator.count
            if orders_page_obj.has_next():
                response['HX-Trigger-After-Swap'] = 'loadMoreCustomerOrdersAvailable'
            else:
                response['HX-Trigger-After-Swap'] = 'loadMoreCustomerOrdersUnavailable'
            return response

    elif active_tab == DEFAULT_PARCELS_TAB:
        parcels_qs = Parcel.objects.select_related(
            'order__warehouse', 'order__imported_by', 'created_by'
        ).prefetch_related(
            Prefetch('items_in_parcel', queryset=ParcelItem.objects.select_related(
                'order_item__product', 'shipped_from_batch'
            ).order_by('order_item__product__name'))
        ).all()

        # Get distinct courier names for the filter
        # Coalesce empty strings to NULL first, then filter out NULLs, then get distinct.
        # This handles cases where courier_name might be an empty string.
        courier_companies = Parcel.objects.annotate(
            actual_courier_name=Coalesce('courier_name', Value(None))
        ).filter(actual_courier_name__isnull=False).values_list('actual_courier_name', flat=True).distinct().order_by('actual_courier_name')
        context['courier_companies'] = courier_companies

        warehouses_for_parcel_filters_ui = all_warehouses_qs
        actual_selected_warehouse_id_for_query_and_ui = request.GET.get('parcel_warehouse')
        selected_parcel_courier = request.GET.get('parcel_courier') # New filter

        if not user.is_superuser:
            if user.warehouse:
                warehouses_for_parcel_filters_ui = Warehouse.objects.filter(pk=user.warehouse.pk)
                parcels_qs = parcels_qs.filter(order__warehouse=user.warehouse)
                actual_selected_warehouse_id_for_query_and_ui = str(user.warehouse.pk)
            else:
                warehouses_for_parcel_filters_ui = Warehouse.objects.none()
                parcels_qs = parcels_qs.none()
                actual_selected_warehouse_id_for_query_and_ui = None

        context['warehouses'] = warehouses_for_parcel_filters_ui

        if user.is_superuser and actual_selected_warehouse_id_for_query_and_ui:
            parcels_qs = parcels_qs.filter(order__warehouse_id=actual_selected_warehouse_id_for_query_and_ui)

        if selected_parcel_courier: # Apply courier filter
            parcels_qs = parcels_qs.filter(courier_name=selected_parcel_courier)

        parcel_query_param = request.GET.get('parcel_q', '').strip()
        page_number = request.GET.get('page', 1)
        logger.debug(f"[ParcelTab] Filters: parcel_warehouse='{actual_selected_warehouse_id_for_query_and_ui}', parcel_courier='{selected_parcel_courier}', parcel_q='{parcel_query_param}', page='{page_number}'")

        if parcel_query_param:
            parcels_qs = parcels_qs.filter(
                Q(parcel_code_system__icontains=parcel_query_param) |
                Q(tracking_number__icontains=parcel_query_param) |
                Q(order__erp_order_id__icontains=parcel_query_param) |
                Q(items_in_parcel__order_item__product__name__icontains=parcel_query_param) |
                Q(items_in_parcel__order_item__product__sku__icontains=parcel_query_param) |
                Q(items_in_parcel__shipped_from_batch__batch_number__icontains=parcel_query_param) |
                Q(items_in_parcel__shipped_from_batch__location_label__icontains=parcel_query_param)
            ).distinct()

        parcels_qs = parcels_qs.order_by('-created_at')
        parcel_paginator = Paginator(parcels_qs, 20)
        try:
            parcels_page = parcel_paginator.page(page_number)
        except PageNotAnInteger:
            parcels_page = parcel_paginator.page(1)
        except EmptyPage:
            parcels_page = parcel_paginator.page(parcel_paginator.num_pages if parcel_paginator.num_pages > 0 else 1)

        context.update({
            'parcels': parcels_page,
            'total_parcels_count': parcel_paginator.count,
            'selected_parcel_warehouse': actual_selected_warehouse_id_for_query_and_ui,
            'selected_parcel_courier': selected_parcel_courier, # Pass to template
            'parcel_query': parcel_query_param,
            'page_title': "Parcel Details",
        })
        if is_ajax:
             template_to_render = 'operation/partials/parcels_table.html'
             response = render(request, template_to_render, context)
             response['X-Total-Parcels-Count'] = parcel_paginator.count
             return response

    return render(request, 'operation/order_management_base.html', context)


@login_required
def load_more_customer_orders(request):
    logger.debug(f"[LoadMoreCustomerOrders] Request GET params: {request.GET}")
    user = request.user

    orders_qs = Order.objects.select_related(
        'warehouse',
        'imported_by'
    ).prefetch_related(
        Prefetch('items', queryset=OrderItem.objects.select_related('product', 'suggested_batch_item', 'warehouse_product').order_by('product__name')),
        Prefetch('parcels', queryset=Parcel.objects.all().order_by('-created_at'))
    ).all()

    if not user.is_superuser and user.warehouse:
        orders_qs = orders_qs.filter(warehouse=user.warehouse)
    elif not user.is_superuser:
        orders_qs = orders_qs.none()

    selected_warehouse_id = request.GET.get('warehouse')
    selected_status = request.GET.get('status')
    query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page', 1)

    if user.is_superuser and selected_warehouse_id:
        orders_qs = orders_qs.filter(warehouse_id=selected_warehouse_id)

    if selected_status:
        orders_qs = orders_qs.filter(status=selected_status)
    if query:
        orders_qs = orders_qs.filter(
            Q(erp_order_id__icontains=query) |
            Q(customer_name__icontains=query) |
            Q(items__product__sku__icontains=query)
        ).distinct()

    orders_qs = orders_qs.order_by('-order_date', '-imported_at')

    paginator = Paginator(orders_qs, 10)
    try:
        orders_page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        logger.warning(f"PageNotAnInteger for page '{page_number}' in load_more_customer_orders. Defaulting to page 1.")
        orders_page_obj = paginator.page(1)
        if not orders_page_obj.object_list:
            return HttpResponse("")
    except EmptyPage:
        logger.info(f"EmptyPage for page '{page_number}' in load_more_customer_orders. No more items.")
        return HttpResponse("")

    for order_instance in orders_page_obj.object_list:
        for item_instance in order_instance.items.all():
            item_instance.quantity_notionally_removed = order_instance.get_total_removed_quantity_for_item(item_instance.id)

    context = {
        'orders': orders_page_obj.object_list,
        'request': request,
    }

    html_rows = render_to_string('operation/partials/_customer_orders_list_items_only.html', context)

    response = HttpResponse(html_rows)
    if orders_page_obj.has_next():
        response['HX-Trigger'] = 'loadMoreCustomerOrdersAvailable'
    else:
        response['HX-Trigger'] = 'loadMoreCustomerOrdersUnavailable'
    return response

# ... (rest of your views: import_orders_from_excel, get_order_items_for_packing, etc.) ...
# Ensure they are not accidentally truncated or modified if not intended.
@login_required
def import_orders_from_excel(request):
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            file_name_lower = excel_file.name.lower()

            try:
                workbook_data = None
                sheet = None
                is_xlsx = False

                if file_name_lower.endswith('.xlsx'):
                    workbook_data = openpyxl.load_workbook(excel_file, data_only=True)
                    sheet = workbook_data.active
                    is_xlsx = True
                elif file_name_lower.endswith('.xls'):
                    file_contents = excel_file.read()
                    workbook_data = xlrd.open_workbook(file_contents=file_contents)
                    sheet = workbook_data.sheet_by_index(0)
                    is_xlsx = False
                else:
                    messages.error(request, "Unsupported file format. Please upload .xlsx or .xls files.")
                    return redirect('operation:order_list')

                headers = []
                data_rows_iterator = None

                if is_xlsx:
                    if sheet.max_row > 0:
                        headers = [cell.value for cell in sheet[1]]
                        data_rows_iterator = sheet.iter_rows(min_row=2, values_only=True)
                    else:
                        headers = []
                else:
                    if sheet.nrows > 0:
                        headers = [sheet.cell_value(0, col_idx) for col_idx in range(sheet.ncols)]
                        def xlrd_rows_iterator(sheet_obj, book_datemode):
                            for r_idx in range(1, sheet_obj.nrows):
                                row_values_xls = []
                                for c_idx in range(sheet_obj.ncols):
                                    cell_type = sheet_obj.cell_type(r_idx, c_idx)
                                    cell_value = sheet_obj.cell_value(r_idx, c_idx)
                                    if cell_type == xlrd.XL_CELL_DATE:
                                        date_tuple = xlrd.xldate_as_datetime(cell_value, book_datemode)
                                        row_values_xls.append(date_tuple)
                                    elif cell_type == xlrd.XL_CELL_NUMBER and cell_value == int(cell_value):
                                        row_values_xls.append(int(cell_value))
                                    else:
                                        row_values_xls.append(cell_value)
                                yield row_values_xls
                        data_rows_iterator = xlrd_rows_iterator(sheet, workbook_data.datemode)
                    else:
                        headers = []

                if not headers:
                     messages.error(request, "The Excel file is empty or has no header row.")
                     return redirect('operation:order_list')

                header_mapping_config = {
                    'erp_order_id': 'order id',
                    'order_date': 'order date',
                    'warehouse_name': 'warehouse name',
                    'customer_name': 'address name',
                    'company_name': 'company',
                    'address_line1': 'address',
                    'country': 'country',
                    'city': 'city',
                    'state': 'state',
                    'zip_code': 'zip',
                    'phone': 'phone',
                    'vat_number': 'vat number',
                    'product_name_from_excel': 'product name',
                    'quantity_ordered': 'product quantity',
                    'is_cold': 'iscold',
                    'title_notes': 'title',
                    'shipping_notes': 'comment',
                }

                normalized_actual_headers = {str(h).strip().lower(): str(h).strip() for h in headers if h is not None}
                header_map_for_indexing = {}
                missing_headers_from_config = []
                critical_internal_keys = ['erp_order_id', 'order_date', 'warehouse_name', 'product_name_from_excel', 'quantity_ordered']

                for internal_key, excel_header_normalized_target in header_mapping_config.items():
                    found_actual_header = None
                    for actual_header_normalized, actual_header_original_case in normalized_actual_headers.items():
                        if actual_header_normalized == excel_header_normalized_target.lower():
                            found_actual_header = actual_header_original_case
                            break
                    if found_actual_header:
                        header_map_for_indexing[internal_key] = found_actual_header
                    elif internal_key in critical_internal_keys:
                         missing_headers_from_config.append(f"'{excel_header_normalized_target}' (expected for '{internal_key}')")


                if missing_headers_from_config:
                    messages.error(request, f"Required headers not found in Excel: {', '.join(missing_headers_from_config)}. Available headers found: {', '.join(filter(None,headers))}")
                    return redirect('operation:order_list')

                final_header_to_index_map = {}
                for internal_key, mapped_excel_header_original_case in header_map_for_indexing.items():
                    try:
                        idx = headers.index(mapped_excel_header_original_case)
                        final_header_to_index_map[internal_key] = idx
                    except ValueError:
                        messages.error(request, f"Configuration error: Mapped Excel header '{mapped_excel_header_original_case}' (for internal key '{internal_key}') not found in the original headers list. This is an internal logic error.")
                        return redirect('operation:order_list')

                orders_data = {}
                last_valid_erp_order_id = None

                for row_idx, row_tuple in enumerate(data_rows_iterator, start=2):
                    if not any(str(cell_val).strip() for cell_val in row_tuple if cell_val is not None):
                        logger.debug(f"Row {row_idx}: Skipped. Entirely empty or whitespace.")
                        continue
                    def get_current_row_value(internal_key_lambda, default=None):
                        idx = final_header_to_index_map.get(internal_key_lambda)
                        if idx is not None and idx < len(row_tuple) and row_tuple[idx] is not None:
                            val_lambda = row_tuple[idx]
                            if isinstance(val_lambda, (datetime, int, float)):
                                return val_lambda
                            val_str = str(val_lambda).strip()
                            return val_str if val_str != "" else default
                        return default

                    current_row_erp_order_id = get_current_row_value('erp_order_id')
                    erp_order_id_to_use = current_row_erp_order_id if current_row_erp_order_id else last_valid_erp_order_id

                    if not erp_order_id_to_use:
                        messages.warning(request, f"Row {row_idx}: Skipped. Missing Order ID and no previous order context.")
                        logger.warning(f"Row {row_idx}: Skipped. Missing Order ID.")
                        continue
                    if current_row_erp_order_id:
                        last_valid_erp_order_id = current_row_erp_order_id
                    erp_order_id_to_use = str(erp_order_id_to_use)

                    product_name_excel = get_current_row_value('product_name_from_excel')
                    quantity_ordered_str = get_current_row_value('quantity_ordered')

                    if not all([product_name_excel, quantity_ordered_str]):
                        messages.warning(request, f"Row {row_idx} (Order ID: {erp_order_id_to_use}): Missing critical item data (Product Name or Quantity). Skipping item.")
                        logger.warning(f"Row {row_idx} (Order ID: {erp_order_id_to_use}): Missing Product Name ('{product_name_excel}') or Qty ('{quantity_ordered_str}'). Skipping item.")
                        continue

                    order_date_for_this_entry = None
                    if erp_order_id_to_use not in orders_data:
                        order_date_val = get_current_row_value('order_date')
                        warehouse_name = get_current_row_value('warehouse_name')
                        customer_name = get_current_row_value('customer_name')
                        if order_date_val:
                            if isinstance(order_date_val, datetime):
                                order_date_for_this_entry = order_date_val.date()
                            elif isinstance(order_date_val, str):
                                order_date_str_cleaned = order_date_val.split(' ')[0]
                                order_date_for_this_entry = parse_date(order_date_str_cleaned)
                                if not order_date_for_this_entry:
                                    for fmt in ('%B %d, %Y', '%b %d %Y', '%d/%m/%Y', '%m/%d/%Y', '%Y%m%d'):
                                        try:
                                            order_date_for_this_entry = datetime.strptime(order_date_val, fmt).date()
                                            if order_date_for_this_entry: break
                                        except ValueError:
                                            continue
                            elif isinstance(order_date_val, (float, int)) and hasattr(workbook_data, 'datemode') and not is_xlsx:
                                try:
                                    order_date_for_this_entry = xlrd.xldate_as_datetime(order_date_val, workbook_data.datemode).date()
                                except: pass

                            if not order_date_for_this_entry and order_date_val:
                                logger.warning(f"Row {row_idx} (Order ID: {erp_order_id_to_use}): Unparseable order_date '{order_date_val}'.")

                        if not all([order_date_for_this_entry, warehouse_name, customer_name]):
                            error_parts = []
                            if not order_date_for_this_entry: error_parts.append("Order Date (missing or invalid format)")
                            if not warehouse_name: error_parts.append("Warehouse Name")
                            if not customer_name: error_parts.append("Customer Name (Address Name)")
                            msg = f"Row {row_idx} (Order ID: {erp_order_id_to_use}): Missing essential order details for first line: {', '.join(error_parts)}. Skipping order."
                            messages.warning(request, msg)
                            logger.warning(msg)
                            last_valid_erp_order_id = None
                            continue

                        orders_data[erp_order_id_to_use] = {
                            'order_details': {
                                'erp_order_id': str(erp_order_id_to_use),
                                'order_date': order_date_for_this_entry,
                                'warehouse_name': warehouse_name,
                                'customer_name': customer_name,
                                'company_name': get_current_row_value('company_name'),
                                'recipient_address_line1': get_current_row_value('address_line1'),
                                'recipient_address_city': get_current_row_value('city'),
                                'recipient_address_state': get_current_row_value('state'),
                                'recipient_address_zip': get_current_row_value('zip_code'),
                                'recipient_address_country': get_current_row_value('country'),
                                'recipient_phone': get_current_row_value('phone'),
                                'vat_number': get_current_row_value('vat_number'),
                                'is_cold_chain': False,
                                'title_notes': get_current_row_value('title_notes'),
                                'shipping_notes': get_current_row_value('shipping_notes'),
                            },
                            'items': []
                        }
                    try:
                        quantity_ordered = int(float(str(quantity_ordered_str)))
                        if quantity_ordered <= 0:
                            raise ValueError("Quantity must be positive.")
                    except (ValueError, TypeError):
                        msg = f"Row {row_idx} (Order ID: {erp_order_id_to_use}): Invalid quantity '{quantity_ordered_str}' for item '{product_name_excel}'. Skipping item."
                        messages.warning(request, msg)
                        logger.warning(msg)
                        continue

                    is_cold_str = get_current_row_value('is_cold')
                    is_cold = str(is_cold_str).strip().lower() == 'yes' if is_cold_str else False

                    if is_cold:
                        orders_data[erp_order_id_to_use]['order_details']['is_cold_chain'] = True

                    orders_data[erp_order_id_to_use]['items'].append({
                        'product_name_from_excel': str(product_name_excel),
                        'quantity_ordered': quantity_ordered,
                        'is_cold_item': is_cold,
                    })

                with transaction.atomic():
                    created_orders_count = 0
                    updated_orders_count = 0
                    created_items_count = 0
                    skipped_orders_db = 0
                    skipped_items_db = 0

                    for erp_id_key, data_dict in orders_data.items():
                        order_details_map = data_dict['order_details']
                        try:
                            warehouse = Warehouse.objects.get(name__iexact=order_details_map['warehouse_name'])
                        except Warehouse.DoesNotExist:
                            messages.error(request, f"Order ID {erp_id_key}: Warehouse '{order_details_map['warehouse_name']}' not found during DB operations. Skipping order.")
                            logger.error(f"Order ID {erp_id_key}: Warehouse '{order_details_map['warehouse_name']}' not found.")
                            skipped_orders_db += 1
                            continue

                        order_field_defaults = {
                            k: v for k, v in order_details_map.items()
                            if k not in ['erp_order_id', 'warehouse_name'] and hasattr(Order, k)
                        }

                        order_field_defaults['warehouse'] = warehouse
                        order_field_defaults['status'] = 'NEW_ORDER'
                        order_field_defaults['imported_by'] = request.user if request.user.is_authenticated else None

                        current_order_erp_id_str = str(order_details_map['erp_order_id'])

                        order, created = Order.objects.update_or_create(
                            erp_order_id=current_order_erp_id_str,
                            defaults=order_field_defaults
                        )
                        order.erp_order_id = current_order_erp_id_str

                        if created:
                            created_orders_count += 1
                        else:
                            order.items.all().delete()
                            updated_orders_count +=1

                        current_order_items_processed_db = 0
                        for item_data in data_dict['items']:
                            product_identifier_from_excel = item_data['product_name_from_excel']
                            product = None
                            try:
                                product = Product.objects.get(sku__iexact=product_identifier_from_excel)
                            except Product.DoesNotExist:
                                try:
                                    product = Product.objects.get(name__iexact=product_identifier_from_excel)
                                except Product.DoesNotExist:
                                    messages.error(request, f"Order ID {erp_id_key}, Item Identifier '{product_identifier_from_excel}': Product not found by SKU or Name. Skipping item.")
                                    logger.error(f"Order ID {erp_id_key}, Item '{product_identifier_from_excel}': Product not found by SKU/Name.")
                                    skipped_items_db +=1
                                    continue
                                except Product.MultipleObjectsReturned:
                                    messages.error(request, f"Order ID {erp_id_key}, Item Name '{product_identifier_from_excel}': Multiple products found by this name. Use unique SKU or ensure unique names. Skipping item.")
                                    logger.error(f"Order ID {erp_id_key}, Item '{product_identifier_from_excel}': Multiple products by name.")
                                    skipped_items_db +=1
                                    continue
                            except Product.MultipleObjectsReturned:
                                messages.error(request, f"Order ID {erp_id_key}, SKU '{product_identifier_from_excel}': Multiple products found with this SKU. SKUs must be unique. Skipping item.")
                                logger.error(f"Order ID {erp_id_key}, SKU '{product_identifier_from_excel}': Multiple products by SKU.")
                                skipped_items_db +=1
                                continue

                            warehouse_product_instance = WarehouseProduct.objects.filter(product=product, warehouse=warehouse).first()
                            if not warehouse_product_instance:
                                messages.warning(request, f"Order ID {erp_id_key}, Item '{product.sku}': WarehouseProduct link not found for warehouse '{warehouse.name}'. Item added without specific stock link.")
                                logger.warning(f"Order ID {erp_id_key}, Item '{product.sku}': WarehouseProduct link not found for WH '{warehouse.name}'.")

                            oi_defaults = {
                                'quantity_ordered': item_data['quantity_ordered'],
                                'erp_product_name': product_identifier_from_excel,
                                'is_cold_item': item_data.get('is_cold_item', False),
                                'status': 'PENDING_PROCESSING',
                                'warehouse_product': warehouse_product_instance
                            }
                            OrderItem.objects.create(order=order, product=product, **oi_defaults)
                            current_order_items_processed_db +=1

                        created_items_count += current_order_items_processed_db
                        order.save()


                final_message_parts = []
                if created_orders_count > 0: final_message_parts.append(f"Orders Created: {created_orders_count}")
                if updated_orders_count > 0: final_message_parts.append(f"Orders Updated: {updated_orders_count}")
                if created_items_count > 0: final_message_parts.append(f"Order Items Processed: {created_items_count}")

                if not final_message_parts and (skipped_orders_db == 0 and skipped_items_db == 0):
                    final_message_parts.append("No new orders or items to import based on ERP IDs. Existing orders might have been updated if their content changed.")

                if skipped_orders_db > 0: final_message_parts.append(f"Orders Skipped (DB): {skipped_orders_db}")
                if skipped_items_db > 0: final_message_parts.append(f"Items Skipped (DB): {skipped_items_db}")

                if final_message_parts:
                    messages.success(request, "Import process complete. " + ", ".join(final_message_parts) + ".")
                else:
                    messages.info(request, "No data found in the Excel file to process orders.")


            except ValueError as ve:
                messages.error(request, f"Error in file structure or critical content: {str(ve)}")
                logger.error(f"Import ValueError: {str(ve)}", exc_info=True)
            except xlrd.XLRDError as xe:
                messages.error(request, f"Error reading .xls file. It might be corrupted or an incompatible version: {str(xe)}")
                logger.error(f"Import XLRDError: {str(xe)}", exc_info=True)
            except Exception as e:
                messages.error(request, f"An unexpected error occurred during the import process: {str(e)}")
                logger.error(f"Import Exception: {str(e)}", exc_info=True)

            return redirect('operation:order_list')
        else:
            for field, errors_list in form.errors.items():
                for error in errors_list:
                    messages.error(request, f"Error in field '{form.fields[field].label if field != '__all__' else 'Form'}': {error}")
    return redirect('operation:order_list')

@login_required
def get_order_items_for_packing(request, order_pk):
    try:
        from inventory.services import get_suggested_batch_for_order_item
    except ImportError as e:
        logger.error(f"Failed to import 'get_suggested_batch_for_order_item': {e}\n{traceback.format_exc()}")
        return JsonResponse({'success': False, 'message': f"Critical server error. ({e})"}, status=500)

    try:
        order = get_object_or_404(Order.objects.prefetch_related(
            Prefetch('items', queryset=OrderItem.objects.filter(
                Q(status='PENDING_PROCESSING') | (Q(status='PACKED') & Q(quantity_packed__lt=F('quantity_ordered')))
            ).select_related('product', 'suggested_batch_item', 'warehouse_product').order_by('product__name'))
        ), pk=order_pk)

        logger.debug(f"[Pack Modal Data] For Order {order.pk}: items_removed_log is: {order.items_removed_log}")

        if not request.user.is_superuser and (not request.user.warehouse or order.warehouse != request.user.warehouse):
            return JsonResponse({'success': False, 'message': 'Permission denied for this order.'}, status=403)

        initial_form_data = []
        for item in order.items.all():
            total_removed_for_this_item = order.get_total_removed_quantity_for_item(item.id)
            quantity_remaining_to_pack_for_this_item = item.quantity_ordered - item.quantity_packed - total_removed_for_this_item

            logger.debug(f"  [Pack Modal Data] Item PK {item.pk} (SKU: {item.product.sku if item.product else 'N/A'}): "
                         f"Ordered={item.quantity_ordered}, Packed={item.quantity_packed}, "
                         f"NotionallyRemoved={total_removed_for_this_item}, "
                         f"ActualBalanceToPack={quantity_remaining_to_pack_for_this_item}")

            if quantity_remaining_to_pack_for_this_item > 0:
                best_suggested_batch = get_suggested_batch_for_order_item(item, quantity_remaining_to_pack_for_this_item)
                initial_item_data = {
                    'order_item_id': item.pk,
                    'product_name': item.product.name if item.product else item.erp_product_name,
                    'sku': item.product.sku if item.product else "N/A",
                    'quantity_to_pack': quantity_remaining_to_pack_for_this_item,
                    'selected_batch_item_id': best_suggested_batch.pk if best_suggested_batch else None,
                }
                initial_form_data.append(initial_item_data)
            else:
                logger.debug(f"  [Pack Modal Data] Item PK {item.pk} has no remaining quantity to pack after considering removals. Skipping.")


        formset_html_content = ""
        message_for_modal = ""

        if not initial_form_data:
            relevant_order_items_exist = OrderItem.objects.filter(order=order).exists()
            if relevant_order_items_exist:
                 message_for_modal = 'All items for this order are already fully packed or no remaining quantity to pack (considering any prior item removals).'
            else:
                 message_for_modal = 'This order has no items to pack.'
            formset_html_content = f'<p class="text-center py-4 text-gray-500">{message_for_modal}</p>'
        else:
            packing_items_formset = InitialParcelItemFormSet(initial=initial_form_data, prefix='packitems')
            formset_html_content = render_to_string(
                'operation/partials/pack_order_formset.html',
                {'formset': packing_items_formset, 'order': order},
                request=request
            )
            message_for_modal = 'Items loaded for packing.'

        logger.debug(f"[get_order_items_for_packing] For Order PK {order.pk} (ERP ID: {order.erp_order_id}), DB value for is_cold_chain: {order.is_cold_chain}")
        logger.info(f"[PackModalData] Order PK {order.pk} - Raw order.shipping_notes from DB: '{order.shipping_notes}'")

        # --- NEW LOGIC FOR COURIER COUNTS ---
        today = timezone.now().date()
        daily_courier_counts_query = Parcel.objects.filter(
            created_at__date=today,
            order__warehouse=order.warehouse # Filter by the order's warehouse
        ).values('courier_name').annotate(count=Count('courier_name')).order_by('-count')

        daily_courier_counts = []
        for entry in daily_courier_counts_query:
            if entry['courier_name']: # Only include if courier_name is not None/empty
                daily_courier_counts.append({
                    'name': entry['courier_name'],
                    'count': entry['count']
                })
        # --- END NEW LOGIC ---

        return JsonResponse({
            'success': True,
            'order_id': order.pk,
            'erp_order_id': order.erp_order_id,
            'customer_name': order.customer_name,
            'formset_html': formset_html_content,
            'message': message_for_modal,
            'shipping_notes_for_parcel': order.shipping_notes or '',
            'is_cold_chain': order.is_cold_chain,
            'daily_courier_counts': daily_courier_counts # Pass the new data
        })
    except Http404:
        logger.warning(f"Order PK {order_pk} not found in get_order_items_for_packing.")
        return JsonResponse({'success': False, 'message': 'Order not found.'}, status=404)
    except Exception as e_view:
        logger.error(f"Unexpected error in get_order_items_for_packing for order_pk {order_pk}: {e_view}\n{traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': 'An unexpected server error occurred while preparing packing information.'
        }, status=500)



@login_required
@transaction.atomic
def process_packing_for_order(request, order_pk):
    if request.method != 'POST':
        logger.warning(f"process_packing_for_order received a {request.method} request for order_pk {order_pk}, expected POST. Full path: {request.get_full_path()}")
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

    order = get_object_or_404(Order.objects.select_related('warehouse'), pk=order_pk)
    if not request.user.is_superuser and (not request.user.warehouse or order.warehouse != request.user.warehouse):
        return JsonResponse({'success': False, 'message': 'Permission denied for this order.'}, status=403)

    parcel_notes_from_form = request.POST.get('parcel-notes', order.shipping_notes or '')
    courier_name_from_form = request.POST.get('parcel-courier_name', None)
    packaging_type_from_form = request.POST.get('parcel-packaging_type', None)

    num_item_forms = int(request.POST.get('packitems-TOTAL_FORMS', 0))
    items_to_pack_data = []
    any_item_actually_packed_this_session = False

    for i in range(num_item_forms):
        order_item_id_str = request.POST.get(f'packitems-{i}-order_item_id')
        qty_to_pack_str = request.POST.get(f'packitems-{i}-quantity_to_pack')
        batch_id_str = request.POST.get(f'packitems-{i}-selected_batch_item_id')

        if not order_item_id_str or not qty_to_pack_str : continue
        try:
            order_item_id = int(order_item_id_str)
            order_item = OrderItem.objects.select_related('product', 'warehouse_product').get(pk=order_item_id, order=order)
            qty_to_pack = int(qty_to_pack_str)

            if qty_to_pack <= 0: continue

            any_item_actually_packed_this_session = True
            batch_item = None
            if batch_id_str and batch_id_str.isdigit():
                batch_item = InventoryBatchItem.objects.get(pk=int(batch_id_str))
                if batch_item.warehouse_product != order_item.warehouse_product:
                    return JsonResponse({'success': False, 'message': f"Batch {batch_item.batch_number} mismatch."}, status=400)
                if qty_to_pack > batch_item.quantity:
                    return JsonResponse({'success': False, 'message': f"Batch {batch_item.batch_number} stock low."}, status=400)
            elif qty_to_pack > 0 :
                 return JsonResponse({'success': False, 'message': f"Batch selection missing for {order_item.product.sku}."}, status=400)

            total_removed_for_item = order.get_total_removed_quantity_for_item(order_item.id)
            effective_ordered_qty = order_item.quantity_ordered - total_removed_for_item
            quantity_remaining_on_order_item = effective_ordered_qty - order_item.quantity_packed

            if qty_to_pack > quantity_remaining_on_order_item:
                return JsonResponse({'success': False, 'message': f"Cannot pack {qty_to_pack} of {order_item.product.sku}. Only {quantity_remaining_on_order_item} effectively left after considering packed and removed quantities."}, status=400)


            items_to_pack_data.append({
                'order_item': order_item,
                'quantity': qty_to_pack,
                'batch': batch_item
            })
        except (OrderItem.DoesNotExist, InventoryBatchItem.DoesNotExist, ValueError) as e:
             logger.error(f"Error processing item data for packing order {order_pk}: {e}", exc_info=True)
             return JsonResponse({'success': False, 'message': f'Invalid item data: {e}'}, status=400)
        except Exception as e:
             logger.error(f"Unexpected error processing item for order {order_pk}: {e}", exc_info=True)
             return JsonResponse({'success': False, 'message': f'Error processing item: {e}'}, status=500)


    if not any_item_actually_packed_this_session:
        return JsonResponse({'success': False, 'message': 'No items specified with quantity > 0 to pack.'}, status=400)

    parcel = Parcel.objects.create(
        order=order,
        created_by=request.user,
        notes=parcel_notes_from_form,
        courier_name=courier_name_from_form if courier_name_from_form else None,
        packaging_type=packaging_type_from_form if packaging_type_from_form else None,
    )

    if not order.order_display_code:
        order.order_display_code = parcel.parcel_code_system
        order.save(update_fields=['order_display_code'])


    for item_data in items_to_pack_data:
        oi_instance = item_data['order_item']
        batch_instance_for_oi = item_data['batch']
        qty_packed_for_oi_in_this_parcel = item_data['quantity']

        ParcelItem.objects.create(
            parcel=parcel,
            order_item=oi_instance,
            quantity_shipped_in_this_parcel=qty_packed_for_oi_in_this_parcel,
            shipped_from_batch=batch_instance_for_oi
        )

        oi_instance.suggested_batch_item = batch_instance_for_oi
        if batch_instance_for_oi:
            oi_instance.suggested_batch_number_display = batch_instance_for_oi.batch_number
            oi_instance.suggested_batch_expiry_date_display = batch_instance_for_oi.expiry_date

        oi_instance.save(update_fields=['suggested_batch_item', 'suggested_batch_number_display', 'suggested_batch_expiry_date_display'], skip_order_update=True)

        if batch_instance_for_oi:
            original_batch_qty = batch_instance_for_oi.quantity
            batch_instance_for_oi.quantity = F('quantity') - qty_packed_for_oi_in_this_parcel
            batch_instance_for_oi.save(update_fields=['quantity'])
            batch_instance_for_oi.refresh_from_db()
            logger.info(f"Stock Deducted: Batch {batch_instance_for_oi.id} Qty: {original_batch_qty} -> {batch_instance_for_oi.quantity} (-{qty_packed_for_oi_in_this_parcel})")

            StockTransaction.objects.create(
                warehouse=batch_instance_for_oi.warehouse_product.warehouse,
                warehouse_product=batch_instance_for_oi.warehouse_product,
                product=batch_instance_for_oi.warehouse_product.product,
                transaction_type='OUT',
                quantity=-qty_packed_for_oi_in_this_parcel,
                reference_note=f"Packed for Order {order.erp_order_id}, Parcel {parcel.parcel_code_system}, Batch {batch_instance_for_oi.batch_number}",
                related_order=order
            )

    order.refresh_from_db()
    logger.info(f"Order {order.erp_order_id} final status after packing: {order.get_status_display()}")

    messages.success(request, f"Parcel {parcel.parcel_code_system} created (Packaging: {parcel.get_packaging_type_display() or 'N/A'}) for order {order.erp_order_id}.")
    return JsonResponse({
        'success': True,
        'message': f'Parcel created with packaging: {parcel.get_packaging_type_display() or "N/A"}!',
        'redirect_url': request.build_absolute_uri(reverse('operation:order_list')) + f"?tab={DEFAULT_CUSTOMER_ORDERS_TAB}"
    })




@login_required
def get_available_batches_for_order_item(request, order_item_pk):
    try:
        order_item = get_object_or_404(OrderItem.objects.select_related('product', 'warehouse_product__warehouse', 'order__warehouse'), pk=order_item_pk)

        logger.info(f"[get_available_batches] Called for OI PK: {order_item_pk}. WP ID: {order_item.warehouse_product_id if order_item.warehouse_product else 'None'}")

        if not request.user.is_superuser and \
           (not request.user.warehouse or order_item.order.warehouse != request.user.warehouse):
            logger.error(f"Permission denied for user {request.user.id} on OI {order_item_pk} in get_available_batches. User WH: {request.user.warehouse}, Order WH: {order_item.order.warehouse}")
            return JsonResponse({'success': False, 'message': 'Permission denied to access batches for this order item.'}, status=403)

        if not order_item.warehouse_product:
            logger.warning(f"OrderItem {order_item.pk} (Product: {order_item.product.sku if order_item.product else 'N/A'}) has no linked WarehouseProduct. Cannot fetch batches.")
            return JsonResponse({'success': True, 'batches': [], 'message': 'Order item is not linked to a specific warehouse product.'})

        warehouse_product_for_item = order_item.warehouse_product
        today = timezone.now().date()

        logger.info(f"Fetching batches for WP: {warehouse_product_for_item.id} (Product: {warehouse_product_for_item.product.sku}), Criteria: Qty > 0, Not Expired (Today: {today})")

        batches_qs = InventoryBatchItem.objects.filter(
            warehouse_product=warehouse_product_for_item,
            quantity__gt=0
        ).exclude(
            expiry_date__isnull=False, expiry_date__lt=today
        ).order_by(
            F('pick_priority').asc(nulls_last=True),
            F('expiry_date').asc(nulls_last=True),
            'date_received'
        )

        logger.info(f"Found {batches_qs.count()} total batches matching quantity/expiry for WP {warehouse_product_for_item.id} before serialization for OI {order_item.pk}.")

        batches_data = []
        for batch in batches_qs:
            priority_label = ""
            if batch.pick_priority == 0: priority_label = " [Default]"
            elif batch.pick_priority == 1: priority_label = " [Secondary]"

            location_display = f"[{batch.location_label}]" if batch.location_label else "NoLoc"
            batch_display = f"Batch: {batch.batch_number}" if batch.batch_number else "NoBatch"
            expiry_display = f"Exp: {batch.expiry_date.strftime('%d/%m/%y')}" if batch.expiry_date else "NoExp"
            qty_display = f"Qty: {batch.quantity}"

            batch_data_entry = {
                'id': batch.pk,
                'display_name': f"{location_display} | {batch_display} | {expiry_display} | {qty_display}{priority_label}",
                'quantity_available': batch.quantity,
                'pick_priority': batch.pick_priority
            }
            batches_data.append(batch_data_entry)
            logger.debug(f"Added batch to dropdown data: ID {batch.pk}, Prio {batch.pick_priority}, Qty {batch.quantity}, Label: ...{priority_label}")


        if not batches_data:
            logger.info(f"No available batches were serialized for OI {order_item.pk} (WP: {warehouse_product_for_item.id}). Dropdown will indicate no batches.")
            return JsonResponse({'success': True, 'batches': [], 'message': 'No available stock batches found for this item.'})

        logger.info(f"Successfully prepared {len(batches_data)} batch options for OI {order_item.pk} dropdown.")
        return JsonResponse({'success': True, 'batches': batches_data})

    except Http404:
        logger.error(f"[get_available_batches] OrderItem PK {order_item_pk} not found.")
        return JsonResponse({'success': False, 'message': 'Order item not found.'}, status=404)
    except Exception as e:
        logger.error(f"[get_available_batches] Unexpected error for OI PK {order_item_pk}: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'An unexpected server error occurred: {str(e)}'}, status=500)

@login_required
def get_order_items_for_editing(request, order_pk):
    logger.debug(f"--- get_order_items_for_editing for order_pk: {order_pk} ---")
    order = get_object_or_404(Order.objects.prefetch_related(
        Prefetch('items', queryset=OrderItem.objects.select_related('product').order_by('product__name'))
    ), pk=order_pk)
    logger.debug(f"Order found: {order.erp_order_id}, Status: {order.get_status_display()}")

    if not request.user.is_superuser and (not request.user.warehouse or order.warehouse != request.user.warehouse):
        logger.warning(f"Permission denied for user {request.user.email} to edit order {order_pk}")
        return JsonResponse({'success': False, 'message': 'Permission denied for this order.'}, status=403)

    if order.status != 'PARTIALLY_SHIPPED':
        logger.info(f"Order {order_pk} status is '{order.get_status_display()}', cannot edit items for removal.")
        return JsonResponse({'success': False, 'message': f'Order status is "{order.get_status_display()}", cannot edit items for removal.'}, status=400)

    initial_form_data = []
    logger.debug(f"Processing order items for order {order.erp_order_id}:")
    order_items_all = order.items.all()
    logger.debug(f"Total items in order: {order_items_all.count()}")

    for item in order_items_all:
        total_removed_for_this_item = order.get_total_removed_quantity_for_item(item.id)
        balance_qty = item.quantity_ordered - item.quantity_packed - total_removed_for_this_item

        logger.debug(f"  Item PK: {item.pk}, Prod: {item.product.sku if item.product else 'N/A'}, "
                     f"Ordered: {item.quantity_ordered}, Packed: {item.quantity_packed}, "
                     f"Total Removed Logged: {total_removed_for_this_item}, Calculated Balance: {balance_qty}")

        if balance_qty > 0:
            initial_form_data.append({
                'order_item_id': item.pk,
                'product_name': item.product.name if item.product else item.erp_product_name,
                'sku': item.product.sku if item.product else "N/A",
            })
            logger.debug(f"    -> ADDED to initial_form_data (PK: {item.pk}, Balance: {balance_qty})")
        else:
            logger.debug(f"    -> SKIPPED for initial_form_data (PK: {item.pk}, Balance: {balance_qty})")

    formset_initial_with_balance = []
    if not initial_form_data:
        logger.debug("initial_form_data is EMPTY. No items will be shown in the formset.")
    else:
        logger.debug("Populating formset_initial_with_balance...")
        for item_data_initial in initial_form_data:
            oi = next((i for i in order_items_all if i.pk == item_data_initial['order_item_id']), None)
            if not oi:
                logger.error(f"Could not find OrderItem with PK {item_data_initial['order_item_id']} in prefetched items.")
                continue

            total_removed_for_oi = order.get_total_removed_quantity_for_item(oi.id)
            balance_for_form = oi.quantity_ordered - oi.quantity_packed - total_removed_for_oi
            logger.debug(f"  For Form (OI PK: {oi.pk}): balance_for_form = {balance_for_form}")
            formset_initial_with_balance.append({
                **item_data_initial,
                'balance_quantity_to_pack': balance_for_form
            })

    logger.debug(f"Final formset_initial_with_balance count: {len(formset_initial_with_balance)}")
    edit_items_formset = RemoveOrderItemFormSet(initial=formset_initial_with_balance, prefix='edititems')
    logger.debug(f"edit_items_formset created. Number of forms: {len(edit_items_formset.forms)}")

    if not edit_items_formset.forms:
        logger.debug("No forms in edit_items_formset. formset_html will likely be empty or show 'no items'.")

    formset_template_name = 'operation/partials/edit_order_formset.html'
    logger.debug(f"Rendering template: {formset_template_name}")
    formset_html_content = render_to_string(
        formset_template_name,
        {'formset': edit_items_formset, 'order': order},
        request=request
    )

    logger.debug(f"Length of rendered formset_html_content: {len(formset_html_content)}")
    if len(formset_html_content) < 200:
        logger.debug(f"Short formset_html_content: '{formset_html_content[:500]}...'")

    removed_items_log_display = order.items_removed_log or []
    logger.debug(f"Removed items log being sent to client: {removed_items_log_display}")

    logger.debug(f"--- get_order_items_for_editing END for order_pk: {order_pk} ---")
    return JsonResponse({
        'success': True,
        'order_id': order.pk,
        'erp_order_id': order.erp_order_id,
        'customer_name': order.customer_name,
        'formset_html': formset_html_content,
        'removed_items_log': removed_items_log_display,
        'message': 'Items loaded for editing/removal.'
    })


@login_required
@transaction.atomic
def process_order_item_removal(request, order_pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

    order = get_object_or_404(Order, pk=order_pk)

    if not request.user.is_superuser and (not request.user.warehouse or order.warehouse != request.user.warehouse):
        return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)

    if order.status != 'PARTIALLY_SHIPPED':
        return JsonResponse({'success': False, 'message': f'Order status is "{order.get_status_display()}", cannot process removals.'}, status=400)

    temp_formset_data_from_post = {}
    for key, value in request.POST.items():
        if key.startswith('edititems-') and key.endswith('-order_item_id'):
            form_idx = key.split('-')[1]
            if form_idx not in temp_formset_data_from_post: temp_formset_data_from_post[form_idx] = {}
            temp_formset_data_from_post[form_idx]['order_item_id'] = value
        elif key.startswith('edititems-') and key.endswith('-quantity_to_remove'):
            form_idx = key.split('-')[1]
            if form_idx not in temp_formset_data_from_post: temp_formset_data_from_post[form_idx] = {}
            temp_formset_data_from_post[form_idx]['quantity_to_remove'] = value

    reconstructed_initial_for_formset = []
    for idx_str, data_dict in temp_formset_data_from_post.items():
        try:
            oi_id = int(data_dict['order_item_id'])
            oi = OrderItem.objects.get(pk=oi_id, order=order)
            total_removed_for_oi = order.get_total_removed_quantity_for_item(oi.id)
            balance = oi.quantity_ordered - oi.quantity_packed - total_removed_for_oi
            reconstructed_initial_for_formset.append({
                'order_item_id': oi_id,
                'product_name': oi.product.name,
                'sku': oi.product.sku,
                'balance_quantity_to_pack': balance,
                'quantity_to_remove': data_dict.get('quantity_to_remove', 0)
            })
        except (OrderItem.DoesNotExist, ValueError, KeyError) as e:
            logger.error(f"Error reconstructing formset initial data for validation: {e}")
            return JsonResponse({'success': False, 'message': 'Error processing form data.'}, status=400)

    is_formset_valid = True
    cleaned_forms_data = []

    for i in range(int(request.POST.get('edititems-TOTAL_FORMS', 0))):
        form_data_for_instance = {k.replace(f'edititems-{i}-', ''): v for k, v in request.POST.items() if k.startswith(f'edititems-{i}-')}
        current_form_oi_id = int(form_data_for_instance.get('order_item_id', 0))
        balance_for_this_form = 0
        for init_data in reconstructed_initial_for_formset:
            if init_data['order_item_id'] == current_form_oi_id:
                balance_for_this_form = init_data['balance_quantity_to_pack']
                break

        form = RemoveOrderItemForm(
            form_data_for_instance,
            balance_quantity_to_pack=balance_for_this_form)

        if form.is_valid():
            cleaned_forms_data.append(form.cleaned_data)
        else:
            is_formset_valid = False
            logger.warning(f"Form {i} errors: {form.errors.as_json()}")
            return JsonResponse({'success': False, 'message': 'Invalid data in removal form. Please check quantities.'}, status=400)

    if is_formset_valid:
        removed_items_summary_for_log = order.items_removed_log or []
        any_actual_removal = False

        for data in cleaned_forms_data:
            order_item_id = data.get('order_item_id')
            qty_removed = data.get('quantity_to_remove', 0)

            if qty_removed > 0:
                try:
                    order_item = OrderItem.objects.get(pk=order_item_id, order=order)
                    removed_items_summary_for_log.append({
                        'order_item_id': order_item.id,
                        'product_sku': order_item.product.sku,
                        'product_name': order_item.product.name,
                        'removed_qty': qty_removed,
                        'removed_at': timezone.now().isoformat()
                    })
                    any_actual_removal = True
                except OrderItem.DoesNotExist:
                    messages.error(request, f"Error: Order Item ID {order_item_id} not found for this order.")
                    return JsonResponse({'success': False, 'message': f"Item ID {order_item_id} not found."}, status=400)

        if any_actual_removal:
            order.items_removed_log = removed_items_summary_for_log
        order.save()
        order.refresh_from_db()

        messages.success(request, "Order items updated successfully. Status refreshed.")
        return JsonResponse({
            'success': True,
            'message': 'Items processed successfully! Order status may have been updated.',
            'new_order_status': order.get_status_display(),
            'order_id': order.pk
        })
    else:
        logger.error(f"Order item removal formset was not valid for order {order_pk}.")
        return JsonResponse({'success': False, 'message': 'Invalid form data submitted.'}, status=400)

