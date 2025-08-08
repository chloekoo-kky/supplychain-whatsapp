# ===== warehouse/admin.py =====
import csv
import io
# import chardet # Ensure chardet is installed in your environment
import openpyxl
import logging
from openpyxl.utils import get_column_letter

from django import forms
from django.contrib import admin, messages
from django.http import HttpResponse
from django.urls import path, reverse
from django.shortcuts import redirect
from django.utils.html import format_html
from django.db import transaction # For atomic operations
from django.contrib.admin.views.main import ChangeList

# Import your models
from .models import (Warehouse,
                     WarehouseProduct,
                     PurchaseOrder,
                     PurchaseOrderItem,
                     PurchaseOrderStatusLog,
                     )

from inventory.models import Product, Supplier, StockTransaction # Make sure StockTransaction is imported

logger = logging.getLogger(__name__)

class WarehouseProductCsvImportForm(forms.Form):
    csv_upload = forms.FileField()

class ExcelImportForm(forms.Form):
    excel_upload = forms.FileField(label="Select Excel file (.xlsx)")


@admin.register(WarehouseProduct)
class WarehouseProductAdmin(admin.ModelAdmin):
    change_list_template = "admin/warehouse/warehouseproduct/changelist.html"

    list_display = (
        'get_warehouse_name', 'code', 'get_product_sku', 'get_product_name',
        'quantity', 'threshold', 'supplier'  # This will now show the code
    )
    list_display_links = ('get_product_name',)

    # Keep this as is
    list_editable = ('supplier',)

    search_fields = (
        'warehouse__name', 'code', 'product__sku', 'product__name',
        'supplier__name', 'supplier__code'
    )
    list_filter = ('warehouse', 'supplier', 'product__name')
    autocomplete_fields = ['product', 'warehouse']
    ordering = ('warehouse__name', 'code', 'product__name')

    fieldsets = (
        (None, {'fields': ('warehouse', 'product', 'code', 'supplier')}),
        ('Stock Details', {'fields': ('quantity', 'threshold')}),
    )

    EXCEL_HEADERS = [
        'Warehouse Name', 'Product SKU', 'Warehouse Product Code',
        'Product Name', 'Quantity', 'Threshold', 'Supplier Code'
    ]


    def get_product_name(self, obj):
        return obj.product.name if obj.product else "-"
    get_product_name.short_description = 'Product Name'
    get_product_name.admin_order_field = 'product__name'

    def get_warehouse_name(self, obj):
        return obj.warehouse.name if obj.warehouse else "-"
    get_warehouse_name.short_description = 'Warehouse'
    get_warehouse_name.admin_order_field = 'warehouse__name'

    def get_product_sku(self, obj):
        return obj.product.sku if obj.product else "-"
    get_product_sku.short_description = 'SKU'
    get_product_sku.admin_order_field = 'product__sku'

    def get_product_name(self, obj):
        return obj.product.name if obj.product else "-"
    get_product_name.short_description = 'Product Name'
    get_product_name.admin_order_field = 'product__name'

    def get_supplier_code_display(self, obj):
        if obj.supplier:
            return obj.supplier.code
        return "-"
    get_supplier_code_display.short_description = 'Supplier Code'
    get_supplier_code_display.admin_order_field = 'supplier__code'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.admin_site.admin_view(self.upload_excel), name='warehouse_warehouseproduct_upload_excel'),
            path('download-excel-template/', self.admin_site.admin_view(self.download_excel_template), name='warehouse_warehouseproduct_download_excel_template'),
            path('export-excel/', self.admin_site.admin_view(self.export_excel), name='warehouse_warehouseproduct_export_excel'),
        ]
        return custom_urls + urls

    def download_excel_template(self, request):
        # ... (This method is fine) ...
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "WarehouseProduct Template"
        sheet.append(self.EXCEL_HEADERS)
        sheet.append(['Main Warehouse', 'SKU001', 'MW-SKU001', 'Sample Product One', 100, 10, 'SUP001'])
        sheet.append(['Secondary Warehouse', 'SKU002', '', 'Another Product', 50, 5, ''])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="warehouseproduct_template.xlsx"'
        workbook.save(response)
        return response

    def export_excel(self, request):
        # ... (This method is fine) ...
        sortable_by = self.get_sortable_by(request)
        cl = ChangeList(
            request, self.model, self.list_display, self.list_display_links,
            self.list_filter, self.date_hierarchy, self.search_fields,
            self.list_select_related, self.list_per_page, self.list_max_show_all,
            self.list_editable, self, sortable_by, search_help_text=self.search_help_text,

        )
        queryset = cl.get_queryset(request).select_related('warehouse', 'product', 'supplier')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Warehouse Products Export"
        sheet.append(self.EXCEL_HEADERS)
        for obj in queryset:
            sheet.append([
                obj.warehouse.name if obj.warehouse else '', obj.product.sku if obj.product else '',
                obj.code if obj.code else '', obj.product.name if obj.product else '',
                obj.quantity, obj.threshold, obj.supplier.code if obj.supplier else '',
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="warehouse_products_export.xlsx"'
        workbook.save(response)
        return response

    def upload_excel(self, request):
        """
        Handles the upload of an Excel file (.xlsx) to bulk-update or create
        warehouse product stock records.
        """
        logger.info("--- Starting Warehouse Product Upload Process (Excel) ---")

        if request.method != "POST":
            self.message_user(request, "This action requires a file upload.", level=messages.WARNING)
            return redirect("..")

        uploaded_file = request.FILES.get("excel_upload")
        if not uploaded_file:
            self.message_user(request, "No file was uploaded. Please select a file.", level=messages.ERROR)
            return redirect("..")

        logger.info(f"Processing uploaded file as Excel: {uploaded_file.name}")

        # --- Step 1: Read the file as an Excel Workbook using openpyxl ---
        try:
            # data_only=True ensures we get cell values instead of formulas
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            sheet = workbook.active
            logger.debug("Successfully loaded Excel workbook.")
        except Exception as e:
            logger.error(f"Failed to load Excel file. Error: {e}", exc_info=True)
            self.message_user(request, f"Error reading Excel file. Ensure it is a valid .xlsx file. Error: {e}", level=messages.ERROR)
            return redirect("..")

        # --- Step 2: Process the worksheet row-by-row ---
        created_count, updated_count, errors = 0, 0, []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            try:
                cell_values = [cell.value for cell in row]
                logger.debug(f"Row {row_idx}: Raw data from cells: {cell_values}")

                # Skip empty rows gracefully
                if not any(cell_values) or cell_values[0] is None:
                    logger.warning(f"Row {row_idx}: Skipped because it is an empty row.")
                    continue

                # --- Data Extraction and Cleaning ---
                warehouse_name = str(cell_values[0]).strip()
                # Correctly parse SKU, Quantity, and Threshold as numbers
                product_sku = str(int(float(cell_values[2])))
                quantity_val = int(float(cell_values[3]))
                threshold_val = int(float(cell_values[4]))
                logger.debug(f"Row {row_idx}: Parsed Data -> Warehouse: '{warehouse_name}', SKU: '{product_sku}', Qty: {quantity_val}, Threshold: {threshold_val}")

                if not product_sku or not warehouse_name:
                    errors.append(f"Row {row_idx}: Skipped because Warehouse Name or Product SKU is missing.")
                    continue

                # --- Database Operation ---
                product_to_update = WarehouseProduct.objects.filter(
                    warehouse__name__iexact=warehouse_name,
                    product__sku__iexact=product_sku
                ).first()

                if product_to_update:
                    # UPDATE existing product
                    product_to_update.quantity = quantity_val
                    product_to_update.threshold = threshold_val
                    product_to_update.save(update_fields=['quantity', 'threshold'])
                    updated_count += 1
                    logger.info(f"Row {row_idx}: Successfully UPDATED product with SKU '{product_sku}'.")
                else:
                    # CREATE new product if it doesn't exist
                    try:
                        warehouse = Warehouse.objects.get(name__iexact=warehouse_name)
                        product = Product.objects.get(sku__iexact=product_sku)
                        WarehouseProduct.objects.create(
                            warehouse=warehouse, product=product, quantity=quantity_val, threshold=threshold_val
                        )
                        created_count += 1
                        logger.info(f"Row {row_idx}: Successfully CREATED new product for SKU '{product_sku}'.")
                    except Warehouse.DoesNotExist:
                        errors.append(f"Row {row_idx}: Could not create. Warehouse '{warehouse_name}' not found.")
                    except Product.DoesNotExist:
                        errors.append(f"Row {row_idx}: Could not create. Product with SKU '{product_sku}' not found.")

            except (ValueError, TypeError, IndexError):
                errors.append(f"Row {row_idx}: Skipped due to invalid data format or missing columns.")
            except Exception as e:
                errors.append(f"Row {row_idx}: An unexpected error occurred: {e}")

        # --- Final Summary ---
        summary_message = f"Process complete. Updated: {updated_count}, Created: {created_count}."
        if created_count > 0 or updated_count > 0:
            self.message_user(request, summary_message, level=messages.SUCCESS)

        if errors:
            error_summary = f"Encountered {len(errors)} issues."
            self.message_user(request, error_summary, level=messages.WARNING)
            for error in errors[:5]:
                self.message_user(request, error, level=messages.WARNING)

        return redirect("..")



@admin.register(Warehouse)
class WarehouseAdmin(admin.ModelAdmin):
    list_display = ('name', 'location')
    search_fields = ('name', 'location')

class PurchaseOrderItemInline(admin.TabularInline):
    model = PurchaseOrderItem
    fields = ('item', 'quantity', 'price', 'received_quantity')
    autocomplete_fields = ['item'] # Use autocomplete for selecting WarehouseProduct
    extra = 1

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('item__product', 'item__warehouse')


@admin.register(PurchaseOrder)
class PurchaseOrderAdmin(admin.ModelAdmin):
    list_display = ('id', 'supplier', 'status', 'eta', 'total_amount', 'created_at', 'last_updated_date')
    list_filter = ('status', 'supplier', 'items__item__warehouse')
    search_fields = ('id', 'supplier__name')
    inlines = [PurchaseOrderItemInline]
    date_hierarchy = 'created_at'


    def last_updated_at_display(self, obj):
        return self._format_date(obj.last_updated_date)
    last_updated_at_display.short_description = 'Last Updated At'


@admin.register(PurchaseOrderItem)
class PurchaseOrderItemAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'get_po_link',
        'get_po_status',
        'get_supplier_code',
        'get_product_sku',
        'get_product_name',
        'get_warehouse_name',
        'quantity',
        'price',
        'get_total_price',
        'received_quantity',
        'get_balance_quantity'
    )
    list_filter = (
        'purchase_order__status',
        'purchase_order__supplier',
        'item__product',
        'item__warehouse',
    )
    search_fields = (
        'purchase_order__id',
        'item__product__sku',
        'item__product__name',
        'item__warehouse__name',
        'purchase_order__supplier__code',
        'purchase_order__supplier__name',
    )
    ordering = ('-purchase_order__last_updated_date', 'item__product__name')
    autocomplete_fields = ['purchase_order', 'item']
    readonly_fields = ('get_total_price_display', 'get_balance_quantity_display') # For detail view

    fieldsets = (
        (None, {'fields': ('purchase_order', 'item')}),
        ('Quantities & Pricing', {'fields': ('quantity', 'price', 'get_total_price_display', 'received_quantity', 'get_balance_quantity_display')}),
    )

    def get_po_link(self, obj):
        if obj.purchase_order:
            link = reverse("admin:warehouse_purchaseorder_change", args=[obj.purchase_order.id])
            return format_html('<a href="{}">PO-{}</a>', link, obj.purchase_order.id)
        return "-"
    get_po_link.short_description = 'Purchase Order'
    get_po_link.admin_order_field = 'purchase_order__id'

    def get_po_status(self, obj):
        return obj.purchase_order.get_status_display() if obj.purchase_order else "-"
    get_po_status.short_description = 'PO Status'
    get_po_status.admin_order_field = 'purchase_order__status'

    def get_supplier_code(self, obj):
        if obj.purchase_order and obj.purchase_order.supplier:
            return obj.purchase_order.supplier.code
        # Fallback logic (less common for POItem, as PO should have a supplier)
        elif obj.item and obj.item.supplier: # Supplier on WarehouseProduct
            return obj.item.supplier.code
        return "-"
    get_supplier_code.short_description = 'Supplier Code'
    get_supplier_code.admin_order_field = 'purchase_order__supplier__code'

    def get_product_sku(self, obj):
        return obj.item.product.sku if obj.item and obj.item.product else "-"
    get_product_sku.short_description = 'SKU'
    get_product_sku.admin_order_field = 'item__product__sku'

    def get_product_name(self, obj):
        return obj.item.product.name if obj.item and obj.item.product else "-"
    get_product_name.short_description = 'Product Name'
    get_product_name.admin_order_field = 'item__product__name'

    def get_warehouse_name(self, obj):
        return obj.item.warehouse.name if obj.item and obj.item.warehouse else "-"
    get_warehouse_name.short_description = 'Warehouse'
    get_warehouse_name.admin_order_field = 'item__warehouse__name'

    def get_total_price(self, obj): # For list_display
        return obj.total_price
    get_total_price.short_description = 'Total Price'

    def get_total_price_display(self, obj): # For readonly_fields in detail view
        return obj.total_price
    get_total_price_display.short_description = 'Total Price'


    def get_balance_quantity(self, obj): # For list_display
        return obj.balance_quantity
    get_balance_quantity.short_description = 'Balance Qty'

    def get_balance_quantity_display(self, obj): # For readonly_fields in detail view
        return obj.balance_quantity
    get_balance_quantity_display.short_description = 'Balance Qty'


    @transaction.atomic
    def delete_model(self, request, obj):
        """
        Override to adjust WarehouseProduct quantity and create StockTransaction
        when a PurchaseOrderItem is deleted.
        This handles deletion of a single item from its change form.
        """
        warehouse_product = obj.item
        quantity_to_reverse = obj.received_quantity # Use the actual received quantity

        if warehouse_product and quantity_to_reverse > 0:
            # Deduct the received quantity from the warehouse product
            warehouse_product.quantity -= quantity_to_reverse
            warehouse_product.save(update_fields=['quantity'])

            # Create a stock transaction to record this reversal
            StockTransaction.objects.create(
                warehouse=warehouse_product.warehouse,
                warehouse_product=warehouse_product,
                product=warehouse_product.product,
                transaction_type='ADJUST', # Or a more specific type like 'PO_ITEM_DEL_ADJ'
                quantity=-quantity_to_reverse, # Negative to indicate deduction
                reference_note=f"Deletion of received PO Item ID {obj.id} for PO-{obj.purchase_order.id}",
                related_po=obj.purchase_order
            )
            messages.success(request, f"Stock for {warehouse_product.product.name} in {warehouse_product.warehouse.name} reduced by {quantity_to_reverse} due to PO Item deletion.")

        # Also, update the PO's status if necessary, e.g., if it was 'DELIVERED'
        # and now it's not fully received anymore.
        po = obj.purchase_order
        super().delete_model(request, obj) # Delete the PurchaseOrderItem

        # After deletion, re-check and update PO status
        if po:
            is_now_fully_received = po.is_fully_received()
            if not is_now_fully_received and po.status == 'DELIVERED':
                po.status = 'PARTIALLY_DELIVERED' # Or back to PAYMENT_MADE if nothing else received
                # Check if any items are received at all for this PO
                if not po.items.filter(received_quantity__gt=0).exists():
                    # If no items are received anymore, maybe revert to an earlier status
                    # This depends on your exact workflow. For instance, if 'PAYMENT_MADE' is the status before any receiving.
                    po.status = 'PAYMENT_MADE' # Example, adjust as per your PO lifecycle
                po.save() # This will also update status dates via model's save method
            elif is_now_fully_received and po.status != 'DELIVERED':
                 po.status = 'DELIVERED' # Should not happen if we just deleted a received item
                 po.save()


    @transaction.atomic
    def delete_queryset(self, request, queryset):
        """
        Override to adjust WarehouseProduct quantities and create StockTransactions
        for bulk deletion of PurchaseOrderItems from the changelist.
        """
        pos_to_recheck_status = set() # Keep track of POs whose status might need update

        for obj in queryset:
            warehouse_product = obj.item
            quantity_to_reverse = obj.received_quantity

            if warehouse_product and quantity_to_reverse > 0:
                warehouse_product.quantity -= quantity_to_reverse
                warehouse_product.save(update_fields=['quantity'])

                StockTransaction.objects.create(
                    warehouse=warehouse_product.warehouse,
                    warehouse_product=warehouse_product,
                    product=warehouse_product.product,
                    transaction_type='ADJUST',
                    quantity=-quantity_to_reverse,
                    reference_note=f"Bulk deletion of received PO Item ID {obj.id} for PO-{obj.purchase_order.id}",
                    related_po=obj.purchase_order
                )
            pos_to_recheck_status.add(obj.purchase_order)

        num_deleted = queryset.count()
        queryset.delete() # Perform the actual deletion

        # After deletion, re-check and update status for affected POs
        for po in pos_to_recheck_status:
            if po: # Ensure PO object still exists (it should)
                is_now_fully_received = po.is_fully_received()
                if not is_now_fully_received and po.status == 'DELIVERED':
                    po.status = 'PARTIALLY_DELIVERED'
                    if not po.items.filter(received_quantity__gt=0).exists():
                         po.status = 'PAYMENT_MADE' # Example
                    po.save()
                # No need to check if is_now_fully_received and status != DELIVERED, as we are deleting items.

        self.message_user(request, f"Successfully deleted {num_deleted} purchase order item(s) and adjusted stock accordingly.", messages.SUCCESS)

@admin.register(PurchaseOrderStatusLog)
class PurchaseOrderStatusLogAdmin(admin.ModelAdmin):
    list_display = ('purchase_order', 'status', 'timestamp', 'user', 'notes')
    list_filter = ('status',)
    search_fields = ('purchase_order__id', 'user__username')
    readonly_fields = ('purchase_order', 'status', 'timestamp', 'user', 'notes') # Log entries should be immutable

    def has_add_permission(self, request):
        # Prevent manual creation of log entries from the admin
        return False

    def has_change_permission(self, request, obj=None):
        # Prevent editing of log entries from the admin
        return False
