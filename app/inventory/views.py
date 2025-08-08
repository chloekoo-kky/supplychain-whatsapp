# app/inventory/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET
from django.db.models import Sum, Q, Prefetch
from django.db import transaction, IntegrityError
from django.contrib import messages # Import messages
from django.forms import inlineformset_factory, formset_factory # Ensure formset_factory is imported
from django.utils.text import slugify
from django.db import transaction as db_transaction



from .models import (
    InventoryBatchItem, Product, Supplier,
    StockTakeSession, StockTakeItem, StockDiscrepancy,
     ErpStockCheckSession, ErpStockCheckItem, WarehouseProductDiscrepancy  # New models
)
from warehouse.models import Warehouse, WarehouseProduct as NewAggregateWarehouseProduct, Warehouse # <<< ADDED Warehouse import
from .forms import (
    InventoryBatchItemForm, StockTakeItemForm, StockTakeItemFormSet, # New forms
    StockTakeSessionSelectionForm,
    ErpStockCheckUploadForm,
    DefaultPickItemFormSet,  # Existing formset for default
    DefaultPickItemForm # To be reused
)
from .services import create_stock_take_session_from_csv

from django.contrib.admin.views.decorators import staff_member_required # For superuser/staff views
from django.http import HttpResponse
import csv
import json
import io
import openpyxl
import xlrd
import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger(__name__)


@login_required
@login_required
def inventory_batch_list_view(request):
    user = request.user
    base_queryset = NewAggregateWarehouseProduct.objects.select_related(
        'product',
        'warehouse',
        'supplier'
    ).prefetch_related(
        Prefetch('batches', queryset=InventoryBatchItem.objects.order_by('expiry_date'))
    )

    # For non-superusers, restrict to their assigned warehouse. Superusers see all.
    if not user.is_superuser and user.warehouse:
        base_queryset = base_queryset.filter(warehouse=user.warehouse)
    elif not user.is_superuser and not user.warehouse:
        base_queryset = base_queryset.none()
        messages.warning(request, "You are not assigned to a warehouse. Please contact an administrator.")

    # This GET parameter is kept in case you want to use it for setting an initial filter state via URL.
    selected_warehouse_id = request.GET.get('warehouse')

    warehouse_products_qs = base_queryset.order_by(
        'product__name',
        'warehouse__name'
    )

    today = timezone.now().date()
    processed_warehouse_products_list = []

    for wp in warehouse_products_qs:
        current_aggregate_stock = wp.quantity if wp.quantity is not None else 0
        sum_of_batched_stock = wp.batches.aggregate(total_qty=Sum('quantity'))['total_qty'] or 0
        wp.calculated_unbatched_quantity = current_aggregate_stock - sum_of_batched_stock
        wp.processed_batches = wp.batches.all()
        processed_warehouse_products_list.append(wp)

    all_batches_qs = InventoryBatchItem.objects.select_related(
        'warehouse_product__product',
        'warehouse_product__warehouse'
    )
    if not user.is_superuser and user.warehouse:
        all_batches_qs = all_batches_qs.filter(warehouse_product__warehouse=user.warehouse)
    elif not user.is_superuser and not user.warehouse:
        all_batches_qs = all_batches_qs.none()

    all_inventory_batches_for_modals = all_batches_qs.order_by(
        'warehouse_product__product__name', 'batch_number'
    )

    add_batch_form_instance = InventoryBatchItemForm(request=request)
    user_warehouse = request.user.warehouse if hasattr(request.user, 'warehouse') and request.user.warehouse else None
    default_pick_formset = DefaultPickItemFormSet(
        prefix='default_picks', initial=[], warehouse=user_warehouse
    )

    # MODIFICATION: Get warehouses for the filter buttons.
    # For superusers, we get only warehouses that have products in the current list.
    # For regular users, this list remains empty, so no filters are shown.
    warehouses_for_filter = []
    if user.is_superuser:
        # This is more efficient and user-friendly than getting all warehouses.
        warehouse_ids = base_queryset.values_list('warehouse_id', flat=True).distinct()
        warehouses_for_filter = Warehouse.objects.filter(pk__in=warehouse_ids).order_by('name')


    context = {
        'warehouse_products': processed_warehouse_products_list,
        'all_inventory_batches': all_inventory_batches_for_modals,
        'add_batch_form': add_batch_form_instance,
        'page_title': 'Batches & Stock Levels',
        'default_pick_formset': default_pick_formset,
        'default_pick_formset_prefix': default_pick_formset.prefix,
        # MODIFICATION: Changed context key to 'warehouses' to match the template.
        'warehouses': warehouses_for_filter,
        'selected_warehouse_id': int(selected_warehouse_id) if selected_warehouse_id else None,
    }
    return render(request, 'inventory/inventory_batch_list.html', context)


@login_required
def export_inventory_batch_to_excel(request):
    """
    Exports an Excel file summarizing inventory batches, including total stock
    and stock discrepancy for each product in each warehouse.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventory Batch Report"

    headers = [
        "Product SKU", "Product Name", "Warehouse", "Total Stock (Batched)",
        "System Stock (On Hand)", "Stock Discrepancy"
    ]
    sheet.append(headers)

    # --- THIS IS THE FIX: Use 'batches' instead of 'inventory_batches' ---
    warehouse_products = NewAggregateWarehouseProduct.objects.filter(
        batches__isnull=False
    ).distinct()

    batch_totals = NewAggregateWarehouseProduct.objects.filter(
        id__in=warehouse_products.values_list('id', flat=True)
    ).annotate(
        batched_stock=Sum('batches__quantity') # Corrected field name here
    )
    # --------------------------------------------------------------------

    batch_stock_dict = {bt.id: bt.batched_stock for bt in batch_totals}

    for wp in warehouse_products:
        total_batched_stock = batch_stock_dict.get(wp.id, 0)
        discrepancy = total_batched_stock - wp.quantity

        sheet.append([
            wp.product.sku,
            wp.product.name,
            wp.warehouse.name,
            total_batched_stock,
            wp.quantity,
            discrepancy
        ])

    # --- Save to an in-memory stream ---
    excel_stream = io.BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)

    response = HttpResponse(
        excel_stream,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="inventory_batch_report.xlsx"'

    return response



@login_required
@require_POST
def add_inventory_batch(request):
    user = request.user
    post_data = request.POST.copy()
    hidden_wp_id = post_data.get('hidden_warehouse_product_id')

    if hidden_wp_id:
        post_data['warehouse_product'] = hidden_wp_id

    selected_wp_id = post_data.get('warehouse_product')
    if selected_wp_id:
        try:
            wp_instance = NewAggregateWarehouseProduct.objects.get(pk=selected_wp_id)
            if not user.is_superuser and user.warehouse and wp_instance.warehouse != user.warehouse:
                return JsonResponse({'success': False, 'error': 'Permission denied. You cannot add a batch to this warehouse product.'}, status=403)
            elif not user.is_superuser and not user.warehouse:
                return JsonResponse({'success': False, 'error': 'Permission denied. No warehouse assigned.'}, status=403)
        except NewAggregateWarehouseProduct.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Selected warehouse product does not exist.'}, status=400)
    else:
        return JsonResponse({'success': False, 'error': 'Warehouse product selection is missing.'}, status=400)

    form = InventoryBatchItemForm(post_data, request=request)

    if form.is_valid():
        try:
            batch_item = form.save(commit=False)
            if not user.is_superuser and user.warehouse and batch_item.warehouse_product.warehouse != user.warehouse:
                return JsonResponse({'success': False, 'error': 'Permission denied. Mismatch in final warehouse assignment.'}, status=403)
            batch_item.save()
            return JsonResponse({'success': True, 'message': 'Batch added successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error saving batch: {str(e)}'}, status=500)
    else:
        errors = {field: error[0] for field, error in form.errors.items()}
        return JsonResponse({'success': False, 'errors': errors, 'message': 'Please correct the form errors.'}, status=400)


@login_required
@require_POST
def edit_inventory_batch(request, batch_pk): # <-- FIX: Changed pk to batch_pk
    try:
        batch_item = get_object_or_404(InventoryBatchItem.objects.select_related('warehouse_product__warehouse'), pk=batch_pk) # <-- FIX: Use batch_pk

        if not request.user.is_superuser and (not request.user.warehouse or request.user.warehouse != batch_item.warehouse_product.warehouse):
            return JsonResponse({'success': False, 'message': 'You do not have permission for this warehouse.'}, status=403)

        form = InventoryBatchItemForm(request.POST, instance=batch_item, request=request)

        if form.is_valid():
            updated_batch_item = form.save(commit=False)
            set_as_default = request.POST.get('set_as_default_pick') == 'on'

            with transaction.atomic():
                if set_as_default:
                    InventoryBatchItem.objects.filter(
                        warehouse_product=updated_batch_item.warehouse_product,
                        pick_priority=0
                    ).exclude(pk=updated_batch_item.pk).update(pick_priority=None)
                    updated_batch_item.pick_priority = 0
                elif updated_batch_item.pick_priority == 0:
                    updated_batch_item.pick_priority = None

                updated_batch_item.save()

            return JsonResponse({
                'success': True,
                'message': 'Batch updated successfully!',
                'batch_item': {'pk': updated_batch_item.pk}
            })
        else:
            errors = {field: error[0] for field, error in form.errors.items()}
            return JsonResponse({'success': False, 'message': 'Please correct the form errors.', 'errors': errors}, status=400)

    except Exception as e:
        logger.error(f"Error in edit_inventory_batch view for pk={batch_pk}: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'An unexpected server error occurred. Please check the logs.'}, status=500)


@login_required
def stock_take_operator_view(request):
    user = request.user
    if not user.warehouse:
        messages.error(request, "You are not assigned to a warehouse. Stock take cannot be performed.")
        return redirect('inventory:inventory_batch_list_view')

    active_stock_take_session = None
    session_id_from_url = request.GET.get('session_id')

    if session_id_from_url:
        try:
            active_stock_take_session = StockTakeSession.objects.get(
                pk=session_id_from_url,
                warehouse=user.warehouse,
                status__in=['PENDING', 'COMPLETED_BY_OPERATOR']
            )
        except StockTakeSession.DoesNotExist:
            messages.error(request, "Invalid or inaccessible stock take session selected.")
            return redirect('inventory:stock_take_operator')

    if request.method == 'POST':
        if 'select_or_create_session' in request.POST:
            selection_form = StockTakeSessionSelectionForm(request.POST, warehouse=user.warehouse)
            if selection_form.is_valid():
                selected_session = selection_form.cleaned_data.get('active_session')
                new_session_name_input = selection_form.cleaned_data.get('session_name')

                if selected_session:
                    active_stock_take_session = selected_session
                    messages.success(request, f"Selected stock take session: {active_stock_take_session.name}")
                elif new_session_name_input: # If a name was provided (even if auto-generated and then submitted)
                    active_stock_take_session = StockTakeSession.objects.create(
                        name=new_session_name_input, # Use the submitted name
                        warehouse=user.warehouse,
                        initiated_by=user,
                        status='PENDING'
                    )
                    messages.success(request, f"New stock take session started: {active_stock_take_session.name}")
                else:
                    messages.error(request, "Invalid session selection or creation attempt.")
                    return redirect('inventory:stock_take_operator')
                return redirect(f"{request.path}?session_id={active_stock_take_session.pk}")

        elif 'submit_stock_take_items' in request.POST or 'mark_session_complete' in request.POST:
            session_pk_from_form = request.POST.get('session_pk')
            if not session_pk_from_form:
                messages.error(request, "Session identifier missing from submission.")
                return redirect('inventory:stock_take_operator')
            try:
                current_session = StockTakeSession.objects.get(pk=session_pk_from_form, warehouse=user.warehouse)
                if current_session.status not in ['PENDING', 'COMPLETED_BY_OPERATOR']:
                     messages.error(request, "This stock take session can no longer be modified.")
                     return redirect(f"{request.path}?session_id={current_session.pk}")
                active_stock_take_session = current_session
            except StockTakeSession.DoesNotExist:
                messages.error(request, "Stock take session not found or not accessible.")
                return redirect('inventory:stock_take_operator')

            if current_session.status == 'COMPLETED_BY_OPERATOR':
                # If the session is ALREADY marked as complete by operator
                if 'mark_session_complete' in request.POST:
                    messages.warning(request, f"Session '{current_session.name}' is already marked as complete. No further action taken.")
                else: # Trying to save items to an already completed session
                    messages.error(request, f"Session '{current_session.name}' is already complete. No further items can be saved.")
                return redirect(f"{request.path}?session_id={current_session.pk}")

            elif current_session.status != 'PENDING':
                # For any other status that isn't PENDING or COMPLETED_BY_OPERATOR (e.g., EVALUATED, CLOSED)
                messages.error(request, f"Session '{current_session.name}' status ({current_session.get_status_display()}) does not allow operator modifications.")
                return redirect(f"{request.path}?session_id={current_session.pk}")

            formset = StockTakeItemFormSet(
                request.POST,
                instance=current_session,
                prefix='stocktakeitems',
                form_kwargs={'user': user, 'warehouse': user.warehouse}
            )
            if formset.is_valid():
                formset.save()
                messages.success(request, f"Stock take items saved for session: {current_session.name}")

                if 'mark_session_complete' in request.POST:
                    current_session.status = 'COMPLETED_BY_OPERATOR'
                    current_session.completed_by_operator_at = timezone.now()
                    current_session.save()
                    messages.success(request, f"Stock take session '{current_session.name}' marked as complete by operator.")

                    return redirect('inventory:inventory_batch_list_view')
                return redirect(f"{request.path}?session_id={current_session.pk}")
            else:
                messages.error(request, "Please correct the errors in the stock take items.")
        else:
            messages.error(request, "Invalid action or stock take session not active.")
            return redirect('inventory:stock_take_operator')

    item_formset = None
    selection_form = None

    # Data for auto-generating session name
    current_date_yyyymmdd = timezone.now().strftime('%Y%m%d')
    warehouse_name_slug = slugify(user.warehouse.name) if user.warehouse else "unknown_wh"
    # Prefer user.name if available, otherwise fallback to email prefix
    username_slug = slugify(user.name.split('@')[0] if user.name else user.email.split('@')[0]) if user.is_authenticated else "unknown_user"
    auto_generated_session_name_prefix = f"{current_date_yyyymmdd}_{warehouse_name_slug}_{username_slug}"


    if active_stock_take_session:
        item_formset = StockTakeItemFormSet(
            instance=active_stock_take_session,
            prefix='stocktakeitems',
            form_kwargs={'user': user, 'warehouse': user.warehouse}
        )
    else:
        selection_form = StockTakeSessionSelectionForm(warehouse=user.warehouse)
        # We can pre-fill the session_name field if it's a new session context
        # However, the JS will handle this dynamically.
        # If you wanted to pre-fill via Django form initial:
        # selection_form = StockTakeSessionSelectionForm(
        #     warehouse=user.warehouse,
        #     initial={'session_name': auto_generated_session_name_prefix + "_COUNT1"} # Example
        # )

    is_readonly_for_template_flag = False
    if active_stock_take_session and active_stock_take_session.status != 'PENDING':
        is_readonly_for_template_flag = True

    context = {
        'selection_form': selection_form,
        'item_formset': item_formset,
        'active_session': active_stock_take_session,
        'warehouse': user.warehouse,
        'page_title': f"Stock Take for {user.warehouse.name}" if user.warehouse else "Stock Take",
        # Pass data for JS auto-generation
        'current_date_yyyymmdd': current_date_yyyymmdd,
        'warehouse_name_slug': warehouse_name_slug,
        'username_slug': username_slug,
        'auto_generated_session_name_prefix': auto_generated_session_name_prefix, # Pass the full prefix
        'is_session_readonly_for_template': is_readonly_for_template_flag,
    }
    return render(request, 'inventory/stock_take_form.html', context)

@staff_member_required # Ensures only staff (including superusers) can access
def stock_take_session_list_view(request):
    """
    View for superusers to list all StockTakeSession objects.
    Allows filtering by warehouse and status.
    """
    if not request.user.is_superuser: # Extra check if only superusers, not just staff
        messages.error(request, "You do not have permission to view this page.")
        return redirect('inventory:inventory_batch_list_view')

    sessions_qs = StockTakeSession.objects.select_related('warehouse', 'initiated_by', 'evaluated_by').all()

    # Filtering
    warehouse_filter = request.GET.get('warehouse')
    status_filter = request.GET.get('status')

    if warehouse_filter:
        sessions_qs = sessions_qs.filter(warehouse_id=warehouse_filter)
    if status_filter:
        sessions_qs = sessions_qs.filter(status=status_filter)

    warehouses = Warehouse.objects.all().order_by('name')
    status_choices = StockTakeSession.STATUS_CHOICES

    context = {
        'sessions': sessions_qs,
        'warehouses': warehouses,
        'status_choices': status_choices,
        'selected_warehouse': warehouse_filter,
        'selected_status': status_filter,
        'page_title': "Stock Take Sessions"
    }
    return render(request, 'inventory/stock_take_session_list.html', context)


@login_required
def upload_stock_take_csv(request):
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_upload')
        if not csv_file:
            messages.error(request, "No file was uploaded.")
            return redirect('inventory:stock_take_session_list')

        session, created_count, errors = create_stock_take_session_from_csv(csv_file, request.user)

        # --- THIS IS THE FIX ---
        # First, check if the session object was created at all.
        if session is None:
            # If the session is None, a fatal error occurred.
            messages.error(request, "Failed to create stock take session. Please check the file and try again.")
            for error in errors:
                messages.error(request, error)
            return redirect('inventory:stock_take_session_list')
        # -----------------------

        # If the session was created, proceed with the original logic.
        if errors:
            messages.warning(request, f"Stock take session #{session.id} created with {len(errors)} issues.")
            for error in errors[:5]: # Show the first 5 errors
                messages.error(request, error)
        else:
            messages.success(request, f"Successfully created stock take session #{session.id} with {created_count} items.")

        # Automatically run the evaluation after creating the session
        if hasattr(session, 'evaluate_stock_take') and callable(session.evaluate_stock_take):
            session.evaluate_stock_take()
            messages.info(request, f"Stock take session #{session.id} has been evaluated.")
        else:
            messages.warning(request, "Evaluation function not found on session object.")

        return redirect('inventory:stock_take_session_list')

    return redirect('inventory:stock_take_session_list')


@staff_member_required
def stock_take_session_detail_view(request, session_pk):
    """
    View for superusers to see the details and items of a specific StockTakeSession.
    """
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to view this page.")
        return redirect('inventory:stock_take_session_list')

    session = get_object_or_404(
        StockTakeSession.objects.select_related('warehouse', 'initiated_by', 'evaluated_by')
                           .prefetch_related(
                               'items',
                               'items__warehouse_product__product', # For display
                               'items__warehouse_product__warehouse' # For display consistency
                           ),
        pk=session_pk
    )

    # If you want to allow superusers to edit items from this view, you'd use the formset.
    # For now, this view is read-only for items.
    # item_formset = StockTakeItemFormSet(instance=session, prefix='stocktakeitems_detail', form_kwargs={'user': request.user, 'warehouse': session.warehouse})


    context = {
        'session': session,
        # 'item_formset': item_formset, # Uncomment if making items editable here
        'page_title': f"Details for Stock Take: {session.name}"
    }
    return render(request, 'inventory/stock_take_session_detail.html', context)


@staff_member_required
def download_stock_take_session_csv(request, session_pk):
    """
    Allows superusers to download the items of a stock take session as a CSV file.
    """
    if not request.user.is_superuser:
        return HttpResponseForbidden("You do not have permission to perform this action.")

    session = get_object_or_404(StockTakeSession, pk=session_pk)
    items = session.items.select_related(
        'warehouse_product__product',
        'warehouse_product__warehouse'
    ).order_by('warehouse_product__product__sku', 'location_label_counted', 'batch_number_counted')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="stock_take_session_{session.pk}_{session.name.replace(" ", "_")}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Session ID', 'Session Name', 'Session Warehouse', 'Session Status',
        'Item ID (System)', 'Product SKU (System)', 'Product Name (System)', 'Warehouse (System)',
        'Location (Counted)', 'Batch No. (Counted)', 'Expiry (Counted)', 'Quantity (Counted)',
        'Item Notes', 'Counted At'
    ])

    for item in items:
        writer.writerow([
            session.pk,
            session.name,
            session.warehouse.name,
            session.get_status_display(),
            item.warehouse_product.pk,
            item.warehouse_product.product.sku,
            item.warehouse_product.product.name,
            item.warehouse_product.warehouse.name, # Should match session.warehouse.name
            item.location_label_counted if item.location_label_counted else '',
            item.batch_number_counted if item.batch_number_counted else '',
            item.expiry_date_counted.strftime('%Y-%m-%d') if item.expiry_date_counted else '',
            item.counted_quantity,
            item.notes,
            item.counted_at.strftime('%Y-%m-%d %H:%M:%S') if item.counted_at else ''
        ])
    return response

@staff_member_required
@require_POST # Evaluation should be a POST request as it changes data
@transaction.atomic # Ensure all discrepancy creations are atomic
def evaluate_stock_take_session_view(request, session_pk):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to perform this action.")
        return redirect('inventory:stock_take_session_list')

    session = get_object_or_404(StockTakeSession, pk=session_pk)

    if session.status not in ['COMPLETED_BY_OPERATOR', 'EVALUATED']: # Allow re-evaluation
        messages.error(request, f"Session '{session.name}' is not ready for evaluation or has already been closed.")
        return redirect('inventory:stock_take_session_detail', session_pk=session.pk)

    # Clear existing discrepancies for this session to make evaluation idempotent
    StockDiscrepancy.objects.filter(session=session).delete()

    counted_items = StockTakeItem.objects.filter(session=session).select_related('warehouse_product')
    system_items = InventoryBatchItem.objects.filter(
        warehouse_product__warehouse=session.warehouse
    ).select_related('warehouse_product', 'warehouse_product__product')

    # --- Normalize and Aggregate Counted Items ---
    # Key: (warehouse_product_id, location_label_counted, batch_number_counted, expiry_date_counted)
    aggregated_counted = {}
    for item in counted_items:
        key = (
            item.warehouse_product_id,
            item.location_label_counted if item.location_label_counted else "", # Normalize None/empty
            item.batch_number_counted if item.batch_number_counted else "",   # Normalize None/empty
            item.expiry_date_counted
        )
        if key not in aggregated_counted:
            aggregated_counted[key] = {
                'quantity': 0,
                'warehouse_product': item.warehouse_product,
                'location': item.location_label_counted,
                'batch': item.batch_number_counted,
                'expiry': item.expiry_date_counted,
                'notes': [], # Collect notes from multiple entries if any
                'original_sti_pks': [] # Keep track of original StockTakeItem pks
            }
        aggregated_counted[key]['quantity'] += item.counted_quantity
        if item.notes:
            aggregated_counted[key]['notes'].append(item.notes)
        aggregated_counted[key]['original_sti_pks'].append(item.pk)


    # --- Normalize and Aggregate System Items (InventoryBatchItem) ---
    # Key: (warehouse_product_id, location_label, batch_number, expiry_date)
    aggregated_system = {}
    for item in system_items:
        key = (
            item.warehouse_product_id,
            item.location_label if item.location_label else "",
            item.batch_number if item.batch_number else "",
            item.expiry_date
        )
        if key not in aggregated_system: # Should be unique by definition of InventoryBatchItem unique_together
            aggregated_system[key] = {
                'quantity': item.quantity,
                'warehouse_product': item.warehouse_product,
                'location': item.location_label,
                'batch': item.batch_number,
                'expiry': item.expiry_date,
                'original_ibi_pk': item.pk
            }
        else:
            # This case should ideally not happen if your InventoryBatchItem has unique constraints
            # on (wp, location, batch, expiry). If it can, sum quantities.
            aggregated_system[key]['quantity'] += item.quantity


    discrepancies_created = 0
    processed_system_keys = set()

    # 1. Iterate through aggregated counted items
    for key, counted_data in aggregated_counted.items():
        system_data = aggregated_system.get(key)
        discrepancy_type = ''
        discrepancy_qty = 0
        system_qty_for_calc = 0
        system_ibi_pk = None

        if system_data: # Match found based on key
            processed_system_keys.add(key) # Mark this system key as processed
            system_qty_for_calc = system_data['quantity']
            system_ibi_pk = system_data.get('original_ibi_pk')

            discrepancy_qty = counted_data['quantity'] - system_data['quantity']
            if discrepancy_qty == 0:
                discrepancy_type = 'MATCH'
            elif discrepancy_qty > 0:
                discrepancy_type = 'OVER'
            else: # discrepancy_qty < 0
                discrepancy_type = 'SHORT'
        else: # No exact match in system for this counted item configuration
            discrepancy_type = 'NOT_IN_SYSTEM'
            discrepancy_qty = counted_data['quantity'] # The entire counted amount is "extra"
            system_qty_for_calc = 0 # System has 0 of this specific configuration

        # Create StockDiscrepancy record
        # For NOT_IN_SYSTEM, system_inventory_batch_item will be None
        # For MATCH/OVER/SHORT, link to the specific InventoryBatchItem if possible
        StockDiscrepancy.objects.create(
            session=session,
            warehouse_product=counted_data['warehouse_product'],
            system_inventory_batch_item_id=system_ibi_pk if system_data else None,
            system_location_label=system_data['location'] if system_data else None,
            system_batch_number=system_data['batch'] if system_data else None,
            system_expiry_date=system_data['expiry'] if system_data else None,
            system_quantity=system_qty_for_calc,
            # For stock_take_item_reference, if multiple STIs aggregated, maybe leave null or link first one
            # For simplicity, let's leave it null if aggregated from multiple, or link if single.
            stock_take_item_reference_id=counted_data['original_sti_pks'][0] if len(counted_data['original_sti_pks']) == 1 else None,
            counted_location_label=counted_data['location'],
            counted_batch_number=counted_data['batch'],
            counted_expiry_date=counted_data['expiry'],
            counted_quantity=counted_data['quantity'],
            discrepancy_quantity=discrepancy_qty,
            discrepancy_type=discrepancy_type,
            notes="Counted notes: " + "; ".join(counted_data['notes']) if counted_data['notes'] else ""
        )
        discrepancies_created += 1

    # 2. Iterate through system items not processed yet (i.e., not found in counted items)
    for key, system_data in aggregated_system.items():
        if key not in processed_system_keys:
            StockDiscrepancy.objects.create(
                session=session,
                warehouse_product=system_data['warehouse_product'],
                system_inventory_batch_item_id=system_data.get('original_ibi_pk'),
                system_location_label=system_data['location'],
                system_batch_number=system_data['batch'],
                system_expiry_date=system_data['expiry'],
                system_quantity=system_data['quantity'],
                counted_quantity=0, # Not counted
                discrepancy_quantity=-system_data['quantity'], # Negative, as it's a shortage vs count
                discrepancy_type='NOT_COUNTED',
                notes="Item found in system but not in stock take count."
            )
            discrepancies_created += 1

    # Update session status
    session.status = 'EVALUATED'
    session.evaluated_by = request.user
    session.evaluated_at = timezone.now()
    session.save()

    messages.success(request, f"Stock take session '{session.name}' evaluated. {discrepancies_created} discrepancy/match records created.")
    return redirect('inventory:stock_take_session_detail', session_pk=session.pk)


@login_required
def search_warehouse_products_for_stocktake_json(request):
    """
    AJAX endpoint to search for WarehouseProduct items by code, SKU, or name
    within a specific warehouse. Used for stock take product selection.
    """
    term = request.GET.get('term', '').strip()
    warehouse_id = request.GET.get('warehouse_id') # Expecting warehouse_id of the current session

    if not warehouse_id:
        return JsonResponse({'error': 'Warehouse ID is required.'}, status=400)

    try:
        warehouse = Warehouse.objects.get(pk=warehouse_id)
    except Warehouse.DoesNotExist:
        return JsonResponse({'error': 'Invalid Warehouse ID.'}, status=400)

    # Ensure the requesting user has access to this warehouse if not a superuser
    if not request.user.is_superuser and request.user.warehouse != warehouse:
        return JsonResponse({'error': 'Access to this warehouse is denied.'}, status=403)

    results = []
    if term:
        # Search by WarehouseProduct.code, Product.sku, or Product.name
        # Prioritize code matches
        query = Q(code__icontains=term, warehouse=warehouse) | \
                Q(product__sku__icontains=term, warehouse=warehouse) | \
                Q(product__name__icontains=term, warehouse=warehouse)

        warehouse_products = NewAggregateWarehouseProduct.objects.filter(query).select_related('product', 'warehouse')[:10] # Limit results

        for wp in warehouse_products:
            label = f"{wp.code if wp.code else 'N/A'} - {wp.product.sku} - {wp.product.name} (@{wp.warehouse.name})"
            results.append({
                'id': wp.id,                     # WarehouseProduct ID
                'value': wp.code if wp.code else wp.product.sku, # Value for the input after selection (can be code or SKU)
                'label': label,                  # Text displayed in suggestions
                'name': wp.product.name,
                'sku': wp.product.sku,
                'wp_code': wp.code
            })

    return JsonResponse(results, safe=False)

@staff_member_required
def download_stock_take_evaluation_excel(request, session_pk):
    """
    Generates and downloads an Excel report for a stock take session's evaluation,
    focusing on discrepancies.
    """
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to perform this action.")
        return redirect('inventory:stock_take_session_list') # Or appropriate redirect

    session = get_object_or_404(
        StockTakeSession.objects.select_related(
            'warehouse', 'initiated_by', 'evaluated_by'
        ).prefetch_related(
            'discrepancies__warehouse_product__product',
            'discrepancies__warehouse_product__warehouse',
            'discrepancies__system_inventory_batch_item', # For more system details if needed
            'discrepancies__stock_take_item_reference' # For more counted details if needed
        ),
        pk=session_pk
    )

    if session.status != 'EVALUATED' and not session.discrepancies.exists():
        messages.warning(request, f"Session '{session.name}' has not been evaluated yet or has no discrepancies to report.")
        return redirect('inventory:stock_take_session_detail', session_pk=session.pk)

    # Create an Excel workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Evaluation Report - {session.id}"

    # Define Headers
    headers = [
        "Product SKU", "Product Name", "Warehouse",
        "Discrepancy Type",
        "System Batch", "System Location", "System Expiry", "System Qty",
        "Counted Batch", "Counted Location", "Counted Expiry", "Counted Qty",
        "Discrepancy Qty",
        "Notes", "Resolved", "Resolution Notes", "Resolved By", "Resolved At"
    ]
    sheet.append(headers)

    # Apply styles to header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(col_num)].width = 20 # Adjust width as needed

    # Populate data rows
    row_num = 2
    for discrepancy in session.discrepancies.all():
        resolved_by_name = ""
        if discrepancy.resolved_by:
            resolved_by_name = discrepancy.resolved_by.name or discrepancy.resolved_by.email

        data_row = [
            discrepancy.warehouse_product.product.sku,
            discrepancy.warehouse_product.product.name,
            discrepancy.warehouse_product.warehouse.name,
            discrepancy.get_discrepancy_type_display(),
            discrepancy.system_batch_number or "-",
            discrepancy.system_location_label or "-",
            discrepancy.system_expiry_date.strftime('%Y-%m-%d') if discrepancy.system_expiry_date else "-",
            discrepancy.system_quantity if discrepancy.system_quantity is not None else "N/A",
            discrepancy.counted_batch_number or "-",
            discrepancy.counted_location_label or "-",
            discrepancy.counted_expiry_date.strftime('%Y-%m-%d') if discrepancy.counted_expiry_date else "-",
            discrepancy.counted_quantity if discrepancy.counted_quantity is not None else "N/A",
            discrepancy.discrepancy_quantity,
            discrepancy.notes or "-",
            "Yes" if discrepancy.is_resolved else "No",
            discrepancy.resolution_notes or "-",
            resolved_by_name or "-",
            discrepancy.resolved_at.strftime('%Y-%m-%d %H:%M') if discrepancy.resolved_at else "-",
        ]
        sheet.append(data_row)

        # Apply alternating row colors for readability (optional)
        fill_color = "DDEBF7" if row_num % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col_num in range(1, len(headers) + 1):
            sheet.cell(row=row_num, column=col_num).fill = row_fill
        row_num += 1

    # Auto-size columns (can be slow for very large datasets)
    # for column_cells in sheet.columns:
    #     length = max(len(str(cell.value) or "") for cell in column_cells)
    #     sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # Prepare the HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    safe_session_name = "".join(c if c.isalnum() else "_" for c in session.name)
    response['Content-Disposition'] = f'attachment; filename="stock_take_evaluation_{session.pk}_{safe_session_name}.xlsx"'
    workbook.save(response)

    return response


@staff_member_required
@transaction.atomic # Keep this decorator for the whole view
def upload_erp_stock_check_view(request):
    page_title = "Upload ERP Stock Check File"
    if request.method == 'POST':
        form = ErpStockCheckUploadForm(request.POST, request.FILES)
        if form.is_valid():
            erp_file = request.FILES['erp_file']
            session_name = form.cleaned_data['session_name']

            # Create the session first, outside the main processing loop's specific error handling
            # This initial save is fine.
            session = ErpStockCheckSession.objects.create(
                name=session_name,
                uploaded_by=request.user,
                source_file_name=erp_file.name,
                status='PROCESSING' # Initial status
            )

            items_created_count = 0
            items_failed_match_count = 0
            processing_log = [] # For notes

            try: # Outer try for file reading and initial setup
                file_name_lower = erp_file.name.lower()
                # ... (file type detection and workbook/sheet loading logic - same as before) ...
                if file_name_lower.endswith('.xlsx'):
                    workbook_data = openpyxl.load_workbook(erp_file, data_only=True)
                    is_xlsx = True
                elif file_name_lower.endswith('.xls'):
                    file_contents = erp_file.read() # xlrd reads from content
                    workbook_data = xlrd.open_workbook(file_contents=file_contents)
                    is_xlsx = False
                else:
                    raise ValueError("Unsupported file format. Please upload .xlsx or .xls files.")

                sheet_name_to_check = "Quantity grouped by product var"
                sheet = None
                actual_headers = []
                data_rows_iterator = None

                if is_xlsx: # openpyxl
                    if sheet_name_to_check not in workbook_data.sheetnames:
                        raise ValueError(f"Required sheet '{sheet_name_to_check}' not found. Available: {', '.join(workbook_data.sheetnames)}")
                    sheet = workbook_data[sheet_name_to_check]
                    actual_headers = [cell.value for cell in sheet[1]]
                    data_rows_iterator = sheet.iter_rows(min_row=2, values_only=True)
                else: # xlrd
                    try:
                        sheet = workbook_data.sheet_by_name(sheet_name_to_check)
                    except xlrd.XLRDError:
                        raise ValueError(f"Required sheet '{sheet_name_to_check}' not found. Available: {', '.join(workbook_data.sheet_names())}")
                    if sheet.nrows > 0:
                        actual_headers = [sheet.cell_value(0, col_idx) for col_idx in range(sheet.ncols)]
                    else: actual_headers = []
                    def xlrd_rows_iterator(sheet_obj): # Define iterator for xlrd
                        for r_idx in range(1, sheet_obj.nrows): yield [sheet_obj.cell_value(r_idx, c_idx) for c_idx in range(sheet_obj.ncols)]
                    data_rows_iterator = xlrd_rows_iterator(sheet)

                # ... (header validation logic - same as before) ...
                expected_header_map = {
                    "Warehouse Name": "warehouse_name", "Product Name": "product_name_erp",
                    "Variant Product ID": "product_sku", "Sales Total On Hand": "erp_quantity"
                }
                normalized_actual_headers = [str(h).strip().lower() if h is not None else "" for h in actual_headers]
                normalized_expected_keys = {str(k).strip().lower(): v for k,v in expected_header_map.items()}
                column_indices = {}
                missing_headers_user_friendly = []
                for expected_header_key_normalized, _ in normalized_expected_keys.items():
                    try: column_indices[expected_header_key_normalized] = normalized_actual_headers.index(expected_header_key_normalized)
                    except ValueError:
                        original_expected_header = next((k_orig for k_orig in expected_header_map if str(k_orig).strip().lower() == expected_header_key_normalized), "")
                        missing_headers_user_friendly.append(original_expected_header)
                if missing_headers_user_friendly:
                    raise ValueError(f"Missing columns: {', '.join(missing_headers_user_friendly)}. Found: {', '.join(filter(None, actual_headers))}")


                # Loop through rows
                for row_idx, row_values in enumerate(data_rows_iterator, start=2):
                    try: # Inner try for each row processing
                        # ... (data extraction from row_values using column_indices - same as before) ...
                        erp_warehouse_name = str(row_values[column_indices['warehouse name']]).strip() if 'warehouse name' in column_indices and column_indices['warehouse name'] < len(row_values) else ""
                        erp_product_sku = str(row_values[column_indices['variant product id']]).strip() if 'variant product id' in column_indices and column_indices['variant product id'] < len(row_values) else ""
                        erp_quantity_val = row_values[column_indices['sales total on hand']] if 'sales total on hand' in column_indices and column_indices['sales total on hand'] < len(row_values) else None
                        erp_product_name_ref = str(row_values[column_indices['product name']]).strip() if 'product name' in column_indices and column_indices['product name'] < len(row_values) else ""

                        if not erp_warehouse_name or not erp_product_sku:
                            processing_log.append(f"Row {row_idx}: Skipped. Missing Warehouse Name or SKU. WH: '{erp_warehouse_name}', SKU: '{erp_product_sku}'")
                            items_failed_match_count += 1
                            continue
                        try: erp_quantity = int(float(erp_quantity_val)) if erp_quantity_val is not None else 0
                        except (ValueError, TypeError):
                            processing_log.append(f"Row {row_idx}: Skipped. Invalid ERP Qty '{erp_quantity_val}' for SKU '{erp_product_sku}'.")
                            items_failed_match_count += 1
                            continue

                        # Database queries for matching
                        warehouse_obj = Warehouse.objects.get(name__iexact=erp_warehouse_name)
                        product_obj = Product.objects.get(sku__iexact=erp_product_sku)
                        warehouse_product_obj = NewAggregateWarehouseProduct.objects.get(
                            warehouse=warehouse_obj, product=product_obj
                        )

                        # Create ErpStockCheckItem
                        ErpStockCheckItem.objects.create(
                            session=session, # session object is from outside this loop
                            warehouse_product=warehouse_product_obj,
                            erp_warehouse_name_raw=erp_warehouse_name,
                            erp_product_sku_raw=erp_product_sku,
                            erp_product_name_raw=erp_product_name_ref,
                            erp_quantity=erp_quantity,
                            is_matched=True
                        )
                        items_created_count += 1

                    # Catch expected data errors for a single row
                    except Warehouse.DoesNotExist:
                        processing_log.append(f"Row {row_idx}: Warehouse '{erp_warehouse_name}' not found for SKU '{erp_product_sku}'.")
                        items_failed_match_count += 1
                    except Product.DoesNotExist:
                        processing_log.append(f"Row {row_idx}: Product SKU '{erp_product_sku}' not found.")
                        items_failed_match_count += 1
                    except NewAggregateWarehouseProduct.DoesNotExist:
                        processing_log.append(f"Row {row_idx}: Link between Warehouse '{erp_warehouse_name}' and SKU '{erp_product_sku}' not found.")
                        items_failed_match_count += 1
                    except IntegrityError: # For duplicate ErpStockCheckItem in the same session
                        processing_log.append(f"Row {row_idx}: Duplicate entry for SKU '{erp_product_sku}' in Warehouse '{erp_warehouse_name}' for this session.")
                        items_failed_match_count += 1
                    except IndexError: # If a row has fewer columns than expected
                        processing_log.append(f"Row {row_idx}: Skipped. Row column count mismatch.")
                        items_failed_match_count += 1
                    except Exception as e_row: # Catch any other error for this specific row
                        processing_log.append(f"Row {row_idx}: Unexpected error for SKU '{erp_product_sku}': {str(e_row)}")
                        items_failed_match_count += 1
                # End of row processing loop

                # After the loop, update the session based on outcomes
                if items_created_count > 0:
                    session.status = 'PENDING_EVALUATION'
                    messages.success(request, f"Successfully processed {items_created_count} items. {items_failed_match_count} rows had issues or were skipped.")
                elif items_failed_match_count > 0 and items_created_count == 0 : # All rows failed or were skipped
                    session.status = 'UPLOAD_FAILED' # Or a more specific status
                    messages.warning(request, f"No items could be processed. {items_failed_match_count} rows had issues or were skipped. Check processing notes.")
                else: # No rows in file or no valid data rows
                    session.status = 'UPLOAD_FAILED' # Or a more specific status
                    messages.warning(request, "The file contained no processable data or was empty after the header.")

            except ValueError as ve_file: # Catches file format, sheet name, header errors
                session.status = 'UPLOAD_FAILED'
                session.processing_notes = str(ve_file)
                messages.error(request, session.processing_notes)
            except Exception as e_file: # Catches other file-level errors (e.g., xlrd read issues)
                session.status = 'UPLOAD_FAILED'
                session.processing_notes = f"Failed to read or process Excel file: {str(e_file)}"
                messages.error(request, session.processing_notes)
            finally:
                # This save will happen regardless of row-level errors,
                # updating status and notes. It's crucial this is the only
                # session.save() after the loop if an error occurred inside the loop.
                session.processing_notes = (session.processing_notes + "\n---\n" if session.processing_notes else "") + "\n".join(processing_log)
                session.processed_at = timezone.now()
                session.save() # Save the session status and notes

            # Redirect based on final session status (or always to detail)
            if session.status == 'UPLOAD_FAILED':
                return redirect('inventory:erp_stock_check_list')
            else:
                return redirect('inventory:erp_stock_check_detail', session_pk=session.pk)

        else: # Form not valid
            for field, field_errors in form.errors.items():
                for error in field_errors:
                    messages.error(request, f"Form error in {field}: {error}")
            if not form.is_multipart():
                 messages.error(request, "Form was not multipart. File upload cannot succeed.")
    else: # GET request
        form = ErpStockCheckUploadForm()
        form.fields['session_name'].initial = f"ERP Snapshot - {timezone.now().strftime('%Y-%m-%d_%H%M')}"

    context = {
        'form': form,
        'page_title': page_title,
    }
    return render(request, 'inventory/erp_stock_check_upload.html', context)


# You'll also need views for listing sessions and viewing details/evaluating:
@staff_member_required
def erp_stock_check_list_view(request):
    # Simple list view for now
    sessions = ErpStockCheckSession.objects.all().order_by('-uploaded_at')
    context = {
        'sessions': sessions,
        'page_title': "ERP Stock Check Sessions"
    }
    return render(request, 'inventory/erp_stock_check_list.html', context)

@staff_member_required
def erp_stock_check_detail_view(request, session_pk):
    session = get_object_or_404(ErpStockCheckSession.objects.prefetch_related(
        'items__warehouse_product__product',
        'items__warehouse_product__warehouse',
        'discrepancies__warehouse_product__product' # If discrepancies are already generated
    ), pk=session_pk)

    items = session.items.all()
    discrepancies = session.discrepancies.all()

    context = {
        'session': session,
        'items': items, # ErpStockCheckItem instances
        'discrepancies': discrepancies, # WarehouseProductDiscrepancy instances
        'page_title': f"ERP Check Details: {session.name}"
    }
    return render(request, 'inventory/erp_stock_check_detail.html', context)

@staff_member_required
@require_POST
@transaction.atomic
def evaluate_erp_stock_check_view(request, session_pk):
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to perform this action.")
        return redirect('inventory:erp_stock_check_list')

    session = get_object_or_404(ErpStockCheckSession, pk=session_pk)

    if session.status not in ['PENDING_EVALUATION', 'EVALUATED']:
        messages.error(request, f"Session '{session.name}' is not ready for evaluation or has already been closed.")
        return redirect('inventory:erp_stock_check_detail', session_pk=session.pk)

    WarehouseProductDiscrepancy.objects.filter(session=session).delete()

    # Matched ERP items (items that exist in your system)
    matched_erp_items = ErpStockCheckItem.objects.filter(
        session=session,
        is_matched=True
    ).select_related('warehouse_product__product', 'warehouse_product__warehouse')

    # Determine the scope of WarehouseProducts in your system to check against
    # These are WPs that *could* have been in the ERP file.
    system_wp_ids_in_erp_scope = set()
    if session.warehouse: # If session was for a specific warehouse
        system_wp_ids_in_erp_scope.update(
            NewAggregateWarehouseProduct.objects.filter(warehouse=session.warehouse).values_list('id', flat=True)
        )
    else: # If session was for all warehouses (or inferred from file)
          # Get IDs of all WPs that actually appeared (matched) in the ERP file for this session
        system_wp_ids_in_erp_scope.update(
            matched_erp_items.values_list('warehouse_product_id', flat=True)
        )
        # Additionally, if you want to consider WPs from warehouses that had *any* item in the ERP file
        # (even if some WPs from that warehouse weren't in the file), this could be expanded.
        # For now, this is a reasonable scope for "NOT_IN_ERP".

    processed_system_wp_ids = set() # Keep track of system WPs that had a corresponding ERP entry
    discrepancies_created_count = 0

    # 1. Process items that were matched between ERP and your system
    for erp_item in matched_erp_items:
        wp = erp_item.warehouse_product
        if not wp: # Should not happen if is_matched is True, but as a safeguard
            continue

        processed_system_wp_ids.add(wp.id)

        system_qty = wp.quantity
        erp_qty = erp_item.erp_quantity
        diff = system_qty - erp_qty
        discrepancy_type = ''

        if diff == 0:
            discrepancy_type = 'MATCH'
        elif diff > 0:
            discrepancy_type = 'OVER_IN_SYSTEM'
        else:
            discrepancy_type = 'SHORT_IN_SYSTEM'

        WarehouseProductDiscrepancy.objects.create(
            session=session,
            warehouse_product=wp,
            erp_stock_check_item=erp_item, # Link to the source ERP item
            system_quantity=system_qty,
            erp_quantity=erp_qty,
            # discrepancy_quantity will be auto-calculated by model's save method
            discrepancy_type=discrepancy_type,
        )
        discrepancies_created_count += 1

    # 2. Process items in your system (within scope) that were NOT in the ERP file (NOT_IN_ERP)
    if system_wp_ids_in_erp_scope: # Only if there's a defined scope of system items
        wps_not_in_erp = NewAggregateWarehouseProduct.objects.filter(id__in=system_wp_ids_in_erp_scope).exclude(id__in=processed_system_wp_ids)
        for wp in wps_not_in_erp.select_related('product', 'warehouse'):
            WarehouseProductDiscrepancy.objects.create(
                session=session,
                warehouse_product=wp,
                system_quantity=wp.quantity,
                erp_quantity=0, # Not found in ERP, so ERP quantity is effectively 0
                discrepancy_type='NOT_IN_ERP',
                # discrepancy_quantity will be auto-calculated
                notes=f"SKU {wp.product.sku} @ {wp.warehouse.name} in system, not in this ERP snapshot."
            )
            discrepancies_created_count += 1

    # 3. Process items from ERP file that were NOT matched to any item in your system (NOT_IN_SYSTEM)
    unmatched_erp_items = ErpStockCheckItem.objects.filter(session=session, is_matched=False)
    for unmatched_item in unmatched_erp_items:
        WarehouseProductDiscrepancy.objects.create(
            session=session,
            warehouse_product=None, # Explicitly None as it's not in our system
            erp_stock_check_item=unmatched_item, # Link to the source ERP item
            erp_warehouse_name_for_unmatched=unmatched_item.erp_warehouse_name_raw,
            erp_product_sku_for_unmatched=unmatched_item.erp_product_sku_raw,
            erp_product_name_for_unmatched=unmatched_item.erp_product_name_raw,
            system_quantity=0, # Not in our system, so system quantity is 0
            erp_quantity=unmatched_item.erp_quantity,
            # discrepancy_quantity will be auto-calculated
            discrepancy_type='NOT_IN_SYSTEM',
            notes=f"ERP item SKU {unmatched_item.erp_product_sku_raw} @ WH {unmatched_item.erp_warehouse_name_raw} not matched to system."
        )
        discrepancies_created_count += 1

    session.status = 'EVALUATED'
    session.evaluated_by = request.user
    session.evaluated_at = timezone.now()
    session.save()

    messages.success(request, f"ERP Stock Check session '{session.name}' evaluated. {discrepancies_created_count} discrepancy/match records created/updated.")
    return redirect('inventory:erp_stock_check_detail', session_pk=session.pk)

@staff_member_required
def download_erp_evaluation_excel(request, session_pk):
    """
    Generates and downloads an Excel report for an ERP stock check session's evaluation,
    focusing on WarehouseProductDiscrepancies.
    """
    if not request.user.is_superuser:
        messages.error(request, "You do not have permission to perform this action.")
        return redirect('inventory:erp_stock_check_list')

    session = get_object_or_404(
        ErpStockCheckSession.objects.select_related(
            'warehouse', 'uploaded_by', 'evaluated_by'
        ).prefetch_related(
            'discrepancies__warehouse_product__product', # For matched items
            'discrepancies__warehouse_product__warehouse', # For matched items
            'discrepancies__erp_stock_check_item' # To get raw ERP data if needed
        ),
        pk=session_pk
    )

    if session.status != 'EVALUATED' and not session.discrepancies.exists():
        messages.warning(request, f"ERP Check Session '{session.name}' has not been evaluated or has no discrepancies.")
        return redirect('inventory:erp_stock_check_detail', session_pk=session.pk)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"ERP_Eval_Report_{session.id}"

    headers = [
        "System Product SKU", "System Product Name", "System Warehouse", # From WarehouseProduct if linked
        "ERP Warehouse (Raw)", "ERP Product SKU (Raw)", "ERP Product Name (Raw)", # From ErpStockCheckItem or Discrepancy
        "Discrepancy Type",
        "System Qty", "ERP Qty", "Discrepancy Qty",
        "Notes", "Resolved", "Resolution Notes", "Resolved By", "Resolved At"
    ]
    sheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Blue
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(col_num)].width = 22


    row_num = 2
    for discrepancy in session.discrepancies.all():
        system_sku = "-"
        system_product_name = "-"
        system_warehouse_name = "-"
        if discrepancy.warehouse_product:
            system_sku = discrepancy.warehouse_product.product.sku
            system_product_name = discrepancy.warehouse_product.product.name
            system_warehouse_name = discrepancy.warehouse_product.warehouse.name

        erp_wh_raw = discrepancy.erp_warehouse_name_for_unmatched or (discrepancy.erp_stock_check_item.erp_warehouse_name_raw if discrepancy.erp_stock_check_item else "-")
        erp_sku_raw = discrepancy.erp_product_sku_for_unmatched or (discrepancy.erp_stock_check_item.erp_product_sku_raw if discrepancy.erp_stock_check_item else "-")
        erp_name_raw = discrepancy.erp_product_name_for_unmatched or (discrepancy.erp_stock_check_item.erp_product_name_raw if discrepancy.erp_stock_check_item else "-")

        resolved_by_name = (discrepancy.resolved_by.name or discrepancy.resolved_by.email) if discrepancy.resolved_by else ""

        data_row = [
            system_sku, system_product_name, system_warehouse_name,
            erp_wh_raw, erp_sku_raw, erp_name_raw,
            discrepancy.get_discrepancy_type_display(),
            discrepancy.system_quantity if discrepancy.system_quantity is not None else "N/A",
            discrepancy.erp_quantity if discrepancy.erp_quantity is not None else "N/A",
            discrepancy.discrepancy_quantity,
            discrepancy.notes or "-",
            "Yes" if discrepancy.is_resolved else "No",
            discrepancy.resolution_notes or "-",
            resolved_by_name or "-",
            discrepancy.resolved_at.strftime('%Y-%m-%d %H:%M') if discrepancy.resolved_at else "-",
        ]
        sheet.append(data_row)
        fill_color = "DDEBF7" if row_num % 2 == 0 else "FFFFFF" # Light blue alternating
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col_num in range(1, len(headers) + 1):
            sheet.cell(row=row_num, column=col_num).fill = row_fill
        row_num += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    safe_session_name = "".join(c if c.isalnum() else "_" for c in session.name)
    response['Content-Disposition'] = f'attachment; filename="erp_stock_evaluation_{session.pk}_{safe_session_name}.xlsx"'
    workbook.save(response)
    return response

@login_required
@require_POST # Ensures this action is done via POST for safety
@db_transaction.atomic # Ensures atomicity if multiple updates were needed (though model save handles it now)
def set_default_pick_view(request, batch_pk):
    user = request.user
    try:
        batch_to_set_default = InventoryBatchItem.objects.select_related(
            'warehouse_product__warehouse',
            'warehouse_product__product' # For the success message
        ).get(pk=batch_pk)

        # Permission Check (optional, but good practice)
        if not user.is_superuser and (not user.warehouse or batch_to_set_default.warehouse_product.warehouse != user.warehouse):
            return JsonResponse({'success': False, 'message': 'You do not have permission to modify this batch item.'}, status=403)

        # The model's save() method now handles unsetting other defaults.
        # We just need to set this one to True and save it.
        if not batch_to_set_default.is_default_pick: # Only proceed if it's not already default
            batch_to_set_default.is_default_pick = True
            batch_to_set_default.save() # This will trigger the logic in the model's save()
            message = f'Batch "{batch_to_set_default.batch_number}" is now the default pick for {batch_to_set_default.warehouse_product.product.name}.'
        else:
            # Optionally, allow unsetting if clicked again, or just do nothing if already default
            # batch_to_set_default.is_default_pick = False
            # batch_to_set_default.save()
            # message = f'Batch "{batch_to_set_default.batch_number}" is no longer the default pick.'
            message = f'Batch "{batch_to_set_default.batch_number}" was already the default pick.'


        return JsonResponse({
            'success': True,
            'message': message,
            'batch_pk': batch_to_set_default.pk,
            'warehouse_product_pk': batch_to_set_default.warehouse_product.pk
        })
    except InventoryBatchItem.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Batch item not found.'}, status=404)
    except Exception as e:
        # Log the error (e.g., import logging; logging.error(f"Error in set_default_pick: {e}", exc_info=True))
        return JsonResponse({'success': False, 'message': f'An error occurred: {str(e)}'}, status=500)


@login_required
def get_default_pick_items(request):
    """
    Fetches all inventory items that are currently designated as a
    default pick location (priority 0).
    """
    try:
        # CORRECTED: Added a filter to only get items with a pick_priority of 0
        default_picks = InventoryBatchItem.objects.filter(
            pick_priority=0
        ).select_related(
            'warehouse_product__product',
            'warehouse_product__warehouse'
        ).order_by('warehouse_product__product__name')

        # Serialize the data into a list of dictionaries
        data = [{
            'id': item.id,
            'location_label': item.location_label,
            'product_name': item.warehouse_product.product.name,
            'batch_number': item.batch_number,
            'expiry_date': item.expiry_date,
            'quantity': item.quantity,
            'warehouse_product_id': item.warehouse_product.id,
        } for item in default_picks]

        return JsonResponse({'success': True, 'default_picks': data})

    except Exception as e:
        logger.error(f"Error in get_default_pick_items: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': 'An error occurred on the server.'}, status=500)


@login_required
@require_GET # For search
def search_batch_by_location_json_view(request):
    location_label_query = request.GET.get('location_label', '').strip()
    warehouse_id_query = request.GET.get('warehouse_id') # Optional: for filtering by warehouse
    user = request.user

    if not location_label_query:
        return JsonResponse([], safe=False)

    # Initial queryset construction (without slicing)
    batch_items_qs = InventoryBatchItem.objects.filter(
        location_label__icontains=location_label_query
    ).select_related(
        'warehouse_product__product',
        'warehouse_product__warehouse'
    )

    # Apply warehouse filter if user is not superuser or if warehouse_id is explicitly passed
    if not user.is_superuser and user.warehouse:
        batch_items_qs = batch_items_qs.filter(warehouse_product__warehouse=user.warehouse)
    elif warehouse_id_query: # Allow superuser to filter by warehouse if they provide it
        try:
            batch_items_qs = batch_items_qs.filter(warehouse_product__warehouse_id=int(warehouse_id_query))
        except ValueError:
            pass # Ignore invalid warehouse_id

    # Apply ordering and THEN slicing
    batch_items_qs = batch_items_qs.order_by('location_label', 'warehouse_product__product__name')[:10]

    results = []
    for item in batch_items_qs:
        # Defensive checks for related objects (kept from previous good suggestion)
        if not item.warehouse_product:
            # logger.warning(f"InventoryBatchItem {item.pk} has no warehouse_product. Skipping for search results.")
            continue
        if not item.warehouse_product.product:
            # logger.warning(f"WarehouseProduct {item.warehouse_product.pk} (for BatchItem {item.pk}) has no product. Skipping.")
            continue

        results.append({
            'id': item.pk,
            'location_label': item.location_label,
            'warehouse_product_id': item.warehouse_product.pk,
            'product_name': item.warehouse_product.product.name,
            'product_sku': item.warehouse_product.product.sku,
            'batch_number': item.batch_number,
            'expiry_date': item.expiry_date.isoformat() if item.expiry_date else None,
            'quantity': item.quantity,
            'current_pick_priority': item.pick_priority,
        })
    return JsonResponse(results, safe=False)

@login_required
@require_POST
@db_transaction.atomic
def update_default_pick_items_view(request):
    user = request.user
    user_warehouse = user.warehouse if hasattr(user, 'warehouse') and request.user.warehouse else None

    formset = DefaultPickItemFormSet(request.POST, prefix='default_picks', warehouse=user_warehouse)

    if formset.is_valid():
        items_newly_set_default_count = 0
        items_cleared_from_default_count = 0

        batch_ids_to_clear_priority = []
        batch_ids_to_set_priority_0 = []

        logger.info(f"--- PASS 1: Collecting Batch IDs for Default Pick Update ---")
        logger.info(f"User: {user.email}")

        # Pass 1: Collect batch_ids based on DELETE flag and ensure form validity
        for form_idx, form in enumerate(formset.forms):
            if not form.is_valid():
                logger.warning(f"Form {form_idx} in DefaultPickItemFormSet is not valid. Errors: {form.errors.as_json()}")
                continue

            cleaned_data = form.cleaned_data
            batch_id = cleaned_data.get('inventory_batch_item_id')

            # Enhanced logging for DELETE flag
            # The field name for DELETE is constructed by Django formsets as '<prefix>-<form_index>-DELETE'
            delete_field_key_in_post = form.add_prefix('DELETE') # Gets the correct key like 'default_picks-0-DELETE'
            raw_delete_value_from_post = request.POST.get(delete_field_key_in_post) # Check raw POST data
            cleaned_should_delete_flag = cleaned_data.get('DELETE', False)

            logger.debug(
                f"Form {form_idx}: Batch ID: {batch_id}, "
                f"Raw POST for '{delete_field_key_in_post}': '{raw_delete_value_from_post}', "
                f"Cleaned 'DELETE' flag: {cleaned_should_delete_flag}"
            )

            if batch_id:
                try:
                    if not InventoryBatchItem.objects.filter(pk=batch_id).exists(): # Minimal check first
                        raise InventoryBatchItem.DoesNotExist
                except InventoryBatchItem.DoesNotExist:
                    messages.warning(request, f"Batch item with ID {batch_id} (from form {form_idx}) not found during collection. Skipped.")
                    logger.warning(f"Batch item with ID {batch_id} (from form {form_idx}) not found during collection.")
                    continue

                # Permission check on the actual item
                # This should be done before deciding to add to lists, but simplified here for focus.
                # Ideally, fetch the item once if needed for permission.
                # The more detailed permission check is in Pass 2 and Pass 3.

                if cleaned_should_delete_flag: # Use the cleaned data value
                    batch_ids_to_clear_priority.append(batch_id)
                    logger.info(f"Collected batch ID {batch_id} for CLEARING priority (DELETE flag was True).")
                else:
                    batch_ids_to_set_priority_0.append(batch_id)
                    logger.info(f"Collected batch ID {batch_id} for SETTING priority to 0 (DELETE flag was False).")
            else:
                logger.debug(f"Form {form_idx} skipped: no batch_id (likely an empty extra form or new row not fully populated). Cleaned DELETE: {cleaned_should_delete_flag}")

        logger.info(f"--- END PASS 1 ---")
        logger.info(f"Collected for clearing: {batch_ids_to_clear_priority}")
        logger.info(f"Collected for setting default: {batch_ids_to_set_priority_0}")

        # ... (Pass 2 and Pass 3 logic remains the same as in the user's uploaded views.py / my previous good version) ...
        # Pass 2: Process items to clear default status (set pick_priority to None)
        if batch_ids_to_clear_priority:
            logger.info(f"--- PASS 2: Clearing Default Picks ---")
            logger.info(f"User: {user.email}, Warehouse: {user_warehouse.name if user_warehouse else 'N/A (Superuser or no WH)'}")
            logger.info(f"Batch IDs collected for clearing: {batch_ids_to_clear_priority}")

            cleared_in_this_pass = 0
            for batch_id_to_clear in batch_ids_to_clear_priority:
                try:
                    item_to_clear = InventoryBatchItem.objects.select_related('warehouse_product__warehouse', 'warehouse_product__product').get(pk=batch_id_to_clear)
                    logger.info(f"Processing batch_id_to_clear: {batch_id_to_clear}. Current DB pick_priority: {item_to_clear.pick_priority} for WP: {item_to_clear.warehouse_product}")

                    if not user.is_superuser and user_warehouse:
                        if item_to_clear.warehouse_product.warehouse != user_warehouse:
                            logger.warning(f"Permission DENIED for user {user.email} on batch ID {batch_id_to_clear} (WH: {item_to_clear.warehouse_product.warehouse.name}). Skipping.")
                            messages.error(request, f"Permission denied for batch {item_to_clear.warehouse_product.product.name} - {item_to_clear.batch_number}.")
                            continue

                    original_priority_for_log = item_to_clear.pick_priority
                    if item_to_clear.pick_priority == 0:
                        item_to_clear.pick_priority = None
                        item_to_clear.save()
                        cleared_in_this_pass += 1
                        logger.info(f"SUCCESS: Cleared pick_priority for batch ID {batch_id_to_clear}. Was: {original_priority_for_log}, Now: None.")
                    elif item_to_clear.pick_priority is None:
                        logger.info(f"INFO: Batch ID {batch_id_to_clear} was marked for delete but ALREADY had pick_priority=None in DB.")
                    else:
                        logger.warning(f"WARNING: Batch ID {batch_id_to_clear} marked for delete had unexpected pick_priority={original_priority_for_log}. Forcing to None.")
                        item_to_clear.pick_priority = None
                        item_to_clear.save()
                        if original_priority_for_log is not None:
                             cleared_in_this_pass += 1
                             logger.info(f"SUCCESS (Forced): Cleared pick_priority for batch ID {batch_id_to_clear}. Was: {original_priority_for_log}, Now: None.")
                except InventoryBatchItem.DoesNotExist:
                    logger.warning(f"ERROR: Batch ID {batch_id_to_clear} not found in DB during clearing pass (Pass 2).")
                    messages.warning(request, f"Item to clear (ID: {batch_id_to_clear}) not found during processing.")
                except Exception as e:
                    logger.error(f"EXCEPTION during clearing batch ID {batch_id_to_clear}: {str(e)}", exc_info=True)
                    messages.error(request, f"Error processing removal for item ID {batch_id_to_clear}: {str(e)}")
            items_cleared_from_default_count = cleared_in_this_pass
            logger.info(f"--- END PASS 2: items_cleared_from_default_count = {items_cleared_from_default_count} ---")

        # Pass 3: Process items to set as default (pick_priority = 0)
        if batch_ids_to_set_priority_0:
            logger.info(f"--- Default Picks - Pass 3: Setting Default Picks (Priority 0) ---")
            logger.info(f"Batch IDs collected for setting default: {batch_ids_to_set_priority_0}")
            for batch_id_to_set in batch_ids_to_set_priority_0:
                if batch_id_to_set in batch_ids_to_clear_priority:
                    logger.debug(f"Default Picks - Batch ID {batch_id_to_set} was in clear list, skipping set default.")
                    continue
                try:
                    batch_item = InventoryBatchItem.objects.select_related('warehouse_product__warehouse', 'warehouse_product__product').get(pk=batch_id_to_set)
                    logger.info(f"Default Picks - Processing batch ID {batch_item.pk}. Current DB pick_priority: {batch_item.pick_priority}")

                    if not user.is_superuser and user_warehouse and batch_item.warehouse_product.warehouse != user_warehouse:
                        logger.warning(f"Default Picks - Permission DENIED setting default for batch ID {batch_id_to_set} (WH mismatch).")
                        messages.error(request, f"Permission denied for item {batch_item.warehouse_product.product.name} ({batch_item.batch_number}).")
                        continue

                    if batch_item.pick_priority != 0: # Only update if not already default
                        logger.info(f"Default Picks - Attempting to set pick_priority=0 for batch ID {batch_item.pk}")
                        batch_item.pick_priority = 0
                        batch_item.save()
                        items_newly_set_default_count += 1
                        logger.info(f"Default Picks - SUCCESS: Set pick_priority=0 for batch ID {batch_item.pk} (WP: {batch_item.warehouse_product_id}).")
                    else:
                        logger.info(f"Default Picks - INFO: Batch ID {batch_item.pk} was already pick_priority=0. No change made.")

                except InventoryBatchItem.DoesNotExist:
                    messages.warning(request, f"Batch item with ID {batch_id_to_set} not found for setting default. Skipped.")
                    logger.warning(f"Batch item with ID {batch_id_to_set} not found for setting default during processing.")
                except IntegrityError as e:
                    messages.error(request, f"Integrity error setting default for batch ID {batch_id_to_set}: {str(e)}.")
                    logger.error(f"Integrity error setting default for batch ID {batch_id_to_set}: {str(e)}", exc_info=True)
                except Exception as e:
                    messages.error(request, f"Error setting default for batch ID {batch_id_to_set}: {str(e)}")
                    logger.error(f"Error setting default for batch ID {batch_id_to_set}: {str(e)}", exc_info=True)
            logger.info(f"--- END PASS 3: items_newly_set_default_count = {items_newly_set_default_count} ---")


        # ... (Message generation logic and response/redirect as in your views.py) ...
        if items_newly_set_default_count > 0 or items_cleared_from_default_count > 0:
            success_msg = "Default pick locations updated: "
            if items_newly_set_default_count > 0:
                success_msg += f"{items_newly_set_default_count} item(s) set/confirmed as default. "
            if items_cleared_from_default_count > 0:
                success_msg += f"{items_cleared_from_default_count} item(s) cleared from default."
            messages.success(request, success_msg.strip())
        else:
            submitted_forms_with_batch_ids = [
                form for form in formset.forms
                if form.is_valid() and form.cleaned_data.get('inventory_batch_item_id')
            ]
            if not submitted_forms_with_batch_ids:
                messages.info(request, "No valid items were submitted for default pick location changes.")
            else:
                tried_to_delete = any(
                    form.cleaned_data.get('DELETE')
                    for form in submitted_forms_with_batch_ids
                )
                tried_to_set_or_keep = any(
                    not form.cleaned_data.get('DELETE')
                    for form in submitted_forms_with_batch_ids
                )
                if tried_to_delete and not items_cleared_from_default_count:
                    messages.warning(request, "Items marked for removal from default picks were already not in a default state (or encountered an issue). Please check logs if unexpected.")
                elif tried_to_set_or_keep and not items_newly_set_default_count:
                    messages.info(request, "Items intended as default were already set as such, or no new items were designated default.")
                else:
                    messages.info(request, "No effective changes were made to default pick locations (items may have already been in the desired state or no valid actions performed).")

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            if items_newly_set_default_count == 0 and items_cleared_from_default_count == 0:
                final_json_message = "No effective changes made."
                # ... (your more detailed final_json_message logic from previous step) ...
                return JsonResponse({'success': True, 'message': final_json_message, 'details': 'No database updates performed.'})
            return JsonResponse({'success': True, 'message': 'Default picks updated successfully.'})
        return redirect('inventory:inventory_batch_list_view')
    else:
        # ... (Formset not valid - existing refined error handling) ...
        logger.error(f"Formset for default picks is not valid. Raw errors: {formset.errors}")
        logger.error(f"Formset non-form errors: {formset.non_form_errors()}")
        json_errors = []
        if hasattr(formset, 'errors') and isinstance(formset.errors, list):
            for i, error_dict in enumerate(formset.errors):
                if error_dict:
                    simple_error_dict = {field: [str(e) for e in elist] for field, elist in error_dict.items()}
                    json_errors.append({f"form-{i}": simple_error_dict})
        json_non_form_errors = []
        if hasattr(formset, 'non_form_errors') and callable(formset.non_form_errors):
            nfe = formset.non_form_errors()
            if nfe:
                json_non_form_errors = list(nfe)
        error_messages_for_display = []
        for i, error_dict in enumerate(formset.errors):
            if error_dict:
                for field, errors_in_field in error_dict.items():
                    error_messages_for_display.append(f"Row {i+1} ({field}): {', '.join(errors_in_field)}")
        if formset.non_form_errors():
            error_messages_for_display.append(f"General errors: {', '.join(formset.non_form_errors())}")
        error_message_str = "Please correct the errors: " + "; ".join(error_messages_for_display) if error_messages_for_display else "Invalid data submitted."
        messages.error(request, error_message_str)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': error_message_str,
                'errors': json_errors,
                'formset_errors': json_non_form_errors
            }, status=400)
        return redirect('inventory:inventory_batch_list_view')

